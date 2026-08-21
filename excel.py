"""Leer el Excel que ya usan en planta, tal como viene.

No hay plantilla. Nadie va a recopiar 43 filas a otro formato: el programa se
adapta al archivo, no al revés.

Tres pasos, y el del medio es el importante:

  1. `leer` abre el archivo y encuentra sola la fila de títulos.
  2. `detectar` propone qué columna es cada cosa. **Es una propuesta**: la
     pantalla la muestra y se puede corregir antes de guardar.
  3. `armar` convierte las filas en datos, y lo que no entiende lo devuelve
     aparte con el motivo. Nunca adivina: una fila dudosa se descarta y se
     muestra, no se carga a medias.
"""
from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime

from openpyxl import Workbook, load_workbook

# Cuántas filas miramos buscando los títulos antes de rendirnos.
_FILAS_CABECERA = 15


def _limpio(texto) -> str:
    """Minúsculas, sin tildes, sin puntuación. Para comparar títulos."""
    s = str(texto or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9 ]+", " ", s).strip()


def hojas(ruta: str) -> list[str]:
    wb = load_workbook(ruta, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def desde_texto(texto: str, ruta: str) -> None:
    """Guarda como .xlsx lo que se copió de una planilla y se pegó en pantalla.

    Existe porque no siempre se puede adjuntar el archivo: la planilla está en
    otra computadora, o el navegador no deja elegirlo. Copiar las celdas y
    pegarlas siempre se puede.

    Se convierte a un .xlsx de verdad a propósito, para que de acá en adelante
    haya UN solo camino: el mismo lector, la misma pantalla de revisión y la
    misma confirmación que un archivo subido. Un segundo camino sería un
    segundo lugar donde equivocarse.
    """
    lineas = [l for l in texto.replace("\r\n", "\n").replace("\r", "\n").split("\n")
              if l.strip()]
    if len(lineas) < 2:
        raise ValueError("Pegá la fila de títulos y al menos una fila de datos.")

    # El separador lo decide la fila de títulos, y vale para todas. Si cada
    # fila eligiera el suyo, una fila con una coma adentro partiría distinto
    # que el resto y las columnas quedarían corridas sin que se note.
    separador = next((s for s in ("\t", ";", ",") if s in lineas[0]), None)
    if separador is None:
        raise ValueError(
            "Cada fila tiene que traer al menos dos columnas. "
            "Copiá las celdas desde el Excel, no una lista escrita a mano.")

    wb = Workbook()
    ws = wb.active
    for linea in lineas:
        ws.append([c.strip() for c in linea.split(separador)])
    wb.save(ruta)


def leer(ruta: str, hoja: str | None = None) -> tuple[list[str], list[list]]:
    """Devuelve (títulos, filas). Encuentra la fila de títulos sola.

    Muchos Excel de planta tienen un logo o un título arriba. La fila de
    títulos es la primera con al menos dos celdas con texto.
    """
    wb = load_workbook(ruta, read_only=True, data_only=True)
    try:
        ws = wb[hoja] if hoja and hoja in wb.sheetnames else wb[wb.sheetnames[0]]
        crudas = [list(f) for f in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    inicio = 0
    for i, fila in enumerate(crudas[:_FILAS_CABECERA]):
        textos = [c for c in fila if isinstance(c, str) and c.strip()]
        if len(textos) >= 2:
            inicio = i
            break

    titulos = [str(c).strip() if c is not None else "" for c in crudas[inicio]] if crudas else []
    filas = [f for f in crudas[inicio + 1:] if any(c not in (None, "") for c in f)]
    ancho = len(titulos)
    filas = [list(f) + [None] * (ancho - len(f)) for f in filas]
    return titulos, filas


# --------------------------------------------------------------------------
# Qué columna es cada cosa
# --------------------------------------------------------------------------
PISTAS = {
    "numero": ("maquina", "maq", "mq", "tejedora", "numero", "nro", "n"),
    "marca": ("marca",),
    "modelo": ("modelo",),
    "galga": ("galga", "gg"),
    "diametro": ("diametro", "diam"),
    "alimentadores": ("alimentador", "caida", "caidas", "sistemas"),
    "agujas": ("cantidad de agujas", "total agujas", "agujas totales", "agujas"),
    "anio": ("anio", "ano", "year", "fabricacion"),
    "nota": ("nota", "observacion", "observaciones", "comentario"),
}

_FECHA = ("fecha", "ultim", "cuando")
_KG = ("kg", "kilo", "cada")


def detectar(titulos: list[str], tipos: list[dict]) -> dict[str, int]:
    """Propone qué columna es cada cosa. Devuelve {campo: índice de columna}."""
    limpios = [_limpio(t) for t in titulos]
    mapa: dict[str, int] = {}
    usadas: set[int] = set()

    def tomar(campo, prueba):
        if campo in mapa:
            return
        for i, t in enumerate(limpios):
            if i in usadas or not t:
                continue
            if prueba(t):
                mapa[campo] = i
                usadas.add(i)
                return

    # Primero los tipos de mantenimiento: sus títulos llevan el nombre del tipo
    # ("último cambio de agujas"). Si no los reservamos, la palabra "agujas" se
    # la lleva la columna de la ficha y la fecha queda sin mapear.
    for tipo in tipos:
        palabras = [p for p in _limpio(tipo["nombre"]).split() if len(p) > 3]
        if not palabras:
            continue

        def del_tipo(t, palabras=palabras):
            return any(p in t for p in palabras)

        tomar(f"fecha_{tipo['id']}", lambda t, d=del_tipo: d(t) and any(k in t for k in _FECHA))
        tomar(f"kg_{tipo['id']}", lambda t, d=del_tipo: d(t) and any(k in t for k in _KG))

    for campo, pistas in PISTAS.items():
        tomar(campo, lambda t, p=pistas: any(t == x or t.startswith(x + " ") or f" {x} " in f" {t} " for x in p))

    return mapa


# --------------------------------------------------------------------------
# Convertir las celdas en datos
# --------------------------------------------------------------------------
def a_numero(valor) -> int | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return int(valor)
    encontrados = re.findall(r"\d+", str(valor).replace(".", "").replace(",", ""))
    return int(encontrados[-1]) if encontrados else None


def a_decimal(valor) -> float | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip().replace(",", ".")
    m = re.search(r"-?\d+(\.\d+)?", s)
    return float(m.group()) if m else None


def a_decimal_es(valor) -> float | None:
    """Un número escrito como se escribe acá: punto de miles, coma decimal.

    `a_decimal` toma el punto como coma decimal, así que «1.410» rpm entraba
    como 1,41 y la cuenta de cuánto debería dar salía mal sin que nadie se
    diera cuenta. Éste guarda el signo, que `a_kilos` se lleva puesto.
    """
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    negativo = texto.startswith("-")
    limpio = re.sub(r"[^\d,.]", "", texto)
    if not limpio:
        return None
    limpio = limpio.replace(".", "").replace(",", ".")
    try:
        numero = float(limpio)
    except ValueError:
        return None
    return -numero if negativo else numero


def a_kilos(valor) -> float | None:
    """'30.000 kg' es treinta mil. El punto es de miles: en castellano nadie
    escribe 30.000 queriendo decir treinta."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    s = re.sub(r"[^\d,.]", "", str(valor))
    if not s:
        return None
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def a_fecha(valor) -> date | None:
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    s = str(valor).strip()
    for formato in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, formato).date()
        except ValueError:
            continue
    return None


def _texto(valor) -> str | None:
    s = str(valor).strip() if valor not in (None, "") else ""
    s = re.sub(r"\s+", " ", s)
    # Las celdas de planta traen el año pegado al final ("PILOTELLI 2013").
    s = re.sub(r"\s+(19|20)\d{2}$", "", s)
    return s or None


def _resumen(fila) -> str:
    partes = [str(c).strip() for c in fila[:4] if c not in (None, "")]
    return " · ".join(partes)[:80] or "(vacía)"


def armar(titulos, filas, mapa, maquinas, tipos, hoy=None):
    """Convierte las filas del Excel en datos listos para guardar.

    Devuelve (listas, descartes). Cada descarte dice qué fila y por qué, y se
    muestra en pantalla antes de confirmar. Lo que no se entiende no entra.
    """
    hoy = hoy or date.today()
    por_numero = {m["numero"]: m for m in maquinas if m.get("numero") is not None}
    listas, descartes = [], []

    def celda(fila, campo):
        i = mapa.get(campo)
        if i is None or i >= len(fila):
            return None
        return fila[i]

    for n, fila in enumerate(filas, start=1):
        numero = a_numero(celda(fila, "numero"))
        if numero is None:
            descartes.append({"fila": n, "texto": _resumen(fila),
                              "motivo": "No tiene número de máquina"})
            continue
        maquina = por_numero.get(numero)
        if not maquina:
            descartes.append({"fila": n, "texto": _resumen(fila),
                              "motivo": f"La máquina {numero} no está en Asinfo"})
            continue

        item = {
            "fila": n,
            "maquina": maquina,
            "ficha": {
                "marca": _texto(celda(fila, "marca")),
                "modelo": _texto(celda(fila, "modelo")),
                "galga": a_numero(celda(fila, "galga")),
                "diametro": a_decimal(celda(fila, "diametro")),
                "alimentadores": a_numero(celda(fila, "alimentadores")),
                "agujas": a_numero(celda(fila, "agujas")),
                "anio": a_numero(celda(fila, "anio")),
                "nota": _texto(celda(fila, "nota")),
            },
            "mantenimientos": [],
            "avisos": [],
        }

        for tipo in tipos:
            fecha = a_fecha(celda(fila, f"fecha_{tipo['id']}"))
            kg = a_kilos(celda(fila, f"kg_{tipo['id']}"))
            if fecha is None and kg is None:
                continue
            if fecha and fecha > hoy:
                item["avisos"].append(f"{tipo['nombre']}: la fecha es futura, no se carga")
                fecha = None
            if fecha is None and kg is not None:
                # Sin fecha no hay desde cuándo contar, y ponerle hoy no es
                # gratis: inventa un mantenimiento que le borra los kilos
                # acumulados a una máquina que ya venía contando. Guardamos
                # sólo el tope. Para empezar el conteo está /arranque.
                item["avisos"].append(
                    f"{tipo['nombre']}: sólo se guarda el tope, no se carga un mantenimiento")
            item["mantenimientos"].append({"tipo": tipo, "fecha": fecha, "cada_kg": kg})

        if not item["mantenimientos"] and not any(v for v in item["ficha"].values()):
            descartes.append({"fila": n, "texto": _resumen(fila), "motivo": "La fila está vacía"})
            continue
        listas.append(item)

    # Una máquina repetida en el Excel es un error de carga, no dos máquinas.
    veces: dict[int, int] = {}
    for item in listas:
        veces[item["maquina"]["id"]] = veces.get(item["maquina"]["id"], 0) + 1
    for item in listas:
        if veces[item["maquina"]["id"]] > 1:
            item["avisos"].append("Esta máquina aparece más de una vez en el Excel")

    return listas, descartes


# --------------------------------------------------------------------------
# La planilla que usan en planta: UNA HOJA POR MÁQUINA
# --------------------------------------------------------------------------
# Arriba la ficha de la máquina, abajo el historial desde 2018. Ninguna hoja
# está armada igual que la otra: los mismos datos aparecen en filas distintas,
# con nombres distintos ("AGUJAS" y "Cantidad de agujas"), y el diámetro y la
# galga a veces van juntos en una celda ("D 32   G24") y a veces separados
# bajo el título "DIAMETRO/ GG" ("32 / 24").
#
# Por eso no se lee por posición: se busca cada dato por su nombre, y lo que
# no aparece queda vacío. Vacío es correcto; inventado no.

# RELANIT y JVCE no van: son modelos, no marcas.
_MARCAS = ("MAYER", "TERROT", "FUKUHARA", "PAILUNG", "JIUNN LONG", "ORIZIO",
           "MONARCH", "UNITEX", "SANTONI", "WELLKNIT", "KEUMYONG")

# La misma marca escrita de varias formas. Relanit es una máquina MAYER, y
# JUNNLOG es JIUNN LONG mal tipeado: dejarlas como están parte la planta en
# marcas distintas que son la misma.
_MISMA_MARCA = {"relanit": "MAYER", "junnlog": "JIUNN LONG",
                "junn log": "JIUNN LONG", "jiunnlog": "JIUNN LONG",
                "jiunn log": "JIUNN LONG"}


def _marca_pareja(marca):
    """La marca escrita siempre igual. «MAYER MV4» es MAYER: MV4 es el modelo."""
    if not isinstance(marca, str):
        return marca
    texto = _limpio(marca)
    for canonica in _MARCAS:
        if texto.startswith(canonica.lower()):
            return canonica
    for escrita, canonica in _MISMA_MARCA.items():
        if texto.startswith(escrita):
            return canonica
    return marca

# El tipo no está en una columna: está escrito adentro del texto de la
# actividad. Cada uno se reconoce por su frase, no por la palabra suelta:
# «limpiesa de cilindro» dice cilindro y NO es un cambio de cilindro.
_ES_AGUJAS = re.compile(r"aguja", re.IGNORECASE)
# Tiene que decir CAMBIO: «limpiesa de cilindro» nombra al cilindro y no es un
# cambio de cilindro. El «(algo) y» del medio es para «cambio de agujas y
# platinas», que son dos cosas en un renglón. «ci,indro» y «cilndro» existen:
# la planilla está tipeada a mano.
_CAMBIO = r"cambi\w*\s+(?:de\s+)?(?:\w+\s+y\s+)?"
_ES_CILINDRO = re.compile(_CAMBIO + r"ci\W?l?\W?[ií]ndro", re.IGNORECASE)
_ES_PLATINAS = re.compile(_CAMBIO + r"platina", re.IGNORECASE)


def _clasificar(tipos):
    """Qué tipo de la base corresponde a cada frase de la planilla.

    Los tipos los define el mecánico en la pantalla de Tipos, así que acá no se
    inventa ninguno: se busca cuál de los que EXISTEN habla de agujas, de
    cilindro o de platinas. Si alguno no está cargado, esa frase cae en
    limpieza, como caía antes.
    """
    def buscar(palabra):
        return next((t for t in tipos if palabra in t["nombre"].lower()), None)

    especiales = [(_ES_AGUJAS, buscar("aguja")),
                  (_ES_CILINDRO, buscar("cilindro")),
                  (_ES_PLATINAS, buscar("platina"))]
    especiales = [(regla, tipo) for regla, tipo in especiales if tipo]
    ids = {t["id"] for _, t in especiales}
    # Limpieza es la que se llama así; si no está, la primera que no sea una de
    # las otras. Nunca se elige "la que sobra" a ciegas: con cuatro tipos
    # cargados, la que sobraba podía ser «Cambio de cilindro».
    limpieza = next((t for t in tipos if "limpi" in t["nombre"].lower()), None)
    if limpieza is None:
        limpieza = next((t for t in tipos if t["id"] not in ids), None)
    return especiales, limpieza


def _tipos_del_texto(texto, especiales, limpieza):
    """Todos los mantenimientos que menciona una fila, no sólo el primero.

    Un renglón puede decir «limpiesa de cilindro · cambio de platinas galga
    28»: ese día se hicieron dos cosas y son dos mantenimientos. Contar sólo el
    primero borraba el otro del historial de la máquina.
    """
    encontrados = []
    for regla, tipo in especiales:
        if texto and regla.search(texto) and tipo not in encontrados:
            encontrados.append(tipo)
    # La limpieza va si la nombra, o si la fila no dice ninguna otra cosa: la
    # planilla nació siendo un registro de limpiezas.
    if limpieza and (not encontrados or (texto and re.search(r"limpi", texto, re.I))):
        encontrados.insert(0, limpieza)
    return encontrados


def _celdas(filas, hasta):
    """El bloque de arriba como lista de (fila, columna, texto)."""
    salida = []
    for i, fila in enumerate(filas[:hasta]):
        for j, celda in enumerate(fila):
            if celda not in (None, ""):
                salida.append((i, j, celda))
    return salida


def _valor_de(bloque, etiquetas, filas, salteo=(), preferir="abajo"):
    """El valor que acompaña a una etiqueta: abajo o a la derecha.

    Se saltean las celdas que son otra etiqueta — si no, "Responsable:" se
    lleva el titulo de al lado en vez del nombre.
    """
    etiquetas = tuple(_limpio(e) for e in etiquetas)
    todas = tuple(_limpio(s) for s in salteo)
    for i, j, celda in bloque:
        if not isinstance(celda, str):
            continue
        texto = _limpio(celda)
        if not any(texto == e or texto.startswith(e) for e in etiquetas):
            continue
        abajo, derecha = [], []
        for di in (1, 2, 3):
            if i + di < len(filas) and j < len(filas[i + di]):
                abajo.append(filas[i + di][j])
        for dj in (1, 2, 3):
            if j + dj < len(filas[i]):
                derecha.append(filas[i][j + dj])
        candidatas = derecha + abajo if preferir == "derecha" else abajo + derecha
        for c in candidatas:
            if c in (None, ""):
                continue
            if isinstance(c, str):
                limpio = _limpio(c)
                if not limpio or limpio in todas or any(limpio.startswith(e) for e in etiquetas):
                    continue
            return c
    return None


def _fila_de_titulos(filas):
    """Dónde arranca el historial: la fila que dice «Fecha»."""
    for i, fila in enumerate(filas):
        for celda in fila:
            if isinstance(celda, str) and _limpio(celda) == "fecha":
                return i
    return None


def _segunda_ficha(filas, hasta):
    """En qué columna empieza la ficha de OTRA máquina, si está pegada al lado.

    Seis hojas tienen dos encabezados uno junto al otro: el de la máquina y el
    de la de al lado, copiado. Todo lo que está a la derecha del segundo no es
    de esta máquina — la MQ 61 salía con la marca de una y las agujas de la
    otra. Se corta ahí y no se mira más a la derecha.
    """
    columnas = {}
    for fila in filas[:hasta]:
        for j, celda in enumerate(fila):
            if not isinstance(celda, str):
                continue
            texto = _limpio(celda)
            for etiqueta in ("equipo", "tipo de aguja", "modelo",
                             "cantidad de agujas", "responsable"):
                if texto.startswith(etiqueta):
                    columnas.setdefault(etiqueta, set()).add(j)
    segundas = [sorted(cols)[1] for cols in columnas.values() if len(cols) > 1]
    # Con UNA etiqueta repetida no alcanza: en dos hojas «CANTIDAD DE AGUJAS»
    # está escrito dos veces en el mismo encabezado y cortar ahí se llevaba
    # puesto el número de serie. Una ficha entera repite varias.
    return min(segundas) if len(segundas) >= 2 else None


def _columna_entera(filas, hasta, desde_fila, columna):
    """Todo lo que cuelga de una etiqueta, hacia abajo, en su misma columna.

    «TIPO DE AGUJAS» no trae un código: trae los cuatro del cilindro y el del
    plato, uno por renglón. Guardar sólo el primero es guardar un cuarto de la
    ficha. Dos renglones vacíos seguidos cortan: ahí terminó la lista.
    """
    valores, vacios = [], 0
    for i in range(desde_fila + 1, hasta):
        celda = filas[i][columna] if columna < len(filas[i]) else None
        if celda in (None, ""):
            vacios += 1
            if vacios >= 2 and valores:
                break
            continue
        vacios = 0
        texto = str(celda).strip()
        if texto and texto not in valores:
            valores.append(texto)
    return valores


def _lista_de(filas, hasta, etiquetas, corte_col=None):
    """Los valores que cuelgan de una etiqueta, juntos en un solo texto."""
    etiquetas = tuple(_limpio(e) for e in etiquetas)
    for i, fila in enumerate(filas[:hasta]):
        for j, celda in enumerate(fila):
            if corte_col is not None and j >= corte_col:
                continue
            if not isinstance(celda, str):
                continue
            if not any(_limpio(celda).startswith(e) for e in etiquetas):
                continue
            valores = _columna_entera(filas, hasta, i, j)
            if valores:
                return " · ".join(valores)
    return None


def _espejo(filas, hasta, corte_col, etiquetas, saltar, sirve=None):
    """El dato que está sin título, leído con los títulos de la ficha de al lado.

    Las dos fichas pegadas son la misma plantilla corrida `corte_col` columnas.
    A la MQ 61 se le borraron los títulos y quedaron sólo los de la copia: sin
    esto, su número de serie se leía del de la máquina de al lado —un dato de
    otra máquina, que es peor que no tener ninguno.
    """
    etiquetas = tuple(_limpio(e) for e in etiquetas)
    saltar = tuple(_limpio(s) for s in saltar)
    for fila in filas[:hasta]:
        for j, celda in enumerate(fila):
            if j < corte_col or not isinstance(celda, str):
                continue
            if not any(_limpio(celda).startswith(e) for e in etiquetas):
                continue
            columna = j - corte_col
            for otra in filas[:hasta]:
                if columna >= len(otra):
                    continue
                valor = otra[columna]
                if valor in (None, ""):
                    continue
                if isinstance(valor, str):
                    limpio = _limpio(valor)
                    # «R01 Registro de Mantenimiento» es el título de la hoja y
                    # está en cualquier columna: nunca es el dato.
                    if (not limpio or limpio.startswith("r01")
                            or any(limpio.startswith(s) for s in saltar)):
                        continue
                if sirve and not sirve(valor):
                    continue
                return valor
    return None


def _ficha_de_hoja(filas, corte):
    # Primero se saca de encima la ficha de la máquina de al lado, si la hay:
    # todo lo que sigue lee este bloque como si fuera la única.
    filas_todas = filas
    corte_col = _segunda_ficha(filas, corte)
    if corte_col is not None:
        filas = [[c for j, c in enumerate(fila) if j < corte_col] for fila in filas]

    bloque = _celdas(filas, corte)
    etiquetas = ("equipo", "modelo", "numero", "responsable", "tipo de aguja",
                 "tipo de agujas", "cantidad de agujas", "agujas", "aguujas",
                 "codigo agujas", "codigo de agujas",
                 "año de fabricacion", "diametro gg", "fecha", "observaciones",
                 "repuestos", "actividad realizada", "tipo de mantenimiento")

    texto_todo = " ".join(str(c) for _, _, c in bloque)

    # La marca esta al lado de "CIRCULAR": asi esta armada la planilla. Buscarla
    # suelta en todo el encabezado se equivoca, porque los nombres aparecen
    # tambien adentro de los codigos de aguja.
    # A la derecha, no abajo: abajo de "CIRCULAR" hay un codigo interno.
    marca = _valor_de(bloque, ("circular",), filas, etiquetas, preferir="derecha")
    if not isinstance(marca, str) or _limpio(marca) in ("", "circular"):
        marca = next((m for m in _MARCAS if m.lower() in texto_todo.lower()), None)
    marca = _marca_pareja(marca)
    modelo = _valor_de(bloque, ("modelo",), filas, etiquetas)
    serie = _valor_de(bloque, ("numero", "n de serie", "serie"), filas, etiquetas)
    responsable = _valor_de(bloque, ("responsable",), filas, etiquetas)
    anio = a_numero(_valor_de(bloque, ("año de fabricacion", "ano de fabricacion"),
                              filas, etiquetas))
    if anio is not None and not (1970 <= anio <= 2035):
        anio = None
    if anio is None:
        # Muchas hojas ponen el año suelto, sin etiqueta. Un numero entre 1990
        # y hoy es un año; la cantidad de agujas (2460, 3168) queda afuera del
        # rango, y el numero de serie tiene cinco cifras.
        for _, _, celda in bloque:
            if isinstance(celda, int) and 1990 <= celda <= date.today().year:
                anio = celda
                break

    # Diámetro y galga: juntos ("D 32  G24") o bajo "DIAMETRO/ GG" ("32 / 24").
    diametro = galga = None
    junto = re.search(r"D\s*(\d{2})\s*[/xX]?\s*G+\s*(\d{1,2})", texto_todo)
    if junto:
        diametro, galga = int(junto.group(1)), int(junto.group(2))
    else:
        crudo = _valor_de(bloque, ("diametro gg", "diametro"), filas, etiquetas)
        partido = re.findall(r"\d{1,2}", str(crudo or ""))
        if len(partido) >= 2:
            diametro, galga = int(partido[0]), int(partido[1])

    alimentadores = None
    ali = re.search(r"(\d{2,3})\s*ALIMENTADORES", texto_todo, re.IGNORECASE)
    if ali:
        alimentadores = int(ali.group(1))

    # Agujas: el numero grande que acompaña a "agujas".
    agujas = None
    for etiqueta in ("cantidad de agujas", "agujas", "n agujas"):
        agujas = a_numero(_valor_de(bloque, (etiqueta,), filas, etiquetas))
        if agujas and 500 <= agujas <= 20000:
            break
        agujas = None

    # Los códigos de aguja van uno abajo del otro: se guardan todos. Hay hojas
    # que a la columna la llaman «CODIGO AGUJAS» en vez de «TIPO DE AGUJAS».
    nombres_agujas = ("tipo de agujas", "tipo de aguja", "codigo agujas",
                      "codigo de agujas")
    tipo_agujas = _lista_de(filas, corte, nombres_agujas)
    if not tipo_agujas:
        tipo_agujas = _valor_de(bloque, nombres_agujas, filas, etiquetas)

    # Lo que quedó sin llenar y la ficha de al lado sí tiene titulado.
    if corte_col is not None:
        if not modelo:
            modelo = _espejo(filas_todas, corte, corte_col, ("modelo",), etiquetas)
        if not serie:
            serie = _espejo(filas_todas, corte, corte_col,
                            ("numero", "n de serie", "serie"), etiquetas)
        if not agujas:
            def _es_cantidad(valor):
                n = a_numero(valor)
                return n is not None and 500 <= n <= 20000
            agujas = a_numero(_espejo(filas_todas, corte, corte_col,
                                      ("cantidad de agujas", "agujas"), etiquetas,
                                      sirve=_es_cantidad))

    return {
        "marca": _texto(marca),
        "modelo": _texto(modelo),
        "galga": galga,
        "diametro": diametro,
        "alimentadores": alimentadores,
        "agujas": agujas,
        "anio": anio,
        "serie": _texto(serie),
        "tipo_agujas": _texto(tipo_agujas),
        "nota": _texto(responsable) and f"Responsable: {_texto(responsable)}",
    }


# Una fecha escrita a mano en una celda de texto. En la planilla hay doce, y
# vienen de todas las formas: «01/19/2021», «23/12,/2018», «6/8//2018».
_FECHA_A_MANO = re.compile(r"^\s*(\d{1,2})\D+(\d{1,2})\D+(\d{2,4})\s*$")


def fecha_escrita(texto, hoy=None, antes=None, despues=None):
    """Una fecha tipeada a mano, resuelta con las fechas de arriba y de abajo.

    La planilla mezcla los dos órdenes — «01/19/2021» es mes primero y
    «23/12,/2018» es día primero — y además tiene años cortados: «205», «220».
    Por sí solo, «6/8/2018» puede ser el 6 de agosto o el 8 de junio.

    Pero el historial está en orden: la fila de arriba es anterior y la de
    abajo, posterior. Eso alcanza casi siempre para que quede UNA sola opción
    posible, y no es adivinar — es usar el orden en que está escrita la hoja.

    `antes` y `despues` son las fechas conocidas más cercanas, arriba y abajo.
    Si con eso siguen quedando dos opciones, se devuelve None: elegir movería
    el mantenimiento de mes, y con eso se corren los kilos.
    """
    hoy = hoy or date.today()
    m = _FECHA_A_MANO.match(str(texto or ""))
    if not m:
        return None
    a, b, anio_texto = int(m.group(1)), int(m.group(2)), m.group(3)

    # Qué puede ser día y qué mes. Un número mayor que 12 sólo puede ser día.
    if a > 12 and b <= 12:
        pares = [(a, b)]
    elif b > 12 and a <= 12:
        pares = [(b, a)]
    elif a <= 12 and b <= 12:
        pares = [(a, b), (b, a)]
    else:
        return None

    # Qué puede ser el año. Con cuatro dígitos no hay duda; con dos, 20xx. Si
    # está cortado («205»), lo dicen los vecinos: se prueban todos los años del
    # tramo y sobrevive el que encaja.
    if len(anio_texto) == 4:
        anios = [int(anio_texto)]
    elif len(anio_texto) == 2:
        anios = [2000 + int(anio_texto)]
    else:
        desde = max(1990, antes.year if antes else 1990)
        hasta = min(hoy.year, despues.year if despues else hoy.year)
        anios = list(range(desde, hasta + 1))

    posibles = set()
    for dia, mes in pares:
        for anio in anios:
            try:
                f = date(anio, mes, dia)
            except ValueError:
                continue
            if not (1990 <= f.year and f <= hoy):
                continue
            if antes and f < antes:
                continue
            if despues and f > despues:
                continue
            posibles.add(f)

    return posibles.pop() if len(posibles) == 1 else None


def _columnas_historial(fila):
    """Qué columna es cada cosa en la tabla de abajo de la hoja.

    Los títulos no están siempre en el mismo lugar y a veces hay dos que dicen
    «Fecha». Lo que importa es dónde quedaron la actividad, los repuestos y las
    observaciones: la fecha se busca por su valor, no por su columna.
    """
    mapa = {}
    for i, celda in enumerate(fila):
        t = _limpio(celda)
        if not t:
            continue
        for campo, etiquetas in (
            ("tipo", ("tipo de mantenimiento", "tipo")),
            ("actividad", ("actividad realizada", "actividad")),
            ("repuestos", ("repuestos", "repuesto")),
            ("observaciones", ("observaciones", "observacion")),
            ("responsable", ("responsable", "hecho por", "quien")),
        ):
            if campo not in mapa and any(t == e or t.startswith(e) for e in etiquetas):
                mapa[campo] = i
                break
    return mapa


def leer_historial_por_maquina(ruta, maquinas, tipos, hoy=None):
    """El historial COMPLETO de la planilla de planta: todas las filas.

    `leer_por_maquina` se queda con la ÚLTIMA fecha de cada tipo, porque eso es
    lo único que necesita el semáforo. Pero la planilla tiene el historial
    entero desde 2018 — 1.366 filas — y ése es el valor de la planilla: poder
    mirar qué se le hizo a una máquina y cuándo.

    Devuelve (mantenimientos, fichas, descartes).
    """
    hoy = hoy or date.today()
    por_numero = {m["numero"]: m for m in maquinas if m.get("numero") is not None}
    especiales, limpieza = _clasificar(tipos)

    wb = load_workbook(ruta, read_only=True, data_only=True)
    mantenimientos, fichas, descartes = [], [], []
    try:
        for nombre_hoja in wb.sheetnames:
            numero = a_numero(nombre_hoja)
            maquina = por_numero.get(numero) if numero is not None else None
            if not maquina:
                descartes.append({"donde": nombre_hoja,
                                  "motivo": f"La máquina {numero} no está en Asinfo"})
                continue

            filas = [list(f) for f in wb[nombre_hoja].iter_rows(values_only=True)]
            corte = _fila_de_titulos(filas)
            sin_titulos = corte is None
            if sin_titulos:
                # Seis hojas no tienen la fila que dice «Fecha»: el historial
                # arranca directo, sin encabezado. Antes se las daba por
                # vacías y se perdían enteras — la MQ 22 tiene 22
                # mantenimientos desde 2019. Cuando no hay títulos, el
                # historial empieza en la primera fila que trae una fecha de
                # verdad.
                primera = next((i for i, f in enumerate(filas)
                                if any(isinstance(c, datetime) for c in f)), None)
                if primera is None:
                    descartes.append({"donde": nombre_hoja,
                                      "motivo": "La hoja no tiene ninguna fecha"})
                    continue
                corte = primera - 1

            # Dónde termina la ficha. Sin fila de títulos la ficha llega hasta
            # el primer mantenimiento: con `corte` se perdía el último renglón
            # del encabezado, y si el historial arrancaba en la primera fila,
            # `corte` valía -1 y la ficha se leía de la hoja ENTERA — marca,
            # serie y año salían de renglones de mantenimiento.
            fin_ficha = primera if sin_titulos else corte
            ficha = _ficha_de_hoja(filas, fin_ficha)
            fichas.append((maquina, ficha))
            mapa = {} if sin_titulos else _columnas_historial(filas[corte])

            # El responsable de la hoja NO se copia a las 1.300 filas: es quién
            # tiene la máquina a cargo, no quién hizo cada mantenimiento. La
            # planilla no dice quién hizo cada uno, así que no se inventa.
            # Queda en la ficha, que es donde está escrito.

            def celda(fila, campo):
                i = mapa.get(campo)
                return fila[i] if i is not None and i < len(fila) else None

            cuerpo = filas[corte + 1:]

            # Primero las fechas que Excel guardó como fecha. Se necesitan
            # antes de leer nada, porque son las que ubican en el tiempo a las
            # que están tipeadas a mano.
            ciertas = []
            for fila in cuerpo:
                f = next((c.date() for c in fila if isinstance(c, datetime)), None)
                ciertas.append(f if f and 1990 <= f.year and f <= hoy else None)

            def vecina(i, paso):
                """La fecha conocida más cercana hacia arriba o hacia abajo."""
                j = i + paso
                while 0 <= j < len(ciertas):
                    if ciertas[j]:
                        return ciertas[j]
                    j += paso
                return None

            leidas = 0
            for n, fila in enumerate(cuerpo, start=1):
                fecha = ciertas[n - 1]
                if fecha is None:
                    # Tipeada a mano. El historial está en orden, así que la
                    # fecha tiene que caer entre la de arriba y la de abajo:
                    # con eso casi siempre queda una sola opción posible.
                    for celda_texto in fila[:3]:
                        if not isinstance(celda_texto, str):
                            continue
                        if not _FECHA_A_MANO.match(celda_texto):
                            continue
                        fecha = fecha_escrita(celda_texto, hoy,
                                              antes=vecina(n - 1, -1),
                                              despues=vecina(n - 1, +1))
                        if fecha is None:
                            descartes.append({
                                "donde": f"{nombre_hoja}, fila {n}",
                                "motivo": f"La fecha «{celda_texto.strip()}» está "
                                          "escrita a mano y ni con las fechas de "
                                          "arriba y abajo se puede saber cuál es"})
                        break
                if not fecha:
                    continue

                # El tipo no está en una columna: está escrito adentro del
                # texto. Si dice «aguja» es cambio de agujas; si no, limpieza.
                partes = [_texto(celda(fila, c)) for c in ("tipo", "actividad")]
                actividad = " · ".join(p for p in partes if p) or None
                if actividad is None and not (mapa.get("tipo") or mapa.get("actividad")):
                    # Sólo si la hoja no tiene esas columnas. Si las tiene y
                    # vinieron vacías, la fila no dice qué se hizo — y juntar
                    # todo el texto suelto traería la columna «Próximo», que es
                    # un kilaje, no una actividad.
                    #
                    # Se toma lo que está DESPUÉS de la fecha: a la izquierda
                    # queda el modelo de la máquina repetido en cada renglón,
                    # que no es lo que se hizo.
                    desde_col = next((i for i, c in enumerate(fila)
                                      if isinstance(c, datetime)), -1) + 1
                    actividad = " · ".join(
                        str(c).strip() for c in fila[desde_col:]
                        if isinstance(c, str) and c.strip())[:300] or None
                observaciones = _texto(celda(fila, "observaciones"))
                texto = " · ".join(p for p in (actividad, observaciones) if p)
                del_dia = _tipos_del_texto(texto, especiales, limpieza)
                if not del_dia:
                    continue

                # Una fila puede ser dos mantenimientos. La clave (hoja, orden)
                # tiene que seguir siendo única, así que la fila 7 pasa a ser
                # 70, 71, 72: siguen en orden y no chocan con las de al lado.
                for k, tipo in enumerate(del_dia):
                    mantenimientos.append({
                        "id_maquina": maquina["id"],
                        "maquina_nombre": maquina["nombre"],
                        "tipo_id": tipo["id"],
                        "fecha": fecha,
                        "hecho_por": _texto(celda(fila, "responsable")) or "Planilla de planta",
                        "nota": texto or None,
                        "repuestos": _texto(celda(fila, "repuestos")),
                        "horas": None,
                        "hoja": nombre_hoja,
                        "orden": n * 10 + k,
                    })
                leidas += 1

            if not leidas:
                descartes.append({"donde": nombre_hoja,
                                  "motivo": "La hoja no tiene ninguna fecha cargada"})
    finally:
        wb.close()

    return mantenimientos, fichas, descartes


def leer_por_maquina(ruta, maquinas, tipos, hoy=None):
    """Lee la planilla de planta: una hoja por máquina.

    De cada hoja saca la ficha y la ÚLTIMA fecha de cada tipo de mantenimiento.
    Devuelve (listas, descartes), igual que `armar`, para que la pantalla de
    revisión sea la misma.
    """
    hoy = hoy or date.today()
    por_numero = {m["numero"]: m for m in maquinas if m.get("numero") is not None}
    especiales, limpieza = _clasificar(tipos)

    wb = load_workbook(ruta, read_only=True, data_only=True)
    listas, descartes = [], []
    try:
        for nombre_hoja in wb.sheetnames:
            numero = a_numero(nombre_hoja)
            if numero is None:
                descartes.append({"fila": nombre_hoja, "texto": "",
                                  "motivo": "La hoja no dice de qué máquina es"})
                continue
            maquina = por_numero.get(numero)
            if not maquina:
                descartes.append({"fila": nombre_hoja, "texto": f"MQ {numero}",
                                  "motivo": f"La máquina {numero} no está en Asinfo"})
                continue

            filas = [list(f) for f in wb[nombre_hoja].iter_rows(values_only=True)]
            corte = _fila_de_titulos(filas)
            if corte is None:
                descartes.append({"fila": nombre_hoja, "texto": f"MQ {numero}",
                                  "motivo": "La hoja está vacía"})
                continue

            ficha = _ficha_de_hoja(filas, corte)

            # Última fecha de cada tipo. El tipo no está en una columna: está
            # escrito adentro del texto, así que se decide por lo que dice.
            ultimas: dict[int, date] = {}
            cuantas = 0
            for fila in filas[corte + 1:]:
                fecha = next((c.date() for c in fila if isinstance(c, datetime)), None)
                if not fecha or fecha > hoy:
                    continue
                cuantas += 1
                texto = " ".join(str(c) for c in fila if isinstance(c, str))
                # Una fila puede ser dos mantenimientos: cuentan los dos.
                for tipo in _tipos_del_texto(texto, especiales, limpieza):
                    if fecha > ultimas.get(tipo["id"], date.min):
                        ultimas[tipo["id"]] = fecha

            if not ultimas:
                descartes.append({"fila": nombre_hoja, "texto": f"MQ {numero}",
                                  "motivo": "La hoja no tiene ninguna fecha cargada"})
                continue

            item = {
                "fila": nombre_hoja,
                "maquina": maquina,
                "ficha": ficha,
                "mantenimientos": [],
                "avisos": [],
            }
            for tipo in tipos:
                fecha = ultimas.get(tipo["id"])
                if not fecha:
                    continue
                item["mantenimientos"].append(
                    {"tipo": tipo, "fecha": fecha, "cada_kg": None})
                meses = (hoy - fecha).days // 30
                if meses >= 4:
                    item["avisos"].append(
                        f"{tipo['nombre']}: el último anotado es de hace {meses} meses")
            item["avisos"].append(f"{cuantas} registros en la hoja")
            listas.append(item)
    finally:
        wb.close()

    return listas, descartes
