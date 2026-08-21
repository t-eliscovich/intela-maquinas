"""Tests que corren ANTES de cada deploy. Si algo falla, no se deploya.

Cubren lo que ya nos mordió una vez:
  * que la app se pueda IMPORTAR como la importa Waitress (`app:app`), que es
    donde vivía el bug del pool sin inicializar;
  * que con la base caída avise en vez de tirar 500;
  * la aritmética del semáforo (va por kilos, y el tope puede ser propio
    de cada máquina);
  * que un mantenimiento SIN tope de kilos no prenda ningún color: el cambio
    de agujas no se hace por desgaste, así que no puede pintar de rojo;
  * que la campanita nunca muestre cero cuando Asinfo no contesta;
  * que el Excel de planta se lea aunque tenga los títulos donde quiera, y que
    lo que no se entiende quede afuera en vez de entrar mal;
  * que la planilla de planta traiga el historial ENTERO y que volver a
    cargarla no duplique nada;
  * que el arranque en lote no duplique ni acepte fechas futuras;
  * que ninguna entrada llegue cruda al SQL de Asinfo.
"""
import io, os, sys, tempfile, types
from datetime import date, timedelta, datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("MAQUINAS_DATABASE_URL", "postgresql://fake/fake")

def _stub(pool_ok=True):
    for n in ("psycopg2", "psycopg2.extras", "psycopg2.pool", "requests"):
        sys.modules[n] = types.ModuleType(n)
    sys.modules["psycopg2.extras"].RealDictCursor = object
    if pool_ok:
        class P:
            def __init__(s, *a, **k): pass
        sys.modules["psycopg2.pool"].SimpleConnectionPool = P
    else:
        def roto(*a, **k):
            raise RuntimeError("could not connect to server: Connection refused")
        sys.modules["psycopg2.pool"].SimpleConnectionPool = roto
    sys.modules["psycopg2"].extras = sys.modules["psycopg2.extras"]
    sys.modules["psycopg2"].pool = sys.modules["psycopg2.pool"]

fallos = []
def check(nombre, cond):
    print(("  OK   " if cond else "  FALLA ") + nombre)
    if not cond:
        fallos.append(nombre)


def _mensaje_de(hacer):
    """El texto del error que tira `hacer`, o vacío si no tira ninguno.

    Media docena de controles no alcanza con que frenen: tienen que decir en
    castellano qué pasó. Un ValueError con el mensaje de Python adentro, en
    planta, no le dice nada a nadie.
    """
    try:
        hacer()
    except Exception as exc:  # noqa: BLE001
        return str(exc)
    return ""


def _atrapar(hacer):
    """El error que tira `hacer`, para mirar de qué clase es."""
    try:
        hacer()
    except Exception as exc:  # noqa: BLE001
        return exc
    return None

# --- 1. base caída: importa igual y AVISA ---------------------------------
_stub(pool_ok=False)
import config; config.DATABASE_URL = "postgresql://fake/fake"; config.PASSWORD = ""
import store, asinfo, app as A
# Las de verdad, antes de que las tapemos con las de mentira. Los tests de
# mas abajo prueban lo que hacen ELLAS: el guard de truncado, el corte en
# lotes de 400 pares y el tope de meses. Con la version de mentira puesta no
# se probaria nada.
_ASINFO_REAL = {n: getattr(asinfo, n) for n in
                ("_consultar", "acumulados", "kilos_desde", "produccion_mensual")}
_CONSULTAR_REAL = _ASINFO_REAL["_consultar"]
A.app.config["TESTING"] = True
c = A.app.test_client()
print("Base caída:")
check("importa sin reventar (asi la importa waitress)", A.ERROR_ARRANQUE is not None)
r = c.get("/healthz")
check("healthz devuelve 503 y explica", r.status_code == 503 and "error_arranque" in r.get_json())
# `version` sale de un archivo que escribe el updater y ya se quedó pegado en un
# deploy viejo una vez. `desplegado` sale del código mismo: dice cuándo se
# instaló lo que está corriendo, y no depende de que nadie lo escriba.
check("healthz dice cuando se desplego el codigo que corre",
      r.get_json().get("desplegado") is not None)
r = c.get("/")
check("pantalla explica en vez de 500", r.status_code == 503 and "No hay conexión" in r.get_data(as_text=True))

# --- 2. seguridad del SQL a Asinfo ----------------------------------------
print("Asinfo:")
vistos = []
asinfo._consultar = lambda sql: (vistos.append(sql), [])[1]
asinfo._cache.clear()
asinfo.acumulados([(10, 1, "2026-01-01'; DROP TABLE maquina--")])
check("inyeccion neutralizada", "DROP" not in vistos[0].upper() and "--" not in vistos[0])
for mala in ("no-soy-fecha", ""):
    try:
        asinfo._cache.clear(); asinfo.acumulados([(10, 1, mala)]); ok = False
    except ValueError:
        ok = True
    check(f"fecha invalida rechazada ({mala!r})", ok)
check("guard de truncado existe", hasattr(asinfo, "RespuestaTruncada"))

# --- 3. el semáforo va por kilos ------------------------------------------
print("Semaforo:")
hoy = date.today()
TIPOS = [{"id": 1, "nombre": "Cambio de agujas", "cada_kg": 50000,
          "cada_rollos": None, "cada_dias": None, "activo": True}]
MAQS = [{"id": 10, "codigo": "1", "nombre": "TEJEDURIA-MQ 001", "numero": 1},
        {"id": 11, "codigo": "2", "nombre": "TEJEDURIA-MQ 002", "numero": 2},
        {"id": 12, "codigo": "3", "nombre": "TEJEDURIA-MQ 003", "numero": 3}]
KG = {10: 80000.0, 11: 46000.0, 12: 1000.0}   # pasado / falta poco / tranquilo

store.tipos = lambda incluir_inactivos=False: TIPOS
store.historial = lambda id_maquina=None, limite=200: []
store.responsables = lambda: []
store.ficha = lambda id_maquina: {}
store.topes_por_maquina = lambda: {}
store.ultimos_por_maquina_y_tipo = lambda: {
    (m["id"], 1): {"fecha": hoy - timedelta(days=30), "hecho_por": "x"} for m in MAQS
}
asinfo.maquinas = lambda: (MAQS, datetime.utcnow(), True)
asinfo.acumulados = lambda pares: (
    {(m, t): (KG[m], int(KG[m] / 20)) for m, t, d in pares}, datetime.utcnow(), True)
asinfo.produccion_mensual = lambda id_maquina, meses=12: ([], datetime.utcnow(), True)
A.ERROR_ARRANQUE = None

# Una fila por MÁQUINA: adentro, `principal` es el mantenimiento que le pone el
# color y `por_tipo` los tiene a todos. Antes era una fila por (máquina, tipo) y
# la misma máquina salía dos veces con dos colores distintos.
filas, pend, _, _ = A.armar_semaforo()
por_maquina = {f["maquina"]["id"]: f for f in filas}
check("hay una sola fila por maquina", len(filas) == len(MAQS))
check("lo mas pasado va primero", filas[0]["maquina"]["id"] == 10)
check("pasado de kilos = vencido", por_maquina[10]["estado"] == "vencido")
check("92% = falta poco", por_maquina[11]["estado"] == "por_vencer")
# El aviso arranca al 90%: al 84% todavía está en regla. Antes prendía al 80%.
check("84% todavia esta en regla", A.AVISO == 0.90 and 0.84 < A.AVISO)
check("2% = ok", por_maquina[12]["estado"] == "ok")
check("dice cuanto se paso", round(por_maquina[10]["principal"]["falta"]) == -30000)
check("el color lo pone el mantenimiento principal",
      por_maquina[10]["principal"]["estado"] == "vencido"
      and por_maquina[10]["pct"] == por_maquina[10]["principal"]["pct"])
check("cada tipo queda igual adentro de la fila",
      por_maquina[10]["por_tipo"][1]["kg"] == 80000.0)
check("los dias no prenden nada", por_maquina[12]["principal"]["dias"] == 30
      and por_maquina[12]["estado"] == "ok")

# Sin tope no se inventa un estado: se dice que falta el número, y la fila no
# se pinta de ningún color.
TIPOS[0]["cada_kg"] = None
filas, _, _, _ = A.armar_semaforo()
check("sin tope no pinta nada",
      all(f["estado"] == "sin_tope" and f["pct"] is None for f in filas))
# Pero el dato se muestra igual: cuántos kilos lleva desde la última vez. Un
# guión no dice nada; el número deja decidir a ojo hasta que haya tope.
por_maquina = {f["maquina"]["id"]: f for f in filas}
check("sin tope igual dice cuanto lleva desde la ultima vez",
      por_maquina[10]["principal"]["kg"] == 80000.0
      and por_maquina[10]["principal"]["tope"] is None
      and por_maquina[10]["principal"]["falta"] is None)
TIPOS[0]["cada_kg"] = 50000

# --- 3b. un tipo sin tope de kilos NO prende el semáforo -------------------
# El cambio de agujas no se hace ni por tiempo ni por desgaste: lo pide la tela.
# Así que no puede pintar una máquina de rojo — va al lado, como fecha.
print("Un tipo sin kilos no prende:")
DOS_TIPOS = [{"id": 1, "nombre": "Cambio de agujas", "cada_kg": None,
              "cada_rollos": None, "cada_dias": None, "activo": True},
             {"id": 2, "nombre": "Limpieza", "cada_kg": 50000,
              "cada_rollos": None, "cada_dias": None, "activo": True}]
store.tipos = lambda incluir_inactivos=False: DOS_TIPOS
store.ultimos_por_maquina_y_tipo = lambda: {
    (m["id"], t["id"]): {"fecha": hoy - timedelta(days=30), "hecho_por": "x"}
    for m in MAQS for t in DOS_TIPOS
}
filas, _, _, _ = A.armar_semaforo()
por_maquina = {f["maquina"]["id"]: f for f in filas}
check("el color lo pone el tipo que tiene tope",
      por_maquina[10]["principal"]["tipo"]["id"] == 2)
check("el que no tiene tope queda como dato, sin color",
      por_maquina[10]["por_tipo"][1]["estado"] == "sin_tope"
      and por_maquina[10]["por_tipo"][1]["desde"] == hoy - timedelta(days=30))
check("y el que tiene tope si prende", por_maquina[10]["estado"] == "vencido")
check("la maquina tranquila sigue en verde", por_maquina[12]["estado"] == "ok")

# Una máquina a la que le falta el tope tiene que hablar del mantenimiento que
# PRENDE el semáforo, no del último que se le hizo. La MQ 52 decía «sin tope ·
# Cambio de agujas», que es un tope que nadie va a poner nunca.
DOS_TIPOS[1]["cada_kg"] = None      # el tope va por máquina, no por tipo
store.topes_por_maquina = lambda: {(11, 2): 50000.0}
store.ultimos_por_maquina_y_tipo = lambda: {
    # A la 10 le falta el tope de limpieza y lo último fue un cambio de agujas.
    (10, 1): {"fecha": hoy, "hecho_por": "x"},
    (10, 2): {"fecha": hoy - timedelta(days=200), "hecho_por": "x"},
    # A la 12 nunca le hicieron una limpieza: sólo tiene el cambio de agujas.
    (12, 1): {"fecha": hoy, "hecho_por": "x"},
}
filas, _, _, _ = A.armar_semaforo()
por_maquina = {f["maquina"]["id"]: f for f in filas}
check("sin tope, la fila habla de la limpieza y no del cambio de agujas",
      por_maquina[10]["estado"] == "sin_tope"
      and por_maquina[10]["principal"]["tipo"]["id"] == 2)
check("si nunca le hicieron una limpieza, lo que falta es arrancar",
      por_maquina[12]["estado"] == "sin_arrancar")
DOS_TIPOS[1]["cada_kg"] = 50000

store.ultimos_por_maquina_y_tipo = lambda: {}
filas, _, _, _ = A.armar_semaforo()
check("sin ningun mantenimiento la maquina esta sin arrancar",
      all(f["estado"] == "sin_arrancar" and f["principal"] is None for f in filas))

store.tipos = lambda incluir_inactivos=False: TIPOS
store.ultimos_por_maquina_y_tipo = lambda: {
    (m["id"], 1): {"fecha": hoy - timedelta(days=30), "hecho_por": "x"} for m in MAQS
}

# --- 4. el tope propio de cada máquina manda -------------------------------
print("Tope por maquina:")
store.topes_por_maquina = lambda: {(12, 1): 500.0}   # MQ 3 teje poco: tope chico
filas, _, _, _ = A.armar_semaforo()
por_maquina = {f["maquina"]["id"]: f["principal"] for f in filas}
check("el tope propio pisa al del tipo", por_maquina[12]["tope"] == 500.0)
check("y con eso la maquina chica se prende", por_maquina[12]["estado"] == "vencido")
check("la que no tiene tope propio usa el del tipo", por_maquina[10]["tope"] == 50000)
check("se marca cual es propio", por_maquina[12]["tope_propio"] is True
      and por_maquina[10]["tope_propio"] is False)
store.topes_por_maquina = lambda: {}

# --- 5. la campanita nunca miente -----------------------------------------
print("Campanita:")
with A.app.test_request_context("/"):
    check("cuenta las vencidas", A._campanita()["vencidas"] == 1)

def _asinfo_caido():
    raise asinfo.AsinfoNoDisponible("timeout")

guardado = asinfo.maquinas
asinfo.maquinas = lambda: _asinfo_caido()
with A.app.test_request_context("/"):
    check("con Asinfo caido NO dice cero", A._campanita()["vencidas"] is None)
asinfo.maquinas = guardado

# --- 6. leer el Excel de planta -------------------------------------------
print("Excel:")
import excel
titulos = ["Máquina", "Último cambio de agujas", "Cambio agujas cada kg",
           "Marca", "Modelo", "Galga", "Año"]
mapa = excel.detectar(titulos, TIPOS)
check("encuentra la columna de la maquina", mapa.get("numero") == 0)
check("encuentra la fecha del tipo", mapa.get("fecha_1") == 1)
check("encuentra los kilos del tipo", mapa.get("kg_1") == 2)
check("la ficha no se roba la columna 'agujas'", mapa.get("marca") == 3
      and mapa.get("galga") == 5)

check("30.000 son treinta mil", excel.a_kilos("30.000 kg") == 30000)
check("lee fecha dd/mm/aaaa", excel.a_fecha("12/07/2026") == date(2026, 7, 12))
check("una fecha imposible no revienta", excel.a_fecha("cuando se pueda") is None)
check("MQ 003 es la 3", excel.a_numero("MQ 003") == 3)

filas_excel = [
    ["MQ 1", "12/07/2026", "30.000", "Mayer", "MV4", "24", "2015"],
    ["3",    None,          "5.000",  None,    None,  None, None],   # sin fecha
    ["99",   "01/01/2026", "10.000", None,    None,  None, None],    # no existe
    [None,   None,          None,     None,    None,  None, None],   # vacía
]
listas, descartes = excel.armar(titulos, filas_excel, mapa, MAQS, TIPOS, hoy=date(2026, 8, 19))
check("entran las dos que se entienden", len(listas) == 2)
check("la maquina que no esta en Asinfo queda afuera",
      any("no está en Asinfo" in d["motivo"] for d in descartes))
check("la fila vacia queda afuera", len(descartes) == 2)
check("sin fecha guarda el tope pero no inventa un mantenimiento",
      listas[1]["mantenimientos"][0]["fecha"] is None
      and listas[1]["mantenimientos"][0]["cada_kg"] == 5000
      and any("sólo se guarda el tope" in a for a in listas[1]["avisos"]))
check("lee la ficha", listas[0]["ficha"]["marca"] == "Mayer"
      and listas[0]["ficha"]["galga"] == 24)

futuras, _ = excel.armar(titulos, [["MQ 1", "31/12/2030", "1000", None, None, None, None]],
                         mapa, MAQS, TIPOS, hoy=date(2026, 8, 19))
check("una fecha futura no se carga",
      any("futura" in a for a in futuras[0]["avisos"]))

repetida, _ = excel.armar(titulos,
    [["MQ 1", "12/07/2026", "1000", None, None, None, None],
     ["MQ 1", "13/07/2026", "2000", None, None, None, None]],
    mapa, MAQS, TIPOS, hoy=date(2026, 8, 19))
check("avisa la maquina repetida",
      all(any("más de una vez" in a for a in r["avisos"]) for r in repetida))

# --- 6b. la planilla de planta: una hoja por maquina ----------------------
print("Planilla por maquina:")
from openpyxl import Workbook as _WB
wb2 = _WB()
h1 = wb2.active; h1.title = "MAQ 1"
h1.append(["R01-Registro de Mantenimiento"])
h1.append([])
h1.append([])
h1.append(["Equipo:", "", "Cantidad de agujas", "MODELO", "Responsable", "NUMERO"])
h1.append(["CIRCULAR", "MAYER", 2460, "Relanit 3.2 HS", "", 73830])
h1.append(["17////1", "", "", "D 32   G24"])
h1.append(["", "", "", "", "ANGEL PONCE"])
h1.append(["Año de fabricación:", "", "", 2017])
h1.append(["Fecha", "", "Fecha", "Tipo de mantenimiento", "Actividad realizada"])
h1.append([datetime(2024, 5, 3), "", "", "limpiesa de cilindro", ""])
h1.append([datetime(2026, 7, 17), "", "", "limpieza de cilindro", ""])
h1.append([datetime(2026, 2, 1), "", "", "cambio de agujas del cilindro", ""])
# Dos fechas que no pueden ser un mantenimiento: una todavía no pasó y la otra
# es de antes de que existiera la planilla. En planta se escriben solas.
h1.append([datetime(2030, 1, 1), "", "", "limpieza de cilindro", ""])
h1.append([datetime(1980, 6, 1), "", "", "limpieza de cilindro", ""])
h2 = wb2.create_sheet("MQ 2"); h2.append(["Registro"])   # hoja vacia
h3 = wb2.create_sheet("MAQ.3")
h3.append(["Equipo:", "", "AGUJAS", "MODELO", "DIAMETRO/ GG", "NUMERO"])
h3.append(["CIRCULAR", "JIUNN LONG", 3168, "JLD-T", "36  /  28", 230502])
h3.append(["", "", "100 ALIMENTADORES", "", "", ""])
h3.append(["Fecha", "Tipo de mantenimiento"])
h3.append([datetime(2026, 3, 5), "limpiesa de memminger"])
h4 = wb2.create_sheet("MQ 99")                            # no esta en Asinfo
h4.append(["Fecha", "x"]); h4.append([datetime(2026, 1, 1), "limpieza"])
import io as _io
buf2 = _io.BytesIO(); wb2.save(buf2); buf2.seek(0)
import tempfile as _tmp, os as _os
ruta2 = _os.path.join(_tmp.gettempdir(), "planilla_test.xlsx")
open(ruta2, "wb").write(buf2.getvalue())

TIPOS2 = [{"id": 1, "nombre": "Cambio de agujas", "cada_kg": None, "activo": True},
          {"id": 2, "nombre": "Limpieza", "cada_kg": None, "activo": True}]
pm, pd_ = excel.leer_por_maquina(ruta2, MAQS, TIPOS2, hoy=date(2026, 8, 20))
porn = {i["maquina"]["numero"]: i for i in pm}
check("lee una hoja por maquina", len(pm) == 2)
f1 = porn[1]["ficha"]
check("marca de al lado de CIRCULAR", f1["marca"] == "MAYER")
check("modelo", f1["modelo"] == "Relanit 3.2 HS")
check("diametro y galga de 'D 32  G24'", (f1["diametro"], f1["galga"]) == (32, 24))
check("agujas", f1["agujas"] == 2460)
check("anio", f1["anio"] == 2017)
check("numero de serie", str(f1["serie"]) == "73830")
m1 = {x["tipo"]["nombre"]: x["fecha"] for x in porn[1]["mantenimientos"]}
check("ultima limpieza", m1["Limpieza"] == date(2026, 7, 17))
check("el cambio de agujas se separa por el texto",
      m1["Cambio de agujas"] == date(2026, 2, 1))
f3 = porn[3]["ficha"]
check("diametro y galga de '36 / 28'", (f3["diametro"], f3["galga"]) == (36, 28))
check("alimentadores", f3["alimentadores"] == 100)
check("avisa el mantenimiento viejo",
      any("meses" in a for a in porn[3]["avisos"]))
check("la hoja vacia queda afuera con motivo",
      any(d["fila"] == "MQ 2" and "vacía" in d["motivo"] for d in pd_))
check("la maquina que no esta en Asinfo queda afuera",
      any("99" in str(d["texto"]) for d in pd_))

# --- 6c. el historial COMPLETO de la planilla de planta --------------------
# `leer_por_maquina` se queda con la última fecha de cada tipo, que es lo único
# que necesita el semáforo. La planilla tiene el historial entero desde 2018 y
# ése es su valor: poder mirar qué se le hizo a una máquina y cuándo.
print("Historial completo:")
hm, hf, hd = excel.leer_historial_por_maquina(ruta2, MAQS, TIPOS2, hoy=date(2026, 8, 20))
de_la_1 = [m for m in hm if m["id_maquina"] == 10]
check("trae TODAS las filas, no solo la ultima de cada tipo", len(de_la_1) == 3)
check("la fila que dice «aguja» es cambio de agujas",
      [m["fecha"] for m in de_la_1 if m["tipo_id"] == 1] == [date(2026, 2, 1)])
check("el resto queda como limpieza",
      sorted(m["fecha"] for m in de_la_1 if m["tipo_id"] == 2)
      == [date(2024, 5, 3), date(2026, 7, 17)])
check("una fecha futura no entra", all(m["fecha"] <= date(2026, 8, 20) for m in hm))
check("una fecha anterior a 1990 tampoco", all(m["fecha"].year >= 1990 for m in hm))
check("guarda que maquina es, con su nombre de Asinfo",
      all(m["maquina_nombre"] == "TEJEDURIA-MQ 001" for m in de_la_1))
# El responsable de la hoja es quién tiene la máquina a cargo, no quién hizo
# cada uno de los 1.300 mantenimientos. Copiarlo a cada fila sería inventar.
check("el responsable de la ficha NO se copia a cada fila",
      all(m["hecho_por"] == "Planilla de planta" for m in hm))
check("el responsable queda en la ficha, que es donde esta escrito",
      any("ANGEL PONCE" in (f.get("nota") or "") for _, f in hf))
hm2, _, _ = excel.leer_historial_por_maquina(ruta2, MAQS, TIPOS2, hoy=date(2026, 8, 20))
check("leerla dos veces da las mismas claves (hoja, orden)",
      [(m["hoja"], m["orden"]) for m in hm] == [(m["hoja"], m["orden"]) for m in hm2])
check("la clave no se repite dentro de la planilla",
      len({(m["hoja"], m["orden"]) for m in hm}) == len(hm))
check("la maquina que no esta en Asinfo queda afuera con motivo",
      any("99" in d["motivo"] for d in hd))
# Una hoja SIN la fila de titulos ya no se da por vacia: seis hojas de la
# planilla real arrancan directo con el historial, y asi se perdian enteras —
# la MQ 22 tiene 22 mantenimientos desde 2019.
check("la hoja sin fecha ninguna queda afuera con motivo",
      any(d["donde"] == "MQ 2" and "fecha" in d["motivo"] for d in hd))

# --- 6d. la ficha entera: los codigos de aguja y la ficha de al lado -------
# «TIPO DE AGUJAS» no trae un codigo: trae los cuatro del cilindro y el del
# plato, uno por renglon. Y seis hojas tienen DOS fichas pegadas, la de la
# maquina y la copia de la de al lado: la MQ 61 salia con el numero de serie de
# la otra, que es peor que no tener ninguno.
print("Ficha completa:")
wb3 = _WB()
g1 = wb3.active; g1.title = "MAQ 1"
g1.append(["R01-Registro de Mantenimiento"])
g1.append(["Equipo:", "", "Cantidad de agujas", "MODELO", "Responsable",
           "NUMERO", "TIPO DE AGUJAS"])
g1.append(["CIRCULAR", "MAYER", 2460, "Relanit 3.2 HS", "", 73830,
           "VO LS 140.50 G0036"])
g1.append(["", "", "", "D 32   G24", "", "", "VO LS 140.50 G0037"])
g1.append(["", "", "", "", "ANGEL PONCE", "", "VO LS 140.50 G0038"])
g1.append(["", "", "", "", "", "", "VO LS 140.50 G0039"])
g1.append(["Año de fabricación:", "", "", 2017, "", "", "platina 206001 681k 00"])
g1.append(["Fecha", "Tipo de mantenimiento"])
g1.append([datetime(2026, 7, 17), "limpieza de cilindro"])

# Dos fichas pegadas. A la de la izquierda —la de esta maquina— le borraron los
# titulos: los unicos que quedaron son los de la copia, ocho columnas a la
# derecha. Los datos son los de la izquierda; los titulos, los de la derecha.
g2 = wb3.create_sheet("MAQ 2")
g2.append(["", "", "R01 REGISTRO DE MANTENIMIENTO"])
g2.append(["Equipo:", "", "", "", "", "", "TIPO DE AGUJAS"])
g2.append(["CIRCULAR", "MAYER", 2640, "OV 3.2 QC", "", 72345, "plato VOTA 65.41 G004",
           "", "Equipo:", "", "Cantidad de agujas", "MODELO", "Responsable",
           "NUMERO", "TIPO DE AGUJAS"])
g2.append(["", "", "", "", "", "", "plato VO 65.41 G007",
           "", "CIRCULAR", "JIUNN LONG", "", "JLD-T", "", 230502, "plato VOTA 65.41 G004"])
g2.append(["2", "", "", "D 30   G28", "", "", "cilindro Vota 105,41 G005",
           "", "", "", 3168, "D 36   G28"])
g2.append(["", "", "96 ALIMENTADORES", "", "ANGEL PONCE"])
g2.append(["Año de fabricación:", "", "", 2013])
g2.append(["Fecha", "Tipo de mantenimiento"])
g2.append([datetime(2026, 3, 10), "limpieza de cilindro"])

# Una sola ficha, pero con «CANTIDAD DE AGUJAS» escrito dos veces: pasa en dos
# hojas de verdad. Cortar ahi se llevaba puesto el numero de serie.
g3 = wb3.create_sheet("MAQ.3")
g3.append(["Equipo:", "", "Cantidad de agujas", "MODELO", "Responsable",
           "NUMERO", "TIPO DE AGUJAS"])
g3.append(["CIRCULAR", "MAYER", 1680, "FS 2,2", "", 75508, "VO 91,50 G0011"])
g3.append(["", "", "", "D 30   G18", "", "", "VOTA 62,50 G0011"])
g3.append(["", "", "", "", "", "CANTIDAD DE AGUJAS"])
g3.append(["Fecha", "Tipo de mantenimiento"])
g3.append([datetime(2026, 5, 4), "limpieza de cilindro"])

buf3 = _io.BytesIO(); wb3.save(buf3); buf3.seek(0)
ruta3 = _os.path.join(_tmp.gettempdir(), "planilla_fichas.xlsx")
open(ruta3, "wb").write(buf3.getvalue())
_, fichas3, _ = excel.leer_historial_por_maquina(ruta3, MAQS, TIPOS2,
                                                 hoy=date(2026, 8, 20))
fpn = {m["numero"]: f for m, f in fichas3}
check("guarda los cinco codigos de aguja, no el primero",
      fpn[1]["tipo_agujas"] == ("VO LS 140.50 G0036 · VO LS 140.50 G0037 · "
                                "VO LS 140.50 G0038 · VO LS 140.50 G0039 · "
                                "platina 206001 681k 00"))
check("la serie no se lee de la ficha de al lado", str(fpn[2]["serie"]) == "72345")
check("ni las agujas", fpn[2]["agujas"] == 2640)
check("ni el modelo", fpn[2]["modelo"] == "OV 3.2 QC")
check("de la ficha pegada se leen los codigos de la izquierda",
      fpn[2]["tipo_agujas"].startswith("plato VOTA 65.41 G004 · plato VO 65.41 G007"))
check("el diametro y la galga son los de esta maquina",
      (fpn[2]["diametro"], fpn[2]["galga"]) == (30, 28))
check("una etiqueta repetida en la MISMA ficha no la parte al medio",
      str(fpn[3]["serie"]) == "75508" and fpn[3]["agujas"] == 1680)
check("y sus codigos de aguja salen enteros",
      fpn[3]["tipo_agujas"] == "VO 91,50 G0011 · VOTA 62,50 G0011")

# --- 6e. una fila puede ser dos mantenimientos ----------------------------
# La MQ 22 tiene renglones que dicen «limpiesa de cilindro» y en la columna de
# al lado «cambio de cilindro a galga 28»: ese dia se hicieron DOS cosas.
# Guardar sólo la limpieza borraba el cambio de cilindro del historial.
print("Dos mantenimientos en un renglon:")
TIPOS4 = [{"id": 1, "nombre": "Limpieza", "cada_kg": None, "activo": True},
          {"id": 2, "nombre": "Cambio de agujas", "cada_kg": None, "activo": True},
          {"id": 3, "nombre": "Cambio de cilindro", "cada_kg": None, "activo": True},
          {"id": 4, "nombre": "Cambio de platinas", "cada_kg": None, "activo": True}]
esp4, limp4 = excel._clasificar(TIPOS4)
check("la limpieza es la que se llama limpieza", limp4["id"] == 1)
nombres = lambda t: [x["nombre"] for x in excel._tipos_del_texto(t, esp4, limp4)]
# La trampa: «limpiesa de cilindro» dice cilindro y NO es un cambio de cilindro.
check("limpiar el cilindro es una limpieza",
      nombres("limpiesa de cilindro") == ["Limpieza"])
check("limpieza y cambio de cilindro son dos",
      nombres("limpiesa de cilindro · cambio de cilindro a galga 28")
      == ["Limpieza", "Cambio de cilindro"])
check("el cambio de ci,indro mal tipeado tambien cuenta",
      "Cambio de cilindro" in nombres("limpiesa de cilindro · cambio de ci,indro G 24"))
check("las platinas van aparte",
      nombres("limpiesa de cilindro · cambio de platinas galga 28")
      == ["Limpieza", "Cambio de platinas"])
check("agujas y platinas el mismo dia son las dos",
      set(nombres("cambio de agujas y platinas"))
      == {"Cambio de agujas", "Cambio de platinas"})
check("una fila que no dice nada es una limpieza", nombres("") == ["Limpieza"])
# Los tipos los crea el mecanico en la pantalla de Tipos. Si todavia no estan,
# la planilla se sigue leyendo igual que antes: todo cae en limpieza.
esp2, limp2 = excel._clasificar(TIPOS2)
check("sin los tipos nuevos cargados, sigue siendo una limpieza",
      [x["nombre"] for x in
       excel._tipos_del_texto("limpiesa de cilindro · cambio de platinas", esp2, limp2)]
      == ["Limpieza"])

hm4, _, _ = excel.leer_historial_por_maquina(ruta2, MAQS, TIPOS4, hoy=date(2026, 8, 20))
check("la clave (hoja, orden) sigue sin repetirse",
      len({(m["hoja"], m["orden"]) for m in hm4}) == len(hm4))

# Relanit es una maquina Mayer y JUNNLOG es JIUNN LONG mal tipeado: la misma
# marca escrita de tres formas partia la planta en marcas que son la misma.
check("relanit es mayer", excel._marca_pareja("RELANIT") == "MAYER")
check("mayer mv4 es mayer", excel._marca_pareja("MAYER  MV4") == "MAYER")
check("junnlog es jiunn long", excel._marca_pareja("JUNNLOG") == "JIUNN LONG")
check("una marca bien escrita no se toca", excel._marca_pareja("PILOTELLI") == "PILOTELLI")

# --- 7. la carga entera, de punta a punta ---------------------------------
print("Carga del Excel:")
import io
from openpyxl import Workbook

wb = Workbook(); ws = wb.active
ws.append(["PLANILLA DE MANTENIMIENTO"])          # basura arriba, como en planta
ws.append([])
ws.append(titulos)
ws.append(["MQ 1", "12/07/2026", "30.000", "Mayer", "MV4", 24, 2015])
ws.append(["MQ 2", "01/06/2026", "40.000", "Terrot", "S296", 28, 2011])
buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)

guardado_lote = {}
store.cargar_lote = lambda s_, t_, f_: (
    guardado_lote.update(servicios=s_, topes=t_, fichas=f_),
    {"servicios": len(s_), "topes": len(t_), "fichas": len(f_)})[1]

r = c.post("/carga", data={"archivo": (buffer, "planta.xlsx")},
           content_type="multipart/form-data")
check("subir el Excel lleva a revisar", r.status_code == 302 and "/carga/" in r.headers["Location"])
token = r.headers["Location"].rsplit("/", 1)[-1]

r = c.get(f"/carga/{token}")
cuerpo = r.get_data(as_text=True)
check("la revision encuentra las dos maquinas", r.status_code == 200 and "MQ 1" in cuerpo and "MQ 2" in cuerpo)
check("no guarda nada todavia", "servicios" not in guardado_lote)

r = c.post(f"/carga/{token}", data={"boton": "confirmar", "hecho_por": "Mecánico"})
check("confirmar guarda y vuelve al semaforo", r.status_code == 302)
check("guarda los dos mantenimientos", len(guardado_lote.get("servicios", [])) == 2)
check("guarda los topes por maquina", len(guardado_lote.get("topes", [])) == 2)
check("guarda las dos fichas", len(guardado_lote.get("fichas", [])) == 2)
check("el tope quedo en kilos de verdad", guardado_lote["topes"][0][2] == 30000)

r = c.get(f"/carga/{token}")
check("el archivo se borra despues de guardar", r.status_code == 302)

# Pegar las filas tiene que terminar en la MISMA pantalla de revisión que
# subir el archivo: si fueran dos caminos, serían dos lugares donde fallar.
print("Pegar en vez de subir:")
pegado = "\t".join(titulos) + "\nMQ 1\t12/07/2026\t30.000\tMayer\tMV4\t24\t2015"
r = c.post("/carga", data={"pegado": pegado})
check("pegar lleva a la misma revision", r.status_code == 302 and "/carga/" in r.headers["Location"])
token_p = r.headers["Location"].rsplit("/", 1)[-1]
cuerpo = c.get(f"/carga/{token_p}").get_data(as_text=True)
check("lo pegado se lee igual que el archivo", "MQ 1" in cuerpo)
guardado_lote.clear()
c.post(f"/carga/{token_p}", data={"boton": "confirmar", "hecho_por": "Mecánico"})
check("lo pegado guarda el tope", guardado_lote.get("topes", [[None, None, None]])[0][2] == 30000)

r = c.post("/carga", data={"pegado": "solo una linea"}, follow_redirects=True)
check("una sola linea no pasa", "títulos" in r.get_data(as_text=True))
r = c.post("/carga", data={"pegado": "titulo\notro"}, follow_redirects=True)
check("sin separador no pasa", "dos columnas" in r.get_data(as_text=True))
r = c.post("/carga", data={}, follow_redirects=True)
check("sin archivo ni pegado avisa", "pegá las filas" in r.get_data(as_text=True))

# --- 7b. la planilla de planta entera, de punta a punta --------------------
# Una hoja por máquina se reconoce sola y va por otro camino que la tabla: no
# hay nada que mapear y se guarda el historial ENTERO, no la última fecha.
print("Carga del historial:")
wb3 = _WB()
HOJAS3 = (("MAQ 1", [(datetime(2026, 7, 17), "limpieza de cilindro"),
                     (datetime(2026, 2, 1), "cambio de agujas del cilindro"),
                     (datetime(2025, 3, 4), "limpiesa de memminger")]),
          ("MAQ 2", [(datetime(2026, 6, 1), "limpieza general")]),
          ("MAQ 3", [(datetime(2026, 5, 5), "limpieza de platos"),
                     (datetime(2025, 1, 9), "cambio de agujas")]))
for i, (nombre_hoja, filas_hoja) in enumerate(HOJAS3):
    hoja = wb3.active if i == 0 else wb3.create_sheet(nombre_hoja)
    hoja.title = nombre_hoja
    hoja.append(["Equipo:", "", "", "MODELO"])
    hoja.append(["CIRCULAR", "MAYER", "", "Relanit 3.2"])
    hoja.append(["Fecha", "Tipo de mantenimiento"])
    for cuando, texto in filas_hoja:
        hoja.append([cuando, texto])
buf3 = io.BytesIO(); wb3.save(buf3); buf3.seek(0)

store.tipos = lambda incluir_inactivos=False: TIPOS2
_hist = {}
store.guardar_historial = lambda filas: (
    _hist.update(filas=filas), {"mantenimientos": len(filas), "borrados": 2})[1]
store.cuantos_de_la_carga_vieja = lambda: 2

r = c.post("/carga", data={"archivo": (buf3, "planta.xlsx")},
           content_type="multipart/form-data")
check("la planilla de planta entra por la misma pantalla",
      r.status_code == 302 and "/carga/" in r.headers["Location"])
token_h = r.headers["Location"].rsplit("/", 1)[-1]
cuerpo = c.get(f"/carga/{token_h}").get_data(as_text=True)
check("la revision resume por maquina, no fila por fila",
      "Planilla de mantenimiento" in cuerpo and "MQ 1" in cuerpo and "MQ 3" in cuerpo)
check("el boton dice cuantos mantenimientos son",
      "Guardar estos 6 mantenimientos" in cuerpo)
check("avisa que va a reemplazar lo de la primera carga", "primera carga" in cuerpo)
check("no guarda nada todavia", "filas" not in _hist)

r = c.post(f"/carga/{token_h}", data={"boton": "confirmar"})
check("confirmar guarda el historial entero",
      r.status_code == 302 and len(_hist.get("filas", [])) == 6)
check("guarda la clave que evita duplicar",
      all(f["hoja"] and f["orden"] for f in _hist["filas"]))
check("guarda tambien las fichas de las tres hojas",
      len(guardado_lote.get("fichas", [])) == 3)
check("el archivo se borra despues de guardar",
      c.get(f"/carga/{token_h}").status_code == 302)
store.tipos = lambda incluir_inactivos=False: TIPOS

# --- 8. arranque en lote ---------------------------------------------------
print("Arranque:")
store.ultimos_por_maquina_y_tipo = lambda: {}
capt = {}
store.registrar_muchos = lambda f: (capt.update(f=f), len(f))[1]
r = c.post("/arranque", data={"tipo_id": "1", "fecha": hoy.isoformat(), "hecho_por": "x"})
check("carga las 3 de una vez", r.status_code == 302 and len(capt["f"]) == 3)
r = c.post("/arranque", data={"tipo_id": "1", "fecha": (hoy + timedelta(days=1)).isoformat(), "hecho_por": "x"})
check("rechaza fecha futura", "futura" in r.get_data(as_text=True))
store.ultimos_por_maquina_y_tipo = lambda: {(m["id"], 1): {"fecha": hoy, "hecho_por": "x"} for m in MAQS}
antes = len(capt["f"])
c.post("/arranque", data={"tipo_id": "1", "fecha": hoy.isoformat(), "hecho_por": "x"})
check("no duplica si ya arrancaron", len(capt["f"]) == antes)

# --- 8b. cargar uno escribiendo el numero ---------------------------------
print("Cargar uno:")
guardado_uno = {}
store.registrar_service = lambda *a, **k: guardado_uno.update(args=a, kw=k)
for escrito, esperado in (("1", 10), ("MQ 2", 11), ("003", 12)):
    guardado_uno.clear()
    r = c.post("/registrar", data={"maquina": escrito, "tipo_id": "1",
                                   "fecha": hoy.isoformat(), "hecho_por": "Luis",
                                   "horas": "2"})
    check(f"'{escrito}' es la maquina {esperado}",
          guardado_uno.get("args", [None])[0] == esperado)

# Cuánto llevó no es opcional: sin las horas no se sabe cuánto cuesta parar una
# máquina, que es media razón por la que esto se anota.
guardado_uno.clear()
r = c.post("/registrar", data={"maquina": "1", "tipo_id": "1",
                               "fecha": hoy.isoformat(), "hecho_por": "Luis"})
check("sin horas no guarda y lo dice",
      not guardado_uno and "cuánto llevó" in r.get_data(as_text=True))
r = c.post("/registrar", data={"maquina": "1", "tipo_id": "1", "horas": "un rato",
                               "fecha": hoy.isoformat(), "hecho_por": "Luis"})
check("las horas tienen que ser un numero", "tienen que ser un número" in r.get_data(as_text=True))
# repuestos, horas y los tres encargados
guardado_uno.clear()
c.post("/registrar", data={"maquina": "1", "tipo_id": "1", "fecha": hoy.isoformat(),
                           "hecho_por": "Roberto", "repuestos": "12 agujas",
                           "horas": "2,5", "nota": "se trabo el plato"})
args = guardado_uno.get("args", ())
kw = guardado_uno.get("kw", {})
check("guarda quien lo hizo", args[4] == "Roberto")
check("guarda los repuestos", kw.get("repuestos") == "12 agujas")
check("2,5 horas se entienden como 2.5", kw.get("horas") == 2.5)
r = c.post("/registrar", data={"maquina": "1", "tipo_id": "1", "fecha": hoy.isoformat(),
                               "hecho_por": "Roberto", "horas": "999"})
check("rechaza horas absurdas", "razonable" in r.get_data(as_text=True))
check("los tres encargados estan en la lista",
      tuple(store.ENCARGADOS) == ("Roberto", "Darlin", "Humberto"))

r = c.post("/registrar", data={"maquina": "77", "tipo_id": "1",
                               "fecha": hoy.isoformat(), "hecho_por": "Luis"})
check("una maquina que no existe se avisa", "No hay ninguna máquina 77" in r.get_data(as_text=True))
r = c.post("/registrar", data={"maquina": "", "tipo_id": "1",
                               "fecha": hoy.isoformat(), "hecho_por": "Luis"})
check("sin numero pide el numero", "Poné el número" in r.get_data(as_text=True))

# --- 8c. archivos ----------------------------------------------------------
print("Archivos:")
guardados = {}
store.guardar_archivo = lambda nombre, contenido, **k: (
    guardados.update(nombre=nombre, contenido=contenido, extra=k), 1)[1]
store.archivos = lambda id_maquina=None: [
    {"id": 1, "id_maquina": 10, "nombre": "planilla.xlsx", "descripcion": "la de planta",
     "tamano": 2048, "subido_por": "Roberto", "creado_en": datetime(2026, 8, 20)}]
store.archivo = lambda i: {"id": 1, "id_maquina": 10, "nombre": "planilla.xlsx",
                           "contenido": b"PK\x03\x04hola", "descripcion": None,
                           "tamano": 9, "subido_por": None, "creado_en": None}
store.borrar_archivo = lambda i: guardados.update(borrado=i)

r = c.post("/archivos", data={"archivo": (io.BytesIO(b"unos bytes"), "manual.pdf"),
                              "id_maquina": "3", "descripcion": "manual",
                              "subido_por": "Darlin"},
           content_type="multipart/form-data")
check("sube un archivo", r.status_code == 302 and guardados.get("nombre") == "manual.pdf")
check("lo liga a la maquina por el numero", guardados["extra"]["id_maquina"] == 12)
r = c.post("/archivos", data={"descripcion": "sin archivo"},
           content_type="multipart/form-data")
check("sin archivo avisa", "Elegí un archivo" in r.get_data(as_text=True))
r = c.get("/archivos/1")
check("se puede bajar", r.status_code == 200 and b"hola" in r.data)
check("baja con su nombre", "planilla.xlsx" in r.headers.get("Content-Disposition", ""))
r = c.post("/archivos/1/borrar")
check("se puede borrar", guardados.get("borrado") == 1)

# --- 8d. los kilos se cargan en la ficha de cada maquina -------------------
# Antes habia una pantalla aparte con las 43 juntas. El numero es de la
# maquina, asi que se carga mirandola a ella: un solo lugar, uno que no puede
# quedar viejo respecto del otro.
print("Kilos en la ficha:")
puestos = {}
store.guardar_tope = lambda m, t, kg: puestos.update({(m, t): kg})
store.guardar_ficha = lambda i, d: None
# Al rechazar un número, la ficha se vuelve a dibujar entera: hay que tener
# contestado todo lo que pide.
store.ajustes = lambda id_maquina=None, tela=None, limite=400: []
store.agujas = lambda: {}
store.eficiencias = lambda: {}
store.archivos = lambda id_maquina=None: []
store.ficha = lambda id_maquina: {c: None for c in store.CAMPOS_FICHA}
r = c.post("/maquina/10", data={"marca": "Mayer", "tope_1": "250.000"})
check("guarda el tope de esa maquina", r.status_code == 302 and puestos.get((10, 1)) == 250000)
puestos.clear()
c.post("/maquina/10", data={"marca": "Mayer", "tope_1": ""})
check("el vacio borra el numero propio", puestos.get((10, 1)) is None)
r = c.post("/maquina/10", data={"marca": "Mayer", "tope_1": "-5"},
           follow_redirects=True)
check("rechaza un numero que no es kilos", "mayores que cero" in r.get_data(as_text=True))

# El listado tiene que decir a que maquina le falta algo: es la pantalla desde
# la que se recorren las 43 completando de a una.
store.fichas = lambda: {10: {"marca": "Mayer", "modelo": "Relanit", "galga": 24,
                             "diametro": 32, "alimentadores": 96, "agujas": 2460,
                             "anio": 2017, "serie": "7", "tipo_agujas": None,
                             "nota": None}}
# El tope va por máquina: sólo la 10 lo tiene puesto, las otras dos no.
TIPOS[0]["cada_kg"] = None
store.topes_por_maquina = lambda: {(10, 1): 50000.0}
r = c.get("/maquinas")
cuerpo = r.get_data(as_text=True)
check("el listado muestra el tope de kilos de cada maquina", "Tope kg" in cuerpo)
check("y el numero de la que lo tiene", "50.000" in cuerpo)
check("la que no lo tiene puesto dice que falta", ">falta<" in cuerpo)
TIPOS[0]["cada_kg"] = 50000

# --- 8d-bis. las fechas tipeadas a mano -----------------------------------
# En la planilla hay doce fechas escritas en una celda de texto, con los dos
# ordenes mezclados. Se acepta sólo lo que no da lugar a duda: elegir mal mueve
# el mantenimiento de mes, y con eso se corren los kilos.
print("Fechas escritas a mano:")
check("dia mayor que 12 al final se entiende",
      excel.fecha_escrita("01/19/2021", date(2026, 8, 20)) == date(2021, 1, 19))
check("dia mayor que 12 al principio tambien",
      excel.fecha_escrita("23/12,/2018", date(2026, 8, 20)) == date(2018, 12, 23))
check("con los dos numeros chicos y sin vecinas NO se adivina",
      excel.fecha_escrita("6/8/2018", date(2026, 8, 20)) is None)
# El historial está en orden: la fecha tiene que caer entre la de arriba y la
# de abajo. Con eso, de las dos lecturas posibles sobrevive una sola.
check("las fechas de arriba y abajo resuelven cual es",
      excel.fecha_escrita("6/8/2018", date(2026, 8, 20),
                          antes=date(2018, 7, 1), despues=date(2018, 9, 1))
      == date(2018, 8, 6))
check("y si las dos siguen entrando, no se elige",
      excel.fecha_escrita("6/8/2018", date(2026, 8, 20),
                          antes=date(2018, 1, 1), despues=date(2018, 12, 1)) is None)
check("un anio cortado lo completan las vecinas",
      excel.fecha_escrita("21/4/205", date(2026, 8, 20),
                          antes=date(2015, 1, 1), despues=date(2015, 12, 1))
      == date(2015, 4, 21))
check("un anio cortado sin vecinas no se completa",
      excel.fecha_escrita("21/4/205", date(2026, 8, 20)) is None)
check("una fecha futura no entra",
      excel.fecha_escrita("10/30/2026", date(2026, 8, 20)) is None)
check("lo que no es una fecha no molesta",
      excel.fecha_escrita("limpiesa de cilindro") is None)

# --- 8e. la planilla de control de ajuste ---------------------------------
print("Planilla de control de ajuste:")
import ajustes as AJ

MAQS_AJ = [{"id": 10, "numero": 1, "nombre": "TEJEDURIA-MQ 001"},
           {"id": 11, "numero": 52, "nombre": "TEJEDURIA-MQ 052"}]


def _planilla_ajuste(ruta):
    """Una planilla como la de planta, con sus tres mañas adentro."""
    from openpyxl import Workbook
    wb = Workbook()

    # Hoja normal: tres columnas «Polea» y tres «HILO» que se llaman igual.
    ws = wb.active; ws.title = "MAQ 1"
    ws.append(["CONTROL DE AGUSTES Y RENDIMIENTOS"])
    ws.append(["MQ. 1", "Tipo de  MQ.", "FECHA", "SERIE", "Polea", "Polea", "Polea",
               "ajuste agujas ", "ESTIRAJE", "TIPO DE TELA", "HILO", "HILO", "HILO",
               "G /m2   crudo", "Longitud de Malla manual", "Longitud de Malla",
               "Rendimiento Crudo"])
    ws.append([1, "MAYER 32", datetime(2026, 3, 4), "pro 30", "polea 5", "polea 20",
               None, "lycra pro 6,5", "B/9", "FALSO F. KW", "20/1 KW", "16/1 EP",
               None, 181.5, "30,2 dibujo", "29,0 LM", 4.05])
    # Una fila que sólo tiene el número y el modelo: no es un ajuste.
    ws.append([1, "MAYER 32", None, None, None, None, None, None, None, None])

    # Hoja a la que se le perdieron las primeras columnas: hay que leerla por
    # posición, y avisar.
    ws2 = wb.create_sheet("MAQ 52")
    ws2.append(["CONTROL DE AGUSTES Y RENDIMIENTOS"])
    ws2.append(["Polea", "Polea", "ajuste agujas ", "ESTIRAJE", "TIPO DE TELA"])
    ws2.append([52, "JIUNN LONG", datetime(2024, 5, 15), "pro 25", "dibujo 25",
                "pro 32", "jersey 50", "Z30", "B/8", "JAMES", "75/72"])

    # La hoja vieja, de cuando estaba todo junto: la primera fila ya está en la
    # hoja MAQ 1 y la segunda no está en ninguna. El número de máquina se
    # escribe una vez y después se arrastra, que es como está escrita la hoja.
    ws8 = wb.create_sheet("AGUSTES")
    ws8.append(["CONTROL DE AGUSTES Y RENDIMIENTOS"])
    ws8.append(["MQ.", "Tipo de  MQ.", "FECHA", "SERIE", "Polea", "Polea",
                "ajuste agujas ", "ESTIRAJE", "TIPO DE TELA", "HILO", "HILO"])
    ws8.append([1, "MAYER 32", datetime(2026, 3, 4), "pro 30", "polea 5",
                "polea 20", "lycra pro 6,5", "B/9", "FALSO F. KW",
                "20/1 KW", "16/1 EP"])
    ws8.append([None, "MAYER 32", datetime(2019, 4, 10), "pro 28", "polea 4",
                None, None, "B/7", "PIQUE", "30/1 KW", None])

    ws3 = wb.create_sheet("AGUJAS")
    ws3.append(["TIPO DE AGUJAS POR MAQUINA"])
    ws3.append(["MQ", "CILINDRO", None, "PLATO", "MAQUINA", "PLATINAS"])
    ws3.append([1, "VO LS-140,50 G36", "VO LS-140,50 G37", None, "MAYER", "206085101G00"])

    ws9 = wb.create_sheet("CODIGO DE AGUJAS")
    ws9.append(["MAQUINA", "MARCA", "CODIGO", "CODIGO", "CILINDRO", "PLATINAS",
                "MARCA"])
    ws9.append(["MAYER)1-2-3", "GROZ", "LS-140.50", "LS-141", "cilindro",
                "206085101G00", "GROZ"])

    # Dos tablas al lado de la otra: las bandas Memminger a la izquierda y las
    # de motor a la derecha, pasando la columna del stock.
    ws4 = wb.create_sheet("BANDAS")
    ws4.append([None])
    ws4.append(["MAQUNA", "cantidad de maquinas", "DIAMETRO", "BANDA  1/2",
                "BANDA 3/4", "BANDA LYCRA", None, None, None, None,
                "CODIGO", "CANTIDAD",
                "MAQUINA", "CANTIDAD", "DIAMETRO", "BANDA", "COBRADOR"])
    ws4.append(["MAYER JERSEY", 2, 30, "6.6", "8.2", None, None, None, None, None,
                6.6, 10,
                "MAYER 30", 2, 30, "M-45", "SI", "la corta el mecanico"])
    ws4.append([None, None, None, None, None, None, None, None, None, None,
                "total", 146])

    ws5 = wb.create_sheet("INVENTARIO LEVAS")
    ws5.append(["INVENTARIO DE LEVAS"])
    ws5.append(["MAQUINA ", "CODIGO", "CANTIDAD", "UBICACION", "ACCIONAMIENTO"])
    ws5.append(["MAYER (1 )", "30-32 385953,0", 208, "cilindro", "TRABAJO"])
    ws5.append(["MAYER (4 5 7 )", "NO", 0, "cilindro", "TRABAJO"])

    ws6 = wb.create_sheet("consumo de hilo")
    ws6.append(["TELA", "HILO", "HILO", "HILO", None, None])
    ws6.append(["JAMES", "poliester 75/36f", None, None, 0.4786, "75F36"])

    ws7 = wb.create_sheet("Eficiencia producción ")
    ws7.append([None])
    ws7.append([None, "Máquina", "Velocidad  (rpm)", "Sistema", "Diámetro (inch)",
                "F", "Tamaño de rollo", "Tiempo (min)", "N° Rollos diarios ",
                "Aproximación ", "Peso (kg)"])
    ws7.append([None, 1, 25, "102", 32, 24, 1410, 56.4, 12.76, 12, 270])

    wb.save(ruta)


_ruta_aj = os.path.join(tempfile.gettempdir(), "test_ajuste.xlsx")
_planilla_ajuste(_ruta_aj)

check("reconoce la planilla de ajuste", AJ.es_planilla_ajuste(_ruta_aj))
bloq, desc = AJ.leer(_ruta_aj, MAQS_AJ, hoy=date(2026, 8, 20))

_a1 = [a for a in bloq["ajustes"] if a["hoja"] == "MAQ 1"]
check("lee un ajuste por fila", len(_a1) == 1)
check("la fila sin nada útil no entra", all(a["tela"] for a in _a1))
check("las poleas van juntas en una columna",
      _a1[0]["poleas"] == "polea 5 · polea 20")
check("los hilos van juntos en una columna", _a1[0]["hilos"] == "20/1 KW · 16/1 EP")
check("distingue malla manual de malla",
      _a1[0]["malla_manual"] == "30,2 dibujo" and _a1[0]["malla"] == "29,0 LM")
check("el gramaje entra como número", _a1[0]["gramaje_crudo"] == 181.5)

_a52 = [a for a in bloq["ajustes"] if a["hoja"] == "MAQ 52"]
check("la hoja sin títulos se lee por posición igual", len(_a52) == 1)
check("leída por posición, la fecha cae en su lugar",
      _a52[0]["fecha"] == date(2024, 5, 15))
check("leída por posición, el modelo NO es la fecha",
      _a52[0]["tipo_maquina"] == "JIUNN LONG")
check("avisa que esa hoja se leyó por posición",
      any("por posición" in d["motivo"] for d in desc))

# La hoja vieja parece repetida y no lo es: casi todas sus filas no quedaron en
# ninguna hoja por máquina. Sólo entra lo que no está, y lo repetido se avisa
# para que el número cierre.
_viejos = [a for a in bloq["ajustes"] if a["hoja"] == "AGUSTES"]
check("de la hoja vieja entra sólo lo que no está por máquina", len(_viejos) == 1)
check("la fila que ya estaba en la hoja de la máquina no se duplica",
      _viejos[0]["fecha"] == date(2019, 4, 10))
check("avisa cuántas de la hoja vieja ya estaban",
      any(d["donde"] == "AGUSTES" and "ya estaban" in d["motivo"] for d in desc))
check("el número de máquina se arrastra de la fila de arriba",
      _viejos[0]["id_maquina"] == 10)

check("las agujas quedan pegadas a su máquina",
      len(bloq["agujas"]) == 1 and bloq["agujas"][0]["id_maquina"] == 10)
check("las cuatro columnas de cilindro van juntas",
      "G36 · VO LS-140,50 G37" in bloq["agujas"][0]["cilindro"])

# La aguja por MODELO va a su propia tabla: la fila es un modelo, no una
# máquina, y repartirla entre máquinas sería adivinar.
check("las agujas por modelo salen de su propia hoja",
      len(bloq["agujas_modelo"]) == 1
      and bloq["agujas_modelo"][0]["modelo"] == "MAYER)1-2-3")
check("junta los códigos de aguja del modelo",
      bloq["agujas_modelo"][0]["codigos"] == "LS-140.50 · LS-141")
check("guarda de qué marca es la aguja y de cuál la platina",
      bloq["agujas_modelo"][0]["marca_aguja"] == "GROZ"
      and bloq["agujas_modelo"][0]["marca_platina"] == "GROZ")

check("la leva sin código no entra", len(bloq["levas"]) == 1)
check("la leva guarda para qué sirve", bloq["levas"][0]["accionamiento"] == "TRABAJO")

check("el stock sale de la columna de al lado del código, no de «cantidad de "
      "maquinas»", bloq["banda_stock"] == [{"medida": 6.6, "cantidad": 10}])
check("la fila «total» no es una medida de banda",
      all(s["medida"] != 146 for s in bloq["banda_stock"]))
check("la banda guarda sus tres medidas",
      bloq["bandas"][0]["media"] == "6.6" and bloq["bandas"][0]["diametro"] == 30)

# Las bandas de motor son otro repuesto, con otras columnas: van con clase
# propia. Mezcladas con las Memminger, una medida taparía a la otra.
_memminger = [b for b in bloq["bandas"] if b["clase"] == "memminger"]
_motor = [b for b in bloq["bandas"] if b["clase"] == "motor"]
check("las bandas de motor quedan con clase «motor»",
      len(_motor) == 1 and _motor[0]["banda"] == "M-45")
check("y no se mezclan con las Memminger",
      len(_memminger) == 1 and _motor[0]["media"] is None)
check("la banda de motor guarda si lleva cobrador",
      _motor[0]["cobrador"] == "SI")

check("la producción calculada queda por máquina",
      len(bloq["eficiencia"]) == 1 and bloq["eficiencia"][0]["kg_dia"] == 270)
# El turno de 24 horas son los mismos títulos repetidos más a la derecha. Si no
# están, queda vacío: un número en la columna equivocada es peor que ninguno.
check("sin la tabla de 24 horas el turno largo queda vacío",
      bloq["eficiencia"][0]["kg_dia_24"] is None
      and bloq["eficiencia"][0]["rollos_dia_24"] is None)
check("el consumo de hilo trae el rendimiento",
      bloq["consumo_hilo"][0]["rendimiento"] == 0.4786)

# Volver a leer la misma planilla tiene que dar la misma clave (hoja, fila):
# es lo único que evita que cargarla dos veces duplique todo.
bloq2, _ = AJ.leer(_ruta_aj, MAQS_AJ, hoy=date(2026, 8, 20))
check("cargarla dos veces da las mismas claves",
      [(a["hoja"], a["orden"]) for a in bloq["ajustes"]]
      == [(a["hoja"], a["orden"]) for a in bloq2["ajustes"]])

check("una fecha futura no se carga",
      all(a["fecha"] is None or a["fecha"] <= date(2026, 8, 20)
          for a in bloq["ajustes"]))
check("la planilla de mantenimiento NO se confunde con esta",
      not AJ.es_planilla_ajuste(RUTA_EXCEL) if "RUTA_EXCEL" in dir() else True)

# La carga entera por la pantalla, que es como se hace de verdad.
_guardado = {}
store.guardar_planilla_ajuste = lambda d: (_guardado.update(d) or
                                           {k: len(v) for k, v in d.items()})
with open(_ruta_aj, "rb") as _f:
    _r = c.post("/carga", data={"archivo": (io.BytesIO(_f.read()), "ajuste.xlsx")},
                content_type="multipart/form-data")
check("la planilla de ajuste entra por la misma pantalla", _r.status_code == 302)
_token = _r.headers["Location"].rstrip("/").split("/")[-1]
_r = c.get(f"/carga/{_token}")
check("la revisión muestra qué entendió",
      "Planilla de control de ajuste" in _r.get_data(as_text=True))
_r = c.post(f"/carga/{_token}", data={"boton": "confirmar"})
check("confirmar guarda todo junto", _r.status_code == 302 and _guardado.get("ajustes"))

# --- 8e-bis. cargar un ajuste a mano, desde la pantalla -------------------
# Hasta ahora los ajustes sólo entraban por la planilla: una puesta a punto que
# se hizo hoy tenía que esperar a que alguien volviera a subir el Excel.
print("Cargar un ajuste:")
_ajuste = {}
store.crear_ajuste = lambda d: _ajuste.update(d)
store.ajustes = lambda id_maquina=None, tela=None, limite=400: []
store.telas = lambda: []
store.resumen_ajustes = lambda: {"filas": 0, "maquinas": 0, "con_fecha": 0,
                                 "desde": None, "hasta": None}
store.consumo_hilo = lambda: []
r = c.post("/ajustes", data={"maquina": "1", "tela": "PIQUE",
                             "fecha": hoy.isoformat(), "hilos": "20/1 KW",
                             "cilindro": "pro 24", "estiraje": "B/4",
                             "gramaje_crudo": "180,5"})
check("cargar un ajuste guarda y vuelve a la lista", r.status_code == 302)
check("lo guarda en la maquina que se eligio", _ajuste.get("id_maquina") == 10)
check("con el nombre que tiene en Asinfo",
      _ajuste.get("maquina_nombre") == "TEJEDURIA-MQ 001")
check("la tela y los ajustes quedan como se escribieron",
      _ajuste.get("tela") == "PIQUE" and _ajuste.get("cilindro") == "pro 24")
check("el gramaje con coma se entiende", _ajuste.get("gramaje_crudo") == 180.5)
check("lo que se dejo vacio queda vacio, no en blanco",
      _ajuste.get("poleas") is None)
_ajuste.clear()
r = c.post("/ajustes", data={"maquina": "1", "tela": ""})
check("sin tela no guarda y lo dice",
      not _ajuste and "tela" in c.get("/ajustes").get_data(as_text=True).lower())
r = c.post("/ajustes", data={"maquina": "1", "tela": "PIQUE",
                             "fecha": (hoy + timedelta(days=1)).isoformat()})
check("una fecha futura no entra", not _ajuste)
r = c.post("/ajustes", data={"maquina": "1", "tela": "PIQUE",
                             "gramaje_crudo": "un poco"})
check("un gramaje que no es numero no entra", not _ajuste)

# --- 8e-ter. la segunda tabla de INVENTARIO LEVAS -------------------------
# La hoja tiene DOS tablas al lado: el inventario de levas y, pegada a la
# derecha, cuántas levas lleva cada tela. Esa segunda eran cuarenta renglones
# que no entraban a ningún lado.
print("Levas por tela:")
from openpyxl import Workbook as _WB3
wb5 = _WB3()
h5 = wb5.active; h5.title = "INVENTARIO LEVAS"
h5.append([None] * 14 + ["CANTIDAD DE LEVAS POR TELA"])
h5.append(["MAQUINA", "CODIGO", "CANTIDAD", "UBICACION", "ACCIONAMIENTO"])
h5.append(["MAYER (1 )", "30-32 385953,0", 208, "cilindro", "TRABAJO",
           None, None, None, None, None, None, None, None, None,
           "JUNN LONG", "DIAMETRO 36", "alimentadores 108", "JERSEY",
           "levas de trabajo    432", "levas de retenido   0",
           "levas de anulacion  0"])
h5.append([None] * 14 + ["MAYER", "DIAMETRO 30", "alimentadores 96", "ROMA",
                         "levas de trabajo 192 cilindro",
                         "levas de retenido 16 cilindro",
                         "levas de anulacion 176 cilindro", "desprende mallas"])
buf5 = _io.BytesIO(); wb5.save(buf5); buf5.seek(0)
ruta5 = _os.path.join(_tmp.gettempdir(), "levas_tela.xlsx")
open(ruta5, "wb").write(buf5.getvalue())
from openpyxl import load_workbook as _lw
wb5b = _lw(ruta5, data_only=True)
lt, _ = AJ.leer_levas_tela(wb5b)
wb5b.close()
check("lee la tabla de la derecha", len(lt) == 2)
check("la tela es la tela", [x["tela"] for x in lt] == ["JERSEY", "ROMA"])
# «levas de trabajo 432» → «432». El texto que sigue al número queda: varias
# filas dicen «192 cilindro» y separar el número perdería dónde van.
check("saca la etiqueta y deja el numero", lt[0]["trabajo"] == "432")
check("y lo que aclara donde va, tambien queda",
      lt[1]["trabajo"] == "192 cilindro" and lt[1]["retenido"] == "16 cilindro")
check("lo que sobra a la derecha queda de nota", lt[1]["nota"] == "desprende mallas")
check("el inventario de la izquierda no se mezcla",
      all("30-32" not in str(x) for x in lt))

# --- 8e-quater. los numeros que venian con la unidad pegada ---------------
# En tres columnas —gramaje crudo, gramaje terminado y kg/m— casi nadie
# escribio el numero solo. Con `_decimal` entraba una celda de cada quince.
print("Numeros con la unidad pegada:")
for _txt, _esperado in (("1,80 kg/m", 1.8), ("2,48 kg/m", 2.48), ("4,40  *KG", 4.4),
                        ("3,37 KG /M", 3.37), ("2,86KG*M", 2.86), ("138 g", 138.0),
                        ("165 gr", 165.0), ("183G", 183.0), ("260 /gm2", 260.0),
                        (181.5, 181.5), ("180,5", 180.5)):
    check(f"«{_txt}» entra como {_esperado}", AJ._medida(_txt) == _esperado)
# La misma columna se uso para otra cosa: eso NO es una medida y no entra.
for _txt in ("30 LM dibujo", "28,2 LM", "LM 37,5 dibujo", "29,5 LM cilindro polea 1",
             "spander 108", "plegado", "V 955", "lycra 20/2", "GG 28", "PRUEBA",
             "guia aguja [4]", "1,20*3,20", "12/23/2019"):
    check(f"«{_txt}» no es una medida", AJ._medida(_txt) is None)
# El «G/m2» a secas es el TERMINADO: antes se lo comia gramaje_crudo.
_ti = AJ._mapa_de_titulos(
    ["MQ 28", "TIPO DE MQ", "FECHA", "RENDIMIENTO CRUDO", "G/m2 crudo", "G/m2", "KG/M"])
check("el crudo y el terminado son dos columnas",
      _ti["gramaje_crudo"] == [4] and _ti["gramaje_terminado"] == [5])

# --- 8e-quinquies. las bandas que hay que pedir ---------------------------
# El cuadro del fondo de la hoja BANDAS: el unico lugar de la planilla donde
# esta escrito que hay que comprar. No lo leia nadie.
print("Bandas para pedir:")
wb6 = _WB3()
h6 = wb6.active; h6.title = "BANDAS"
h6.append([None, "cantidad de maquinas", "DIAMETRO", "BANDA  1/2", "BANDA 3/4",
           "BANDA LYCRA", "CANTIDADA DE MAQUINAS CON ESA MEDIDA",
           "CANTIDAD REQUERIDA", None, "CODIGO", "CANTIDAD"])
h6[1][0].value = "MAQUNA"
h6.append(["MAYER JERSEY", 2, 30, "6.6", "8.2", None, "5 MQ", "4 de 6.600",
           "4 de 8,2000", 6.6, 10])
h6.append(["JIUNN LONG JERSEY", 1, 42, None, None, None, "1 MQ", "2 de 8,200",
           "2 de 9,600", 7.2, 20])
h6.append([])
h6.append(["cantidad", "codigo", None, "stok", "pedir", "metros"])
h6.append([10, 6.6, None, 5, 6, 39.6])
h6.append([20, 7.2, None, 12, 10, 72.0])
h6.append([30, 8.8, None, 7, 20, 176.0])
buf6 = _io.BytesIO(); wb6.save(buf6); buf6.seek(0)
ruta6 = _os.path.join(_tmp.gettempdir(), "bandas_pedido.xlsx")
open(ruta6, "wb").write(buf6.getvalue())
from openpyxl import load_workbook as _lw6
wb6b = _lw6(ruta6, data_only=True)
_bandas, _stock, _d = AJ.leer_bandas(wb6b)
_pedido, _dp = AJ.leer_bandas_pedido(wb6b)
wb6b.close()
check("lee el cuadro de pedido", len(_pedido) == 3)
check("con cuantas hay y cuantas pedir",
      _pedido[0] == {"medida": 6.6, "requeridas": 10, "stock": 5, "pedir": 6,
                     "metros": 39.6})
# «4 de 6.600» son cuatro bandas: el numero de atras es la medida.
check("cuantas bandas hacen falta, de «4 de 6.600»",
      _bandas[0]["requerida_media"] == 4 and _bandas[0]["requerida_tres_cuartos"] == 4)
check("y se guarda la frase entera, que en una fila se contradice",
      _bandas[0]["requerida_media_texto"] == "4 de 6.600")
# La de 42 pulgadas tiene las medidas vacias pero dice cuantas necesita: antes
# se caia por eso.
check("la banda sin medidas pero con cantidad requerida entra igual",
      any(b["diametro"] == 42 for b in _bandas))
check("cuantas maquinas usan esa medida", _bandas[0]["maquinas_con_medida"] == "5 MQ")

# --- 8e-sexies. las ocho telas que desaparecian ---------------------------
# De la mitad de la hoja para abajo el porcentaje se escribio adentro del
# nombre del hilo. El lector exigia la columna de rendimiento, asi que ocho
# telas no entraban NI salian en los descartes: desaparecian.
print("Consumo de hilo:")
wb7 = _WB3()
h7 = wb7.active; h7.title = "consumo de hilo"
h7.append(["TELA", "HILO", "HILO", "HILO", None, "rendimiento"])
h7.append(["JAMES", "poliester 75/36f", None, None, 0.4786, "75F36"])
h7.append([])
h7.append(["FALSO LYCRA", None, "HILO 22/1         71 %"])
h7.append([None, None, "LYCRA 20/1  29 %"])
h7.append([])
h7.append([None, None, "HILO 30/1 82%"])
h7.append(["TELA FLEECEC 200(galga 22)", None, "HILO 75F36 18%"])
buf7 = _io.BytesIO(); wb7.save(buf7); buf7.seek(0)
ruta7 = _os.path.join(_tmp.gettempdir(), "consumo.xlsx")
open(ruta7, "wb").write(buf7.getvalue())
wb7b = _lw6(ruta7, data_only=True)
_con, _dc = AJ.leer_consumo_hilo(wb7b)
wb7b.close()
_telas = {c["tela"] for c in _con}
check("entran las telas que tienen el porcentaje adentro del hilo",
      "FALSO LYCRA" in _telas)
# 71 % y 0,71 son la misma idea en dos unidades: se guarda como esta escrito.
_fl = next(c for c in _con if c["tela"] == "FALSO LYCRA")
check("el porcentaje va a su propia columna, sin convertir",
      _fl["porcentaje"] == 71.0 and _fl["rendimiento"] is None)
check("y el hilo queda sin el porcentaje pegado", _fl["hilo"] == "HILO 22/1")
# Una tela tiene el nombre escrito en la SEGUNDA linea de su bloque.
check("la tela nombrada al medio vale para todo el bloque",
      sum(1 for c in _con if c["tela"] == "TELA FLEECEC 200(galga 22)") == 2)

# --- 8e-septies. la planilla subida queda guardada ------------------------
# El archivo temporal se borra a las dos horas. Cada vez que el lector aprendia
# a leer una columna nueva habia que ir a buscar el Excel a la otra
# computadora: guardada, volver a leerla es un boton.
print("La planilla queda guardada:")
_guardados, _borrados = [], []
store.archivos = lambda id_maquina=None: [
    {"id": 7, "nombre": "CONTROL DE AJUSTE.xlsx", "descripcion": None,
     "tamano": 10, "subido_por": None, "creado_en": datetime.utcnow(),
     "id_maquina": None}]
store.guardar_archivo = lambda nombre, contenido, **k: (
    _guardados.append((nombre, len(contenido))), 9)[1]
store.borrar_archivo = lambda id_archivo: _borrados.append(id_archivo)
_wb8 = _WB3(); _wb8.active.append(["Fecha", "x"])
_buf8 = _io.BytesIO(); _wb8.save(_buf8); _buf8.seek(0)
c.post("/carga", data={"archivo": (_buf8, "CONTROL DE AJUSTE.xlsx")},
       content_type="multipart/form-data")
check("subir la planilla la deja guardada", len(_guardados) == 1)
check("con su nombre", _guardados[0][0] == "CONTROL DE AJUSTE.xlsx")
# La planilla es UNA y sigue viva en planta: la copia vieja se reemplaza, o
# quedarian veinte y ninguna forma de saber cual es la buena.
check("y reemplaza a la copia anterior", _borrados == [7])
_cuerpo = c.get("/carga").get_data(as_text=True)
check("la pantalla de subir muestra las guardadas",
      "CONTROL DE AJUSTE.xlsx" in _cuerpo and "Volver a leerla" in _cuerpo)
# Guardarla es una comodidad, no el trabajo: si la base falla, la carga sigue.
def _no_anda(*a, **k):
    raise RuntimeError("sin base")
store.guardar_archivo = _no_anda
_buf9 = _io.BytesIO(); _wb8.save(_buf9); _buf9.seek(0)
_r8 = c.post("/carga", data={"archivo": (_buf9, "otra.xlsx")},
             content_type="multipart/form-data")
check("si no se puede guardar, la carga sigue igual", _r8.status_code == 302)

# --- 8f. el historial de una máquina, agrupado por día --------------------
print("Los dias de una maquina:")
HISTORIAL = [
    {"fecha": date(2026, 8, 1), "tipo_nombre": "Limpieza", "hecho_por": "Roberto",
     "maquina_nombre": "TEJEDURIA-MQ 001", "nota": "se limpió el cilindro",
     "repuestos": None, "horas": None},
    {"fecha": date(2026, 8, 1), "tipo_nombre": "Cambio de agujas", "hecho_por": "Darlin",
     "maquina_nombre": "TEJEDURIA-MQ 001", "nota": None,
     "repuestos": "12 agujas", "horas": 2.5},
    {"fecha": date(2026, 5, 10), "tipo_nombre": "Limpieza", "hecho_por": "Roberto",
     "maquina_nombre": "TEJEDURIA-MQ 001", "nota": None,
     "repuestos": None, "horas": None},
]
# `kilos_desde` devuelve el acumulado desde cada fecha hasta hoy: los kilos
# entre dos paradas son la resta de los dos acumulados.
asinfo.kilos_desde = lambda id_maquina, fechas: {"2026-08-01": 30000.0,
                                                 "2026-05-10": 95000.0}
dias = A._dias_de_mantenimiento(10, HISTORIAL)
check("dos mantenimientos del mismo dia son un solo renglon", len(dias) == 2)
check("y el renglon lleva los dos tipos",
      dias[0]["tipos"] == ["Limpieza", "Cambio de agujas"])
check("junta a los dos que trabajaron ese dia", dias[0]["quien"] == "Darlin, Roberto")
check("junta las notas y los repuestos del dia",
      dias[0]["notas"] == ["se limpió el cilindro"]
      and dias[0]["repuestos"] == ["12 agujas"])
check("lo mas nuevo va primero", dias[0]["fecha"] == date(2026, 8, 1))
check("la ultima parada cuenta los kilos hasta hoy",
      dias[0]["kg"] == 30000.0 and dias[0]["hasta_hoy"] is True)
check("entre dos paradas, los kilos son la resta de los acumulados",
      dias[1]["kg"] == 65000.0)
# Cuánto estuvo parada la máquina ese día: si se hicieron dos cosas, la suma.
check("el tiempo del dia es la suma de lo que llevo cada cosa",
      dias[0]["horas"] == 2.5)
check("un dia sin horas anotadas no inventa un cero", dias[1]["horas"] is None)

def _kilos_caidos(*a, **k):
    raise asinfo.AsinfoNoDisponible("timeout")

asinfo.kilos_desde = _kilos_caidos
dias = A._dias_de_mantenimiento(10, HISTORIAL)
check("sin Asinfo se muestran los dias igual, sin kilos",
      len(dias) == 2 and all(d["kg"] is None for d in dias))
asinfo.kilos_desde = lambda id_maquina, fechas: {"2026-08-01": 30000.0,
                                                 "2026-05-10": 95000.0}

# --- 9. las pantallas abren -----------------------------------------------
print("Pantallas:")
store.historial = lambda id_maquina=None, limite=200: HISTORIAL
store.ficha = lambda id_maquina: {"marca": "Mayer", "modelo": "Relanit 3.2",
                                  "galga": 24, "diametro": 32, "alimentadores": 96,
                                  "agujas": 2460, "anio": 2017, "serie": "73830",
                                  "tipo_agujas": None, "nota": None}
store.agujas_por_modelo = lambda: [
    {"id": 1, "modelo": "MAYER)1-2-3", "marca_aguja": "GROZ",
     "codigos": "LS-140.50 · LS-141", "donde": "cilindro",
     "platinas": "206085101G00", "marca_platina": "GROZ", "nota": None}]
store.ajustes = lambda id_maquina=None, tela=None, limite=400: [
    {"id": 1, "id_maquina": 10, "maquina_nombre": "TEJEDURIA-MQ 001",
     "fecha": date(2026, 3, 4), "tipo_maquina": "MAYER 32", "cilindro": "pro 30",
     "poleas": "polea 5", "ajuste_agujas": None, "estiraje": "B/9",
     "tela": "FALSO F. KW", "hilos": "20/1 KW", "gramaje_crudo": 181.5,
     "malla_manual": None, "malla": "29,0 LM", "rendimiento": 4.05, "kg_m": None,
     "hoja": "MAQ 1", "orden": 1}]
store.telas = lambda: [{"tela": "FALSO F. KW", "veces": 40, "maquinas": 9,
                        "ultima": date(2026, 3, 4)}]
store.resumen_ajustes = lambda: {"filas": 1, "maquinas": 1, "con_fecha": 1,
                                 "desde": date(2026, 3, 4), "hasta": date(2026, 3, 4)}
store.agujas = lambda: {10: {"id_maquina": 10, "descripcion": "MAYER",
                             "cilindro": "VO LS-140,50", "plato": None,
                             "platinas": "206085101G00", "nota": None}}
store.levas = lambda: [{"id": 1, "maquinas": "MAYER (1 )", "codigo": "30-32",
                        "cantidad": 208, "ubicacion": "cilindro",
                        "accionamiento": "TRABAJO"}]
# Las bandas se piden por clase: las Memminger y las de motor son dos tablas
# distintas en la misma pantalla.
store.bandas = lambda clase="memminger": [
    {"id": 1, "maquinas": "MAYER JERSEY", "cantidad_maquinas": 2, "clase": clase,
     "diametro": 30, "media": "6.6", "tres_cuartos": "8.2", "lycra": None,
     "banda": "M-45", "cobrador": "SI", "nota": None}]
store.banda_stock = lambda: [{"medida": 6.6, "cantidad": 10}]
# Repuestos se muestra de a un cuadro por pestaña, y los seis salen del mismo
# lugar. Cada cuadro devuelve su fila de mentira.
FILAS_CUADRO = {
    "agujas": [{"id_maquina": 10, "descripcion": "MAYER", "cilindro": "VO LS-140,50",
                "plato": None, "platinas": "206085101G00", "nota": None}],
    "modelos": [{"id": 1, "modelo": "MAYER)1-2-3", "marca_aguja": "GROZ",
                 "codigos": "LS-140.50", "donde": "cilindro",
                 "platinas": "206085101G00", "marca_platina": "GROZ", "nota": None}],
    "levas": [{"id": 1, "maquinas": "MAYER (1 )", "codigo": "30-32", "cantidad": 208,
               "ubicacion": "cilindro", "accionamiento": "TRABAJO"}],
    "bandas": [{"id": 1, "maquinas": "MAYER JERSEY", "cantidad_maquinas": 2,
                "diametro": 30, "media": "6.6", "tres_cuartos": "8.2", "lycra": None,
                "banda": None, "cobrador": None, "nota": None}],
    "motor": [{"id": 2, "maquinas": "MAYER JERSEY", "cantidad_maquinas": 2,
               "diametro": 30, "media": None, "tres_cuartos": None, "lycra": None,
               "banda": "M-45", "cobrador": "SI", "nota": None}],
    "stock": [{"medida": 6.6, "cantidad": 10}],
    "pedido": [{"medida": 6.6, "requeridas": 10, "stock": 5, "pedir": 6,
                "metros": 39.6}],
    "levas_tela": [{"id": 1, "marca": "MAYER", "diametro": "DIAMETRO 34",
                    "alimentadores": "alimentadores 108", "tela": "JERSEY",
                    "trabajo": "432", "retenido": "0", "anulacion": "0",
                    "nota": None}],
}
store.filas_de = lambda cuadro: list(FILAS_CUADRO[cuadro])
# La 11 está anotada en la planilla pero sin números: pasa de verdad, y sumar
# un vacío rompía la pantalla.
store.eficiencias = lambda: {10: {"id_maquina": 10, "rpm": 25, "sistemas": "102",
                                  "diametro": 32, "alimentadores": 24,
                                  "tamano_rollo": 1410, "minutos_rollo": 56.4,
                                  "rollos_dia": 12, "kg_dia": 270,
                                  "rollos_dia_24": 24, "kg_dia_24": 540,
                                  "galga": 24, "real_rollos_dia": 9,
                                  "real_kg_dia": 202.5,
                                  "real_rollos_24": 17, "real_kg_24": 416.5},
                             11: {"id_maquina": 11, "rpm": None, "sistemas": "96",
                                  "diametro": 30, "alimentadores": 28,
                                  "tamano_rollo": None, "minutos_rollo": None,
                                  "rollos_dia": None, "kg_dia": None,
                                  "rollos_dia_24": None, "kg_dia_24": None}}
store.consumo_hilo = lambda: [{"id": 1, "tela": "JAMES", "hilo": "poliester",
                               "codigo_hilo": "75F36", "rendimiento": 0.4786}]
store.gramajes = lambda id_maquina=None: []
store.fichas = lambda: {10: {"marca": "Mayer", "modelo": "Relanit", "galga": 24,
                            "diametro": 32, "alimentadores": 96, "agujas": 2460,
                            "anio": 2017, "serie": "7", "tipo_agujas": None, "nota": None}}
for ruta in ("/", "/registrar", "/tipos", "/arranque", "/carga", "/maquina/10",
             "/maquinas", "/archivos", "/ajustes", "/ajustes?tela=FALSO",
             "/ajustes?maquina=1", "/repuestos", "/repuestos?ver=levas",
             "/repuestos?ver=motor", "/repuestos?ver=stock",
             "/repuestos?ver=levas&editar=1"):
    r = c.get(ruta)
    check(f"{ruta} abre", r.status_code == 200)
r = c.get("/?solo=vencidas")
check("la campanita filtra las vencidas", r.status_code == 200)
cuerpo = c.get("/maquinas").get_data(as_text=True)
check("la pestaña de maquinas las lista con su ficha",
      "MQ 1" in cuerpo and "MQ 3" in cuerpo and "Mayer" in cuerpo)

# El buscador de la ficha no es una pantalla: lleva a la máquina que se escribió.
r = c.get("/maquina?numero=1")
check("/maquina?numero=1 lleva a la ficha de la 1",
      r.status_code == 302 and r.headers["Location"].endswith("/maquina/10"))
r = c.get("/maquina?numero=77", follow_redirects=True)
# Un dato que falta se muestra como «—»: Jinja devuelve Undefined cuando no
# está la clave, y sin este freno una pantalla entera se caía con un 500.
check("un dato que falta no tira la pantalla abajo",
      A._num(A.jinja2.Undefined(name="x")) == "—"
      and A._fecha_es(A.jinja2.Undefined(name="x")) == "—"
      and A._num("no soy un numero") == "—")
check("una maquina que no esta en Asinfo vuelve al listado, no rompe",
      c.get("/maquina/999999").status_code == 302)

check("un numero que no existe vuelve al listado y lo dice",
      "No hay ninguna máquina 77" in r.get_data(as_text=True))
cuerpo = c.get("/maquina/10").get_data(as_text=True)
check("la ficha muestra los dias, no los registros sueltos",
      "Limpieza" in cuerpo and "Cambio de agujas" in cuerpo)


# --- 10. editar los repuestos a mano ---------------------------------------
print("Repuestos a mano:")
check("las columnas de la pantalla existen en la base",
      all(col["campo"] in store.CUADROS[cu["clave"]]["campos"]
          for cu in A.CUADROS_REPUESTOS for col in cu["columnas"]))

corridas = []
store._ejecutar = lambda sql, args=(): corridas.append((" ".join(sql.split()), args))

store.guardar_repuesto("levas", None, {"maquinas": "MAYER (5)", "codigo": "30-32",
                                       "cantidad": 12, "ubicacion": None,
                                       "accionamiento": None})
check("agregar una leva es un INSERT", corridas[-1][0].startswith("INSERT INTO mantenimiento.leva"))

store.guardar_repuesto("motor", None, {"maquinas": "MAYER", "diametro": 30})
check("la banda nueva se guarda en la clase de la pestaña",
      "clase" in corridas[-1][0] and "motor" in corridas[-1][1])

store.guardar_repuesto("levas", "7", {"maquinas": "MAYER (5)", "codigo": "30-33"})
check("editar una leva es un UPDATE por su id",
      corridas[-1][0].startswith("UPDATE mantenimiento.leva") and corridas[-1][1][-1] == "7")

# La clave es lo que identifica la fila: cambiarla no es corregir un dato.
store.guardar_repuesto("stock", 6.6, {"medida": 9.9, "cantidad": 4})
check("editando no se pisa la clave de la fila",
      corridas[-1][0] ==
      "UPDATE mantenimiento.banda_stock SET cantidad = %s WHERE medida = %s"
      and corridas[-1][1] == (4, 6.6))

# Cargar dos veces la misma máquina es corregir lo que había, no un error.
store.guardar_repuesto("agujas", None, {"id_maquina": 10, "descripcion": "MAYER"})
check("la aguja de una maquina que ya estaba se corrige, no se duplica",
      "ON CONFLICT (id_maquina) DO UPDATE" in corridas[-1][0])

try:
    store.guardar_repuesto("levas", None, {"maquinas": None, "codigo": None})
    vacia = "no avisó"
except ValueError as exc:
    vacia = str(exc)
check("la fila vacia no se agrega y lo dice", "vacía" in vacia)

store.borrar_repuesto("levas", "7")
check("borrar es un DELETE por su id",
      corridas[-1][0] == "DELETE FROM mantenimiento.leva WHERE id = %s")

cuerpo = c.get("/repuestos?ver=levas&cuadro=levas&editar=1").get_data(as_text=True)
check("el lapicito abre la fila como campos",
      'name="codigo"' in cuerpo and 'form="editar-levas"' in cuerpo)
check("siempre hay una fila en blanco para agregar", 'form="nueva-levas"' in cuerpo)
# Tres pestañas y no seis: de las seis hojas de la planilla, dos hablan de lo
# mismo (la aguja por máquina y la misma aguja por modelo, que es como se pide)
# y tres son bandas.
check("las pestañas son tres", len(A.PESTANAS_REPUESTOS) == 3)
check("y llevan a las tres",
      all(f"ver={p['clave']}" in cuerpo for p in A.PESTANAS_REPUESTOS))
check("cada cuadro vive en una pestaña",
      sorted(cu["clave"] for cu in A.CUADROS_REPUESTOS)
      == sorted(c2 for p in A.PESTANAS_REPUESTOS for c2 in p["cuadros"]))
# Las bandas son tres cuadros en UNA pantalla.
cuerpo = c.get("/repuestos?ver=bandas").get_data(as_text=True)
check("las tres tablas de bandas van juntas",
      all(x in cuerpo for x in ("Bandas Memminger", "Bandas de motor",
                                "Cuántas bandas hay en bodega")))
# Cuatro códigos de aguja pegados con un punto son un renglón ilegible.
check("los codigos de aguja van uno por renglon",
      'class="codigo"' in c.get("/repuestos?ver=agujas").get_data(as_text=True))

# Cuánto debería dar cada máquina tiene pantalla propia: estaba sólo adentro
# de la ficha, así que comparar dos máquinas era entrar y salir de las dos.
store.eficiencias = lambda: {10: {"id_maquina": 10, "rpm": 25, "sistemas": "102",
                                  "diametro": 32, "alimentadores": 24,
                                  "tamano_rollo": 1410, "minutos_rollo": 56.4,
                                  "rollos_dia": 12, "kg_dia": 270,
                                  "rollos_dia_24": 24, "kg_dia_24": 540,
                                  "galga": 24, "real_rollos_dia": 9,
                                  "real_kg_dia": 202.5,
                                  "real_rollos_24": 17, "real_kg_24": 416.5}}
store.gramajes = lambda id_maquina=None: [
    {"id": 1, "id_maquina": 10, "fecha": date(2021, 11, 10), "tela": "FALSO F",
     "hilos": "20/1 KW", "peso": 4.39, "orden": 1}]
cuerpo = c.get("/produccion").get_data(as_text=True)
check("produccion es una pantalla propia", "Cuánto debería dar cada máquina" in cuerpo)
check("con los kilos de cada maquina", "270" in cuerpo)
check("y las que no tienen el calculo se ven aparte",
      "Sin el cálculo" in cuerpo and "MQ 2" in cuerpo)
# El peso medido estaba cargado y no se veía en ninguna pantalla.
check("el peso medido de la tela tambien se ve", "4,39" in cuerpo)
# Lo que la máquina DIO, al lado de lo que debería dar. Estaba en la planilla
# desde el principio y no se traía.
check("y lo que dio de verdad, al lado", "202" in cuerpo and "Dio de verdad" in cuerpo)
check("con el rinde en porcentaje", "75%" in cuerpo)
check("produccion esta en el menu", "/produccion" in c.get("/").get_data(as_text=True))

# --- 11. la ficha: los ultimos 5 y cuanto deberia dar ----------------------
print("Ficha:")
store.historial = lambda id_maquina=None, limite=200: [
    {"fecha": date(2026, 8, 20) - timedelta(days=30 * i), "tipo_nombre": "Limpieza",
     "hecho_por": "Roberto", "nota": f"parada {i}", "repuestos": None, "horas": 1,
     "maquina_nombre": "TEJEDURIA-MQ 001"} for i in range(9)]
cuerpo = c.get("/maquina/10").get_data(as_text=True)
check("se muestran los 5 mas nuevos y el resto atras de la flecha",
      "parada 4" in cuerpo and "Ver los 4 anteriores" in cuerpo)

guardado = {}
store.guardar_eficiencia = lambda id_maquina, datos: guardado.update(
    {"id": id_maquina, **datos})
r = c.post("/maquina/10", data={"que": "eficiencia", "rpm": "25,0",
                                "sistemas": "102", "tamano_rollo": "1.410",
                                "minutos_rollo": "56,4", "rollos_dia": "12",
                                "kg_dia": "270", "rollos_dia_24": "25",
                                "kg_dia_24": "612"})
check("se puede editar cuanto deberia dar",
      r.status_code == 302 and guardado["rpm"] == 25 and guardado["tamano_rollo"] == 1410
      and guardado["kg_dia_24"] == 612)
check("el tamaño de rollo con punto son mil cuatrocientas diez vueltas",
      guardado["tamano_rollo"] == 1410)

# La cuenta de la planilla, con los números de la MQ 1 de la hoja de verdad:
# 1410 vueltas ÷ 25 rpm = 56,4 min; en 12 h entran 12 rollos de 22,5 kg = 270,
# y en 24 h, 25 rollos de 24,5 = 612,5. Si alguien cambia los pesos en la
# pantalla, esto se entera.
minutos = 1410 / 25
pantalla = open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                             "templates", "maquina.html")).read()
check("la cuenta de cuanto deberia dar es la de la planilla",
      minutos == 56.4 and int(720 // minutos) == 12 and int(1440 // minutos) == 25
      and int(720 // minutos) * 22.5 == 270 and int(1440 // minutos) * 24.5 == 612.5)
check("el peso del rollo es el de la planilla, no uno inventado",
      "PESO_12 = 22.5" in pantalla and "PESO_24 = 24.5" in pantalla)

# 270 no es 27: los ceros del final no se recortan.
check("el valor para escribir a mano no pierde los ceros",
      A._editable(270) == "270" and A._editable(22.5) == "22,5"
      and A._editable(None) == "")


# --- 14. un tope de kilos inventado no puede dejar la maquina en verde ------
# El tope se escribe a mano en la ficha. `float()` pelado acepta «nan» e
# «inf», y `nan <= 0` es FALSO, asi que los dos pasaban el control y quedaban
# guardados. Despues ninguna comparacion contra nan es verdadera, y los kilos
# divididos por infinito dan cero: esa maquina decia «En regla» para siempre,
# con los kilos ya pasados. Un falso verde es justo lo que este programa
# existe para evitar.
print("Un tope inventado:")
check("un casillero vacio no es un tope",
      A._kilos_escritos("", "Limpieza") is None
      and A._kilos_escritos(None, "Limpieza") is None)
check("20.000 son veinte mil", A._kilos_escritos("20.000", "Limpieza") == 20000)
check("20.000,5 se entiende con la coma",
      A._kilos_escritos("20.000,5", "Limpieza") == 20000.5)


def _rechaza(valor):
    try:
        A._kilos_escritos(valor, "Limpieza")
        return False
    except ValueError:
        return True


for malo in ("nan", "NaN", "inf", "-inf", "1e400", "-5", "-20.000", "0", "abc"):
    check(f"{malo!r} no es un tope de kilos", _rechaza(malo))
check("y el aviso se lee en castellano",
      "mayores que cero" in _mensaje_de(lambda: A._kilos_escritos("nan", "Limpieza")))
check("lo que no es un numero tambien lo dice",
      "no es un número de kilos"
      in _mensaje_de(lambda: A._kilos_escritos("abc", "Limpieza")))

# Por que hay que frenarlo en la puerta: una vez guardado ya no hay
# comparacion que lo agarre.
_nan = float("nan")
check("contra nan ninguna comparacion prende un color",
      not (_nan >= 1.0) and not (_nan >= A.AVISO))
check("y dividir por infinito da cero, que es verde",
      80000.0 / float("inf") == 0.0)

# Y por la pantalla: no se guarda nada y se avisa.
puestos.clear()
r = c.post("/maquina/10", data={"marca": "Mayer", "tope_1": "nan"},
           follow_redirects=True)
check("la ficha no guarda un tope nan",
      not puestos and "mayores que cero" in r.get_data(as_text=True))
puestos.clear()
c.post("/maquina/10", data={"marca": "Mayer", "tope_1": "1e400"}, follow_redirects=True)
check("ni uno infinito", not puestos)

# Los topes se revisan TODOS antes de escribir ninguno: uno malo no puede
# dejar la ficha con la mitad guardada.
_DOS_TIPOS = [{"id": 1, "nombre": "Limpieza", "cada_kg": None, "activo": True},
              {"id": 2, "nombre": "Cambio de agujas", "cada_kg": None, "activo": True}]
_tipos_antes = store.tipos
store.tipos = lambda incluir_inactivos=False: _DOS_TIPOS
puestos.clear()
c.post("/maquina/10", data={"marca": "Mayer", "tope_1": "20.000", "tope_2": "nan"},
       follow_redirects=True)
check("un tope malo no deja los otros a medio guardar", not puestos)
puestos.clear()
c.post("/maquina/10", data={"marca": "Mayer", "tope_1": "20.000", "tope_2": "8.000"},
       follow_redirects=True)
check("y los dos buenos entran juntos",
      puestos.get((10, 1)) == 20000 and puestos.get((10, 2)) == 8000)
store.tipos = _tipos_antes

# La pantalla de Tipos escribe en la misma columna y no controlaba nada.
tipos_creados = []
store.crear_tipo = lambda *a: tipos_creados.append(a)
for malo in ("-20.000", "nan", "inf"):
    tipos_creados.clear()
    c.post("/tipos", data={"nombre": "Limpieza", "cada_kg": malo}, follow_redirects=True)
    check(f"tipos tampoco acepta {malo!r}", not tipos_creados)
tipos_creados.clear()
c.post("/tipos", data={"nombre": "Limpieza", "cada_kg": "20.000"}, follow_redirects=True)
check("y un numero de verdad si", bool(tipos_creados) and tipos_creados[0][1] == 20000)

# --- 15. un solo numero de maquina -----------------------------------------
# Se tomaba el ULTIMO numero escrito: «12,5» cargaba el mantenimiento en la
# MQ 5 y «MQ 12 galga 28» en la 28, sin que nada lo dijera. El mantenimiento
# queda en la ficha de otra maquina y las dos cuentan mal los kilos.
print("Un solo numero de maquina:")
for escrito, esperado in (("1", 10), ("01", 10), ("MQ 2", 11),
                          ("TEJEDURIA-MQ 003", 12), (" 3 ", 12)):
    check(f"{escrito!r} es la maquina {esperado}",
          A._buscar_maquina(escrito, MAQS) == esperado)
for confuso in ("12,5", "MQ 12 galga 28", "1 2"):
    check(f"{confuso!r} no elige ninguna",
          "un solo número" in _mensaje_de(lambda: A._buscar_maquina(confuso, MAQS)))
check("sin numero pide el numero",
      "Poné el número" in _mensaje_de(lambda: A._buscar_maquina("", MAQS)))
check("un numero que no existe se dice con el numero puesto",
      "No hay ninguna máquina 77" in _mensaje_de(lambda: A._buscar_maquina("77", MAQS)))
guardado_uno.clear()
r = c.post("/registrar", data={"maquina": "12,5", "tipo_id": "1", "horas": "1",
                               "fecha": hoy.isoformat(), "hecho_por": "Luis"})
check("y por la pantalla tampoco carga en la equivocada",
      not guardado_uno and "un solo número" in r.get_data(as_text=True))

# --- 16. la contraseña ------------------------------------------------------
# /healthz tiene que quedar SIEMPRE abierto: es lo que mira el auto-updater
# del server para decidir si la version nueva levanto. Si pidiera contraseña,
# cada actualizacion se desharia sola.
print("La contraseña:")
config.PASSWORD = "la clave"
try:
    puerta = A.app.test_client()
    r = puerta.get("/")
    check("sin contraseña el semaforo manda al login",
          r.status_code == 302 and "/login" in r.headers["Location"])
    check("y se acuerda de a donde iba", "next=" in r.headers["Location"])
    check("healthz no pide contraseña nunca", puerta.get("/healthz").status_code == 200)
    r = puerta.post("/login", data={"password": "otra cosa"})
    check("una clave que no es no entra", "incorrecta" in r.get_data(as_text=True))
    check("y sigue sin dejar pasar", puerta.get("/").status_code == 302)
    r = puerta.post("/login?next=/maquinas", data={"password": "la clave"})
    check("con la clave entra y vuelve a donde iba",
          r.status_code == 302 and r.headers["Location"] == "/maquinas")
    check("y ahora si abre", puerta.get("/maquinas").status_code == 200)
    # Una direccion de afuera pegada en ?next= mandaria a la gente a otro sitio
    # justo despues de escribir la contraseña.
    for afuera in ("//otro.com", "https://otro.com", "http://otro.com/x"):
        with A.app.test_request_context(f"/login?next={afuera}"):
            check(f"{afuera} no es una pantalla de acá", A._adonde_iba() is None)
    with A.app.test_request_context("/login?next=/maquina/10"):
        check("una pantalla de acá si", A._adonde_iba() == "/maquina/10")
    r = puerta.get("/salir")
    check("salir cierra la sesion",
          r.status_code == 302 and puerta.get("/").status_code == 302)
finally:
    config.PASSWORD = ""

# --- 17. los templates se parsean ------------------------------------------
# `py_compile` no ve los .html. Un `{% endif %}` de menos, o un filtro con el
# nombre mal escrito, aparecen recien en runtime — con la pantalla ya rota en
# produccion. Esto lo hacia solo el workflow; aca corre ANTES de commitear,
# que es donde sirve.
print("Los templates:")
import jinja2.nodes

CARPETA_T = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                         "templates")
plantillas = sorted(f for f in os.listdir(CARPETA_T) if f.endswith(".html"))
check("estan todas las pantallas", len(plantillas) >= 15)
filtros_usados = set()
for nombre in plantillas:
    with open(os.path.join(CARPETA_T, nombre), encoding="utf-8") as f:
        fuente = f.read()
    try:
        arbol = A.app.jinja_env.parse(fuente, filename=nombre)
        filtros_usados |= {n.name for n in arbol.find_all(jinja2.nodes.Filter)}
        bien = True
    except jinja2.TemplateSyntaxError as exc:
        bien = False
        print(f"         {nombre}: {exc}")
    check(f"{nombre} se parsea", bien)
faltan = sorted(f for f in filtros_usados if f not in A.app.jinja_env.filters)
check("todos los filtros que usan existen" + (f" (falta {', '.join(faltan)})" if faltan else ""),
      not faltan)
check("y los nuestros estan puestos",
      {"num", "pct", "fecha_es", "editable", "mq"} <= set(A.app.jinja_env.filters))

# --- 18. el esquema, una sentencia por transaccion -------------------------
# Si van todas juntas y una falla, `init_pool` revienta, /healthz devuelve 503
# y el auto-update del server deshace el deploy entero sin decir cual fallo.
print("El esquema:")
partes = store._sentencias(store.ESQUEMA)
check("el esquema son muchas sentencias, no una", len(partes) > 20)
check("ninguna queda vacia", all(p.strip() for p in partes))
check("los comentarios no viajan adentro", not any("--" in p for p in partes))
check("el schema se crea una sola vez",
      sum(1 for p in partes if p.upper().startswith("CREATE SCHEMA")) == 1)
check("un punto y coma adentro de un comentario no parte la sentencia",
      store._sentencias("SELECT 1; -- ojo; acá\nSELECT 2;") == ["SELECT 1", "SELECT 2"])
check("y una sentencia sola sigue siendo una",
      store._sentencias("SELECT 1") == ["SELECT 1"])
# Un ON CONFLICT contra un indice PARCIAL tiene que repetir el WHERE del
# indice, o Postgres contesta que no encuentra ninguna restriccion que encaje.
check("el indice de (hoja, orden) es parcial y el ON CONFLICT lo repite",
      "WHERE hoja IS NOT NULL" in store.ESQUEMA
      and "ON CONFLICT (hoja, orden) WHERE hoja IS NOT NULL" in open(
          os.path.join(os.path.dirname(CARPETA_T), "store.py"), encoding="utf-8").read())

# --- 19. la planilla que espera para ser revisada --------------------------
# El token va en la direccion. Sin control, un «../../algo» leeria cualquier
# archivo del server.
print("La planilla que espera:")
for feo in ("", "zzz", "../../etc/passwd", "a" * 15, "A" * 16, "a" * 17,
            "aaaaaaaa/aaaaaaa", "0123456789abcde."):
    check(f"{feo!r} no abre ningun archivo",
          "ya no está" in _mensaje_de(lambda: A._ruta_de(feo)))
os.makedirs(A.CARPETA_CARGA, exist_ok=True)
_token = "0123456789abcdef"
_suyo = os.path.join(A.CARPETA_CARGA, _token + ".xlsx")
with open(_suyo, "wb") as f:
    f.write(b"PK\x03\x04")
check("un token bueno abre el suyo", A._ruta_de(_token) == _suyo)
os.remove(_suyo)
check("y si ya no esta, lo dice",
      "ya no está" in _mensaje_de(lambda: A._ruta_de(_token)))

# --- 20. lo que se le pide a Asinfo ----------------------------------------
print("Lo que se le pide a Asinfo:")
check("el numero sale del nombre", asinfo._numero("TEJEDURIA-MQ 003") == 3)
check("y si no esta en el nombre, del codigo", asinfo._numero("MAQUINA", "MQ 12") == 12)
check("y si no esta en ninguno, no se inventa", asinfo._numero("MAQUINA", None) is None)

pedidos = []
for _nombre, _real in _ASINFO_REAL.items():
    setattr(asinfo, _nombre, _real)
asinfo._consultar = lambda sql: (pedidos.append(sql), [])[1]
asinfo._cache.clear()
asinfo.produccion_mensual(10, meses=999)
check("no se piden mas de 60 meses", "TOP 60" in pedidos[0])
asinfo._cache.clear(); pedidos.clear()
asinfo.produccion_mensual(10, meses=0)
check("ni menos de uno", "TOP 1" in pedidos[0])

# El VALUES de SQL Server no admite mil filas: los pares van de a 400.
asinfo._cache.clear(); pedidos.clear()
asinfo.acumulados([(i, 1, "2026-01-01") for i in range(900)])
check("900 pares se piden en tres consultas", len(pedidos) == 3)
asinfo._cache.clear(); pedidos.clear()
datos, _, _ = asinfo.acumulados([(10, 1, "2026-01-01")])
check("una maquina que no produjo nada da cero de verdad", datos[(10, 1)] == (0.0, 0))
# La produccion del dia del service es anterior al service, no desgaste
# posterior: la fecha es exclusiva.
check("la fecha del ultimo service no se cuenta", "mi.fecha > CONVERT(date" in pedidos[0])

asinfo._cache.clear(); pedidos.clear()
kg = asinfo.kilos_desde(10, [date(2026, 1, 1), date(2026, 1, 1), None, "2026-03-01"])
check("una fecha repetida se pide una sola vez", len(kg) == 2)
check("y una vacia no se pide", None not in kg)

# Metabase corta en 2.000 filas sin avisar. Un resultado truncado no es un
# dato con menos filas: es un dato equivocado, y se trata como error.
class _RespuestaAlTope:
    status_code = 200

    def raise_for_status(self):
        pass

    def json(self):
        return {"data": {"rows": [[1, 1, 1, 1]] * asinfo._TOPE_FILAS}}


asinfo._consultar = _CONSULTAR_REAL
asinfo._session_token = "un token"
config.METABASE_URL = "http://metabase"
config.METABASE_USERNAME = "u"
config.METABASE_PASSWORD = "p"
sys.modules["requests"].post = lambda *a, **k: _RespuestaAlTope()
check("una respuesta con el tope de filas es un error, no un dato",
      "incompleto" in _mensaje_de(lambda: asinfo._consultar("SELECT 1")))
check("y se levanta como RespuestaTruncada",
      isinstance(_atrapar(lambda: asinfo._consultar("SELECT 1")),
                 asinfo.RespuestaTruncada))
asinfo._consultar = lambda sql: (pedidos.append(sql), [])[1]

# --- 21. cuanto deberia dar, con los numeros como salen de la base ---------
# Postgres devuelve `numeric` como Decimal, no como float. La pantalla divide
# uno por otro para sacar el rinde, y ahi los dos tipos no se mezclan.
print("Cuanto deberia dar:")
from decimal import Decimal as D

store.eficiencias = lambda: {
    10: {"id_maquina": 10, "rpm": D("25.00"), "sistemas": "102",
         "diametro": D("32.00"), "galga": 24, "tamano_rollo": D("1410.00"),
         "minutos_rollo": D("56.40"), "rollos_dia": D("12.00"), "kg_dia": D("270.00"),
         "rollos_dia_24": D("25.00"), "kg_dia_24": D("612.50"),
         "real_rollos_dia": D("9.00"), "real_kg_dia": D("202.50"),
         "real_rollos_24": D("17.00"), "real_kg_24": D("416.50")},
    # La 11 esta anotada en la planilla sin numeros: pasa de verdad, y sin
    # kilos de 12 h no hay rinde que sacar.
    11: {"id_maquina": 11, "rpm": None, "sistemas": None, "diametro": None,
         "galga": None, "tamano_rollo": None, "minutos_rollo": None,
         "rollos_dia": None, "kg_dia": D("0.00"), "rollos_dia_24": None,
         "kg_dia_24": None, "real_rollos_dia": None, "real_kg_dia": D("10.00"),
         "real_rollos_24": None, "real_kg_24": None},
}
store.gramajes = lambda id_maquina=None: []
cuerpo = c.get("/produccion").get_data(as_text=True)
check("la pantalla abre con los numeros como los devuelve la base", "202" in cuerpo)
check("202,5 de 270 es 75%", "75%" in cuerpo)
check("la que no tiene el calculo sale igual y no divide por cero",
      "MQ 2" in cuerpo and "%" in cuerpo)
check("la ficha de la maquina tambien abre con Decimales",
      c.get("/maquina/10").status_code == 200)

# --- 22. con todo apagado, ninguna pantalla se cae -------------------------
# Asinfo no contesta y todavia no hay ningun tipo cargado: es exactamente como
# arranca el programa en una base nueva. Media pantalla es mejor que un 500.
print("Con todo apagado:")


def _asinfo_caido(*a, **k):
    raise asinfo.AsinfoNoDisponible("Asinfo no contesta")


asinfo.maquinas = _asinfo_caido
asinfo.kilos_desde = _asinfo_caido
asinfo.acumulados = lambda pares: ({}, datetime.utcnow(), True)
asinfo.produccion_mensual = lambda i, meses=12: ([], datetime.utcnow(), True)
store.tipos = lambda incluir_inactivos=False: []
store.historial = lambda id_maquina=None, limite=200: []
store.responsables = lambda: []
store.ficha = lambda i: {}
store.fichas = lambda: {}
store.topes_por_maquina = lambda: {}
store.ultimos_por_maquina_y_tipo = lambda: {}
store.archivos = lambda i=None: []
store.archivo = lambda i: None
store.ajustes = lambda id_maquina=None, tela=None, limite=400: []
store.telas = lambda: []
store.resumen_ajustes = lambda: {}
store.consumo_hilo = lambda: []
store.agujas = lambda: {}
store.eficiencias = lambda: {}
store.gramajes = lambda i=None: []
store.filas_de = lambda cuadro: []
for ruta in ("/", "/registrar", "/tipos", "/arranque", "/carga", "/maquinas",
             "/archivos", "/ajustes", "/repuestos", "/repuestos?ver=bandas",
             "/produccion", "/maquina/10", "/?solo=vencidas"):
    check(f"{ruta} abre igual", c.get(ruta).status_code == 200)
check("y el semaforo dice por que esta vacio",
      "Asinfo" in c.get("/").get_data(as_text=True))
# La campanita NUNCA puede decir cero por un problema de red: un cero ahi
# significa «esta todo bien», que es el error que este programa evita.
with A.app.test_request_context("/"):
    check("la campanita no dice cero cuando no se sabe",
          A._campanita()["vencidas"] is None)
for ruta in ("/archivos/999999", "/maquina?numero=1"):
    check(f"{ruta} vuelve al listado, no rompe", c.get(ruta).status_code == 302)
check("healthz contesta igual", c.get("/healthz").status_code in (200, 503))


# --------------------------------------------------------------------------
# Lo que se leía a medias de la planilla de ajuste
# --------------------------------------------------------------------------
class _FalsoCursor:
    def __init__(self, guardado): self.guardado = guardado
    def execute(self, sql, args=()): self.guardado.update(sql=sql, args=args)
    def __enter__(self): return self
    def __exit__(self, *a): return False


class _FalsoConn:
    """Una conexión de mentira: guarda la consulta en vez de correrla."""
    def __init__(self, guardado): self.guardado = guardado
    def cursor(self, **k): return _FalsoCursor(self.guardado)
    def commit(self): pass
    def __enter__(self): return self
    def __exit__(self, *a): return False


def _hoja_con_dos_tablas():
    """Una hoja como la MAQ 53: dos tablas pegadas y notas al costado."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active; ws.title = "MAQ 1"
    titulos = ["MQ.  1", "Tipo de  MQ.", "FECHA", "SERIE", "Polea", "Polea",
               "Polea", "ajuste agujas ", "ESTIRAJE", "TIPO DE TELA", "HILO",
               "HILO", "HILO", "G /m2   crudo", "Longitud de Malla",
               "Rendimiento Crudo", "(G/ m2)", "KG/M"]
    ws.append(["CONTROL DE AGUSTES Y RENDIMIENTOS"])
    # A la derecha de la tabla, dos columnas sin título (18 y 19) y en la 20
    # arranca otra tabla entera con los mismos títulos.
    ws.append(titulos + [None, None, "MQ.  1"] + titulos[1:])
    ws.append([1, "MAYER 32", datetime(2026, 3, 4), "pro 30", "polea 5", None,
               None, "lycra pro 6,5", "B/9", "FALSO F", "20/1 KW", None, None,
               181.5, "29,0 LM", 4.05, None, None,
               "menos 5 cm solicitado por Oscar 11/8/2020", None,
               1, "MAYER 30", datetime(2021, 11, 20), None, "polea 115", None,
               None, None, None, "KIANA", "75F36"])
    return wb


wb_dos = _hoja_con_dos_tablas()
aj_dos, desc_dos = AJ.leer_ajustes(wb_dos, [MAQS_AJ[0]], hoy=date(2026, 8, 20))
izq = [a for a in aj_dos if a["orden"] < 1000]
der = [a for a in aj_dos if a["orden"] >= 1000]

# 1. La nota al costado
check("la nota de la derecha entra en el ajuste",
      izq and izq[0]["nota"] == "menos 5 cm solicitado por Oscar 11/8/2020")
check("la nota no repite lo que ya entro por su columna",
      izq and "FALSO F" not in (izq[0]["nota"] or "")
      and izq[0]["tela"] == "FALSO F")

# 3. La segunda tabla
check("la segunda tabla de la hoja se lee aparte", len(der) == 1)
check("la segunda tabla no se mete en las filas de la primera",
      izq and izq[0]["poleas"] == "polea 5" and izq[0]["tela"] == "FALSO F")
check("las dos tablas de una hoja no comparten (hoja, orden)",
      len({(a["hoja"], a["orden"]) for a in aj_dos}) == len(aj_dos))
check("la segunda tabla se avisa en pantalla",
      any("segunda tabla" in d["motivo"] for d in desc_dos))

# 2. La ficha escrita arriba de la hoja
def _hojas_con_ficha_escrita():
    from openpyxl import Workbook
    wb = Workbook()
    titulos = ["MQ.", "Tipo de  MQ.", "FECHA", "SERIE", "Polea", "ESTIRAJE",
               "TIPO DE TELA", "HILO"]
    ws = wb.active; ws.title = "MAQ 1"
    ws.append(["MAQUINA MAYER OV 3,2 QC 2016 AGUJAS 2976  DIAMETRO 34 "
               "GALGA 28  108 ALIMENTADORES  6 TRAK "])
    ws.append(titulos)
    # Dos hojas con la MISMA frase: está copiada y pegada.
    for nombre in ("MAQ 52", "MAQ 53"):
        otra = wb.create_sheet(nombre)
        otra.append(["MAQUINA JIUNN LONG  DIAMETRO 36 GALGA 28  "
                     "100 ALIMENTADORES  6 TRAK "])
        otra.append(titulos)
    return wb


MAQS_FICHA = MAQS_AJ + [{"id": 12, "numero": 53, "nombre": "TEJEDURIA-MQ 053"}]
escritas, _ = AJ.leer_fichas_escritas(_hojas_con_ficha_escrita(), MAQS_FICHA)
una = escritas.get(10) or {}
check("la ficha escrita arriba de la hoja se lee entera",
      (una.get("marca"), una.get("diametro"), una.get("galga"),
       una.get("alimentadores")) == ("MAYER", 34.0, 28, 108))
# El 2016 de «OV 3,2 QC 2016» es el año: las agujas son las 2976 que dice
# la palabra AGUJAS. Al revés, la ficha diría que la máquina tiene 2016.
check("el ano no se guarda como cantidad de agujas",
      una.get("agujas") == 2976 and una.get("anio") == 2016)

cambios, choques = AJ.completar_ficha(
    escritas, {10: {"marca": "MAYER", "galga": 24}})
solo_uno = next((c for c in cambios if c["id_maquina"] == 10), {})
check("la ficha escrita completa lo vacio y no pisa lo cargado",
      solo_uno.get("diametro") == 34.0 and "galga" not in solo_uno
      and any("galga" in d["motivo"] for d in choques))
check("una frase copiada en varias hojas no completa ninguna ficha",
      not [c for c in cambios if c["id_maquina"] in (11, 12)])

# Y al guardar, se escribe SÓLO lo que está vacío. Lo hace el motor con
# `coalesce`: entre leer la ficha y escribirla, alguien puede estar cargándola
# desde la pantalla, y lo que escribió a mano tiene que ganar siempre.
_sql = {}
store._conn = lambda: _FalsoConn(_sql)
store.completar_ficha_vacia(10, {"marca": "MAYER", "galga": None, "anio": 2016})
check("completar la ficha no pisa lo que ya esta cargado",
      "coalesce(marca, %s)" in _sql.get("sql", ""))
check("y no escribe los campos vacios", "galga" not in _sql.get("sql", ""))
check("con los valores de la hoja", ("MAYER", 2016) == tuple(_sql["args"][1:3]))

# 4. El número sin título de INVENTARIO LEVAS
def _levas_con_numero_suelto():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active; ws.title = "INVENTARIO LEVAS"
    ws.append(["INVENTARIO DE  LEVAS", None, None, None, None, None, None,
               None, None, None, None, None, None, None,
               " CANTIDAD DE LEVAS POR TELA"])
    ws.append(["MAQUINA ", "CODIGO", "CANTIDAD", "UBICACION", "ACCIONAMIENTO"])
    ws.append(["MAYER (1-2-3 )", "30-32  274080-1", 142, "cilindro",
               "ANULACION", 288])
    return wb


levas_sueltas, desc_levas = AJ.leer_levas(_levas_con_numero_suelto())
check("el numero sin titulo de las levas se avisa",
      any("288" in d["motivo"] for d in desc_levas))
check("el numero sin titulo no se guarda como si fuera otra cosa",
      levas_sueltas and levas_sueltas[0]["cantidad"] == 142
      and 288 not in levas_sueltas[0].values())

print()
if fallos:
    print("FALLARON:", ", ".join(fallos)); sys.exit(1)
print("Todos los tests OK")
