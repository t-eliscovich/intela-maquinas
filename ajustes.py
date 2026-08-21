"""Leer la planilla CONTROL DE AJUSTE, tal como viene.

Es otra planilla que ya usan en planta, distinta de la de mantenimiento. Tiene
cincuenta hojas y adentro conviven ocho cosas que no se parecen en nada:

  * 42 hojas de AJUSTES, una por máquina — cómo se pone la máquina para tejer
    cada tela. Es el bloque grande: ~1.200 filas desde 2021.
  * AGUJAS — qué aguja lleva cada máquina.
  * INVENTARIO LEVAS — las levas, agrupadas por modelo.
  * BANDAS — las bandas Memminger y cuántas hay de cada medida.
  * Eficiencia producción — cuánto debería dar cada máquina en 12 horas.
  * consumo de hilo — cuánto hilo lleva cada tela.
  * Hoja3 — el peso medido de la tela de cada máquina.
  * CODIGO DE AGUJAS — la misma información que AGUJAS, agrupada por modelo.
    No se carga: sería el mismo dato dos veces, y dos copias del mismo dato es
    el camino más corto a que una quede vieja.

Las reglas que valen para todo el módulo:

  * **Cada dato se busca por su nombre, no por su posición.** Ninguna hoja está
    armada igual: hay hojas con tres columnas «Polea» y otras con cuatro, y una
    que directamente perdió los títulos.
  * **Lo que no se entiende no se carga.** Se devuelve aparte, con el motivo, y
    la pantalla lo muestra antes de guardar. Vacío es correcto; inventado no.
  * **La fecha es la primera celda con fecha de verdad de la fila**, no la
    columna que dice «FECHA». En varias hojas los datos quedaron corridos una
    columna y la columna del título ya no es la que tiene el dato.
"""
from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime

from openpyxl import load_workbook

import excel

# Cuántas filas miramos arriba de cada hoja buscando los títulos.
_FILAS_CABECERA = 6


def _limpio(texto) -> str:
    """Minúsculas, sin tildes, sin puntuación. Para comparar títulos."""
    s = str(texto or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9 ]+", " ", s)


def _apretado(texto) -> str:
    return re.sub(r"\s+", " ", _limpio(texto)).strip()


def _texto(valor) -> str | None:
    """El texto de una celda. Una fecha NO es texto: devuelve vacío.

    Suena raro, pero es lo que evita el error más feo de esta planilla. Cuando
    una hoja quedó corrida una columna, la fecha aterriza en el casillero del
    modelo de la máquina, y sin este freno la ficha termina diciendo que la
    MQ 27 es una «14/11/2025». Mejor vacío.
    """
    if valor in (None, "") or isinstance(valor, (datetime, date)):
        return None
    s = re.sub(r"\s+", " ", str(valor).strip())
    return s or None


def _numero(valor) -> int | None:
    if isinstance(valor, bool) or valor in (None, ""):
        return None
    if isinstance(valor, (int, float)):
        return int(valor)
    hallados = re.findall(r"\d+", str(valor))
    return int(hallados[0]) if hallados else None


def _decimal(valor) -> float | None:
    """Sólo acepta números de verdad.

    A propósito no lee «30 LM dibujo» como 30: esa celda es una anotación, no
    un gramaje. Un número sacado a la fuerza de un texto entra a la base sin
    avisar y después nadie sabe de dónde salió.
    """
    if isinstance(valor, bool) or valor in (None, ""):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip().replace(",", ".")
    return float(s) if re.fullmatch(r"-?\d+(\.\d+)?", s) else None


# Una medida escrita a mano: el número primero y detrás su unidad. La unidad
# sólo puede estar hecha con las letras de gramos, kilos y metros —g, gr, kg,
# m, m2— más barras, asteriscos, puntos, comas y espacios. Ni una letra más:
# la «L» de «28,2 LM» es lo único que hace falta para que las longitudes de
# malla que alguien escribió en la columna del gramaje se queden afuera.
_MEDIDA = re.compile(
    r"""^\s*(-?\d+(?:[.,]\d+)?)               # el número, y va primero
         (?:[\s/*.,]*(?:[mM]2|[gGrRkKmM]))*     # la unidad, con sus separadores
         [\s/*.,]*$""",
    re.VERBOSE,
)


def _medida(valor) -> float | None:
    """El número de una celda que trae la unidad pegada: «1,80 kg/m», «138 g».

    En tres columnas —gramaje crudo, gramaje terminado y kg/m— casi nadie
    escribió el número solo. Hay «2,48 kg/m», «4,40 *KG», «165 gr», «183G».
    Con `_decimal` entraba una celda de cada quince: el resto se perdía por
    tener al lado la unidad, que es justo lo que confirma que eso ES una medida.

    Pero la misma columna se usó también para otra cosa: «spander 108»,
    «V 955», «plegado», «guia aguja [4]» y, sobre todo, cuarenta longitudes de
    malla («30 LM dibujo», «28,2 LM»). Ninguna de ésas es un gramaje.

    La regla que separa las dos cosas, mirando la planilla: **la celda es el
    número y su unidad, y nada más**. El número va adelante —eso ya deja afuera
    «spander 108» y «V 955»— y atrás sólo puede venir una unidad escrita con
    las letras de gramos, kilos y metros. La «L» de LM alcanza para que las
    longitudes de malla no entren. Un segundo número tampoco pasa: «1,20*3,20»
    son dos medidas de un cilindro y «12/23/2019» es una fecha.

    El número se guarda tal como está escrito. No se convierte nada: la misma
    columna trae «3,50 kg/m» y «3,50 m/kg», que son unidades inversas, y elegir
    cuál quiso decir el que la escribió sería inventar el dato.
    """
    if isinstance(valor, bool) or valor in (None, ""):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    hallado = _MEDIDA.match(str(valor).strip())
    return float(hallado.group(1).replace(",", ".")) if hallado else None


def _fecha(valor, hoy: date) -> date | None:
    if isinstance(valor, datetime):
        f = valor.date()
    elif isinstance(valor, date):
        f = valor
    else:
        return None
    # 1990 corta las fechas escritas mal (hay un 2002 que quería ser 2022) y
    # las que Excel guardó como número. Una fecha futura tampoco existe.
    return f if date(1990, 1, 1) <= f <= hoy else None


def _juntar(valores) -> str | None:
    """Varias columnas que dicen lo mismo, en un solo texto.

    Las poleas y los hilos son tres o cuatro columnas que se llaman todas
    igual y no significan siempre lo mismo. Se guardan juntas, separadas por
    «·», en vez de repartirlas en columnas fijas: partirlas sería inventar una
    estructura que el papel no tiene.
    """
    partes = [t for t in (_texto(v) for v in valores) if t]
    return " · ".join(partes) or None


def _filas_de(ws) -> list[list]:
    return [list(f) for f in ws.iter_rows(values_only=True)]


def _vacia(fila) -> bool:
    return all(c in (None, "") for c in fila)


# --------------------------------------------------------------------------
# ¿Es esta planilla?
# --------------------------------------------------------------------------
_HOJAS_PROPIAS = ("agujas", "inventario levas", "bandas", "consumo de hilo",
                  "eficiencia produccion", "codigo de agujas")


def es_planilla_ajuste(ruta: str) -> bool:
    """Se reconoce sola por los nombres de las hojas.

    Con dos hojas propias alcanza: el nombre del archivo no sirve de pista
    porque cada uno lo guarda como quiere.
    """
    try:
        wb = load_workbook(ruta, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return False
    try:
        nombres = {_apretado(n) for n in wb.sheetnames}
    finally:
        wb.close()
    return sum(1 for h in _HOJAS_PROPIAS if h in nombres) >= 2


# --------------------------------------------------------------------------
# Las 42 hojas de ajustes
# --------------------------------------------------------------------------
# Cada campo con los títulos que puede llegar a tener. Se prueba el más largo
# primero: «longitud de malla manual» tiene que ganarle a «longitud de malla»,
# y «g m2 crudo» a «g m2». Si no, el título corto se lleva las dos columnas.
_TITULOS_AJUSTE = [
    ("malla_manual",  ("longitud de malla manual", "malla manual")),
    ("malla",         ("longitud de malla", "malla")),
    # Crudo y terminado son dos columnas distintas y se llaman casi igual. El
    # crudo va PRIMERO y con su etiqueta completa: al revés, «g m2 crudo»
    # entraría por el startswith de «g m2» y el crudo se cargaría en la columna
    # del terminado. La que dice sólo «G/m2» es la del terminado: en las 42
    # hojas va siempre después de «rendimiento crudo», que es donde las cuatro
    # hojas que sí lo escribieron dicen «G/m2 terminado».
    ("gramaje_crudo", ("g m2 crudo", "gr m2 crudo")),
    ("gramaje_terminado", ("g m2 terminado", "gr m2 terminado", "g m2", "gr m2")),
    ("rendimiento",   ("rendimiento crudo", "rendimiento")),
    ("kg_m",          ("kg m terminado", "kg m")),
    ("tipo_maquina",  ("tipo de mq", "tipo de maquina")),
    ("cilindro",      ("serie",)),
    ("poleas",        ("polea",)),
    ("ajuste_agujas", ("ajuste agujas", "ajuste de agujas")),
    ("estiraje",      ("estiraje", "stiraje")),
    ("tela",          ("tipo de tela", "tela")),
    ("hilos",         ("hilo",)),
    ("fecha",         ("fecha",)),
]

# Los campos que aceptan varias columnas. El resto se queda con la primera.
_REPETIBLES = ("poleas", "hilos")


def _campo_del_titulo(celda) -> str | None:
    """Qué campo es ese título, o vacío si no es ninguno."""
    # Una fecha de verdad en la fila de títulos es la columna de la fecha:
    # alguien escribió «FECHA 5/01/2021» encima y Excel lo guardó así.
    if isinstance(celda, (datetime, date)):
        return "fecha"
    t = _apretado(celda)
    if not t:
        return None
    for campo, etiquetas in _TITULOS_AJUSTE:
        if any(t == e or t.startswith(e) for e in etiquetas):
            return campo
    return None


def _mapa_de_titulos(fila, desde: int = 0, hasta: int | None = None) -> dict[str, list[int]]:
    """{campo: [columnas]} a partir de la fila de títulos.

    `desde` y `hasta` acotan a qué tabla se está mirando. Existen porque la
    hoja MAQ 53 tiene DOS tablas pegadas, con los mismos títulos: sin este
    freno las columnas «Polea» y «HILO» de la tabla de la derecha entraban en
    las filas de la izquierda, y cada ajuste de la MQ 53 salía con las poleas
    de otro.
    """
    mapa: dict[str, list[int]] = {}
    for i, celda in enumerate(fila):
        if i < desde or (hasta is not None and i >= hasta):
            continue
        campo = _campo_del_titulo(celda)
        if campo is None or (campo in mapa and campo not in _REPETIBLES):
            continue
        mapa.setdefault(campo, []).append(i)
    return mapa


def _fila_de_titulos(filas) -> tuple[int | None, dict[str, list[int]]]:
    """La primera fila de arriba que parece los títulos: tres campos o más."""
    mejor, mejor_mapa, mejor_n = None, {}, 0
    for i, fila in enumerate(filas[:_FILAS_CABECERA]):
        mapa = _mapa_de_titulos(fila)
        if len(mapa) > mejor_n:
            mejor, mejor_mapa, mejor_n = i, mapa, len(mapa)
    return (mejor, mejor_mapa) if mejor_n >= 3 else (None, {})


# Cuando una hoja perdió los títulos, se lee por posición con el orden que
# tienen las otras 39. Se avisa en pantalla: leer por posición es adivinar, y
# adivinar en silencio es lo único que no se puede hacer.
_ORDEN_SUPLENTE = {
    "tipo_maquina": [1], "fecha": [2], "cilindro": [3], "poleas": [4, 5, 6],
    "ajuste_agujas": [7], "estiraje": [8], "tela": [9], "hilos": [10, 11, 12],
    "gramaje_crudo": [13], "malla_manual": [14], "malla": [15],
    "rendimiento": [16], "gramaje_terminado": [17], "kg_m": [18],
}


def _ultima_titulada(mapa) -> int:
    """La columna más a la derecha que tiene título. -1 si no hay ninguna."""
    return max((c for cols in mapa.values() for c in cols), default=-1)


def _segunda_tabla(fila_titulos):
    """La otra tabla de ajustes pegada a la derecha, si la hay.

    La hoja MAQ 53 tiene DOS tablas al lado: la de la izquierda es la máquina
    como está hoy y la de la derecha son nueve ajustes con su propio
    encabezado, que también dice «MQ. 53». Se reconoce por el título
    REPETIDO: cuando un título que puede aparecer una sola vez —«FECHA»,
    «TIPO DE TELA», «ESTIRAJE»— vuelve a aparecer más a la derecha, ahí
    empieza otra tabla.

    Que el título se repita no alcanza: en la MAQ 52 hay dos «Longitud de
    Malla» seguidas y son dos columnas de la MISMA tabla. Por eso se exige
    además que de ahí a la derecha haya una tabla entera —cinco campos y la
    fecha—, que es lo que separa una tabla de una columna repetida.

    Devuelve (columna donde empieza, mapa) o (None, {}).
    """
    vistos: set[str] = set()
    for i, celda in enumerate(fila_titulos):
        campo = _campo_del_titulo(celda)
        if campo is None or campo in _REPETIBLES:
            continue
        if campo not in vistos:
            vistos.add(campo)
            continue
        mapa = _mapa_de_titulos(fila_titulos, desde=i)
        if len(mapa) < 5 or "fecha" not in mapa:
            continue
        # La tabla no empieza en el título repetido sino en la columna del
        # número de máquina, que está a su izquierda y sin título propio
        # («MQ. 53»). Es la que dice de quién son esos ajustes.
        inicio = i
        while inicio > 0 and _apretado(fila_titulos[inicio - 1]):
            inicio -= 1
        return inicio, mapa
    return None, {}


def _nota_al_costado(fila, desde: int, hasta: int) -> str | None:
    """Lo que quedó escrito a la derecha de la última columna con título.

    Son ~100 celdas en toda la planilla y hasta ahora se tiraban enteras. Ahí
    está lo único que explica POR QUÉ se cambió un ajuste: «menos 5
    centímetros solicitado por Oscar 11/8/2020, cambio 11/11/2020 a 32,5 LM».
    Al lado hay también medidas sueltas —«tención 30», «spander 132»— que no
    tienen columna propia en ningún lado.

    Se toma sólo lo que está DESPUÉS de la última columna con título, así que
    no puede repetir lo que ya entró por su campo. Y se corta antes de la
    segunda tabla de la MAQ 53, que son ajustes y no una nota.
    """
    return _juntar(_celdas(fila, range(desde, hasta)))


def _celdas(fila, columnas) -> list:
    return [fila[i] if i < len(fila) else None for i in columnas]


def _una(fila, mapa, campo):
    columnas = mapa.get(campo)
    if not columnas:
        return None
    i = columnas[0]
    return fila[i] if i < len(fila) else None


def leer_ajustes(wb, maquinas, hoy=None) -> tuple[list[dict], list[dict]]:
    """Las 42 hojas de ajuste. Devuelve (filas, descartes)."""
    hoy = hoy or date.today()
    por_numero = {m["numero"]: m for m in maquinas if m.get("numero") is not None}
    salida, descartes = [], []

    for nombre_hoja in wb.sheetnames:
        limpio = _apretado(nombre_hoja)
        if limpio in _HOJAS_PROPIAS or limpio == "hoja3":
            continue
        if limpio == "agustes":
            # La hoja vieja, de cuando estaba todo junto antes de partirla en
            # una hoja por máquina. Parecía repetida y NO lo es: de sus 201
            # ajustes, 124 no están en ninguna hoja por máquina — son los más
            # viejos, casi todos sin fecha. Se lee aparte porque el número de
            # máquina está en cada fila, no en el nombre de la hoja.
            continue
        numero = _numero(nombre_hoja)
        if numero is None:
            descartes.append({"donde": nombre_hoja,
                              "motivo": "La hoja no dice de qué máquina es"})
            continue
        maquina = por_numero.get(numero)
        if not maquina:
            descartes.append({"donde": nombre_hoja,
                              "motivo": f"La máquina {numero} no está en Asinfo"})
            continue

        filas = _filas_de(wb[nombre_hoja])
        corte, mapa = _fila_de_titulos(filas)
        # Si hay otra tabla pegada a la derecha, la de la izquierda termina
        # ahí: sus columnas no son de esta.
        inicio2, mapa2 = _segunda_tabla(filas[corte]) if corte is not None else (None, {})
        if inicio2 is not None:
            mapa = _mapa_de_titulos(filas[corte], hasta=inicio2)
        # Sin columna de fecha, los títulos no sirven: quiere decir que a esa
        # hoja se le perdieron las primeras columnas y TODO el resto está
        # corrido. Pasa en una sola hoja, pero si se leyera igual saldrían las
        # poleas donde va la tela. Se lee por posición y se avisa.
        por_posicion = corte is None or "fecha" not in mapa
        if por_posicion:
            corte, mapa = 1, dict(_ORDEN_SUPLENTE)
            inicio2, mapa2 = None, {}

        # Hasta dónde llega la tabla de la izquierda, y qué hay más a la
        # derecha: otra tabla de ajustes (MAQ 53) o anotaciones sueltas.
        ultima = _ultima_titulada(mapa)
        tope_nota = inicio2 if inicio2 is not None else max(
            (len(f) for f in filas), default=ultima + 1)

        leidas = _leer_bloque(filas, corte, mapa, maquina, nombre_hoja, hoy,
                              salida, ultima + 1, tope_nota, 0, por_posicion)

        if inicio2 is not None:
            # La segunda tabla dice de qué máquina es en su primera celda
            # («MQ. 53»). Si no coincide con la hoja, no se adivina: se avisa.
            suya = _numero(filas[corte][inicio2]) if inicio2 < len(filas[corte]) else None
            duena = por_numero.get(suya) if suya is not None else None
            if duena is None:
                descartes.append({
                    "donde": nombre_hoja,
                    "motivo": "Hay una segunda tabla de ajustes a la derecha y "
                              "no dice de qué máquina es: no se leyó"})
            else:
                # El orden arranca en 1.000 para que la clave (hoja, orden)
                # siga siendo única: las dos tablas están en la misma hoja y
                # con la numeración de la izquierda se pisarían entre ellas.
                ultima2 = _ultima_titulada(mapa2)
                otras = _leer_bloque(filas, corte, mapa2, duena, nombre_hoja,
                                     hoy, salida, ultima2 + 1,
                                     max((len(f) for f in filas), default=0),
                                     1000, False)
                descartes.append({
                    "donde": nombre_hoja,
                    "motivo": f"La hoja tiene una segunda tabla pegada a la "
                              f"derecha: {otras} ajustes más de la MQ {suya}. "
                              "Conviene mirarlos."})
                leidas += otras

        if not leidas:
            descartes.append({"donde": nombre_hoja,
                              "motivo": "La hoja no tiene ningún ajuste cargado"})
        elif por_posicion:
            descartes.append({
                "donde": nombre_hoja,
                "motivo": f"La hoja no tiene títulos: se leyeron {leidas} filas "
                          "por posición. Conviene mirarlas."})

    return salida, descartes


def _leer_bloque(filas, corte, mapa, maquina, nombre_hoja, hoy, salida,
                 desde_nota, hasta_nota, base_orden, por_posicion) -> int:
    """Lee una tabla de ajustes y la agrega a `salida`. Devuelve cuántas leyó.

    Es una función aparte porque la MAQ 53 tiene DOS tablas en la misma hoja y
    las dos se leen igual; lo único que cambia es en qué columnas están.
    """
    leidas = 0
    for n, fila in enumerate(filas[corte + 1:], start=1):
        if _vacia(fila):
            continue
        # Cuando se lee por posición no se sabe dónde terminaba el
        # encabezado, así que una fila de títulos puede caer acá abajo. Se
        # reconoce sola: tiene tres nombres de columna y ningún dato.
        if por_posicion and len(_mapa_de_titulos(fila)) >= 3:
            continue
        # La fecha es la primera celda con fecha de verdad de la fila. En
        # varias hojas los datos quedaron corridos una columna y la
        # columna del título ya no es la que tiene el dato. Se mira sólo
        # adentro de la tabla: la de la izquierda no puede quedarse con la
        # fecha de la de la derecha.
        propias = _celdas(fila, range(min((c for cols in mapa.values()
                                           for c in cols), default=0),
                                      desde_nota))
        fecha = next((f for f in (_fecha(c, hoy) for c in propias) if f), None)

        tipo = _texto(_una(fila, mapa, "tipo_maquina"))
        if not tipo:
            # Se corrió una columna: el tipo quedó en la primera.
            primera = _texto(fila[0] if fila else None)
            if primera and not re.fullmatch(r"\d+", primera):
                tipo = primera

        item = {
            "id_maquina": maquina["id"],
            "maquina_nombre": maquina["nombre"],
            "fecha": fecha,
            "tipo_maquina": tipo,
            "cilindro": _texto(_una(fila, mapa, "cilindro")),
            "poleas": _juntar(_celdas(fila, mapa.get("poleas", []))),
            "ajuste_agujas": _texto(_una(fila, mapa, "ajuste_agujas")),
            "estiraje": _texto(_una(fila, mapa, "estiraje")),
            "tela": _texto(_una(fila, mapa, "tela")),
            "hilos": _juntar(_celdas(fila, mapa.get("hilos", []))),
            # Con `_medida` y no con `_decimal`: casi nadie escribió el
            # número solo, y «138 g» es un gramaje igual que 138.
            "gramaje_crudo": _medida(_una(fila, mapa, "gramaje_crudo")),
            "gramaje_terminado": _medida(_una(fila, mapa, "gramaje_terminado")),
            "malla_manual": _texto(_una(fila, mapa, "malla_manual")),
            "malla": _texto(_una(fila, mapa, "malla")),
            "rendimiento": _decimal(_una(fila, mapa, "rendimiento")),
            "kg_m": _medida(_una(fila, mapa, "kg_m")),
            "nota": _nota_al_costado(fila, desde_nota, hasta_nota),
            "hoja": nombre_hoja,
            "orden": base_orden + n,
        }
        # Una fila que sólo trae el número y el modelo de la máquina no es
        # un ajuste: es una fila que quedó empezada. La nota cuenta: una
        # fila que sólo dice «produccion normal» es lo que anotó el
        # mecánico ese día y no hay dónde más guardarlo.
        util = ("cilindro", "poleas", "ajuste_agujas", "estiraje", "tela",
                "hilos", "gramaje_crudo", "gramaje_terminado",
                "malla_manual", "malla", "nota")
        if not fecha and not any(item[c] for c in util):
            continue
        salida.append(item)
        leidas += 1
    return leidas


def leer_agustes(wb, maquinas, hoy=None) -> tuple[list[dict], list[dict]]:
    """La hoja «AGUSTES»: los ajustes viejos, todos juntos.

    Es de antes de partir la planilla en una hoja por máquina. Parece repetida
    y no lo es: la mayoría de sus filas no quedaron en ninguna hoja. La
    diferencia con las otras es que acá el número de máquina va en la primera
    columna de cada fila — y cuando se repite, no se vuelve a escribir: se
    arrastra el de la fila de arriba, que es como está escrita la hoja.
    """
    hoy = hoy or date.today()
    hoja = next((n for n in wb.sheetnames if _apretado(n) == "agustes"), None)
    if not hoja:
        return [], []
    por_numero = {m["numero"]: m for m in maquinas if m.get("numero") is not None}
    filas = _filas_de(wb[hoja])
    corte, mapa = _fila_de_titulos(filas)
    if corte is None or "fecha" not in mapa:
        return [], [{"donde": hoja, "motivo": "No se encontraron los títulos"}]

    # Igual que en las hojas por máquina, a la derecha de la última columna
    # con título quedaron anotaciones sueltas: 40 celdas en esta hoja.
    ultima = _ultima_titulada(mapa)
    tope_nota = max((len(f) for f in filas), default=ultima + 1)

    salida, descartes = [], []
    numero = None
    for n, fila in enumerate(filas[corte + 1:], start=1):
        if _vacia(fila):
            continue
        propio = _numero(fila[0] if fila else None)
        if propio is not None:
            numero = propio
        maquina = por_numero.get(numero) if numero is not None else None
        if not maquina:
            descartes.append({"donde": f"{hoja}, fila {n}",
                              "motivo": f"La máquina {numero} no está en Asinfo"})
            continue

        fecha = next((f for f in (_fecha(c, hoy) for c in fila) if f), None)
        item = {
            "id_maquina": maquina["id"],
            "maquina_nombre": maquina["nombre"],
            "fecha": fecha,
            "tipo_maquina": _texto(_una(fila, mapa, "tipo_maquina")),
            "cilindro": _texto(_una(fila, mapa, "cilindro")),
            "poleas": _juntar(_celdas(fila, mapa.get("poleas", []))),
            "ajuste_agujas": _texto(_una(fila, mapa, "ajuste_agujas")),
            "estiraje": _texto(_una(fila, mapa, "estiraje")),
            "tela": _texto(_una(fila, mapa, "tela")),
            "hilos": _juntar(_celdas(fila, mapa.get("hilos", []))),
            "gramaje_crudo": _medida(_una(fila, mapa, "gramaje_crudo")),
            "gramaje_terminado": _medida(_una(fila, mapa, "gramaje_terminado")),
            "malla_manual": _texto(_una(fila, mapa, "malla_manual")),
            "malla": _texto(_una(fila, mapa, "malla")),
            "rendimiento": _decimal(_una(fila, mapa, "rendimiento")),
            "kg_m": _medida(_una(fila, mapa, "kg_m")),
            "nota": _nota_al_costado(fila, ultima + 1, tope_nota),
            "hoja": hoja,
            "orden": n,
        }
        util = ("cilindro", "poleas", "ajuste_agujas", "estiraje", "tela",
                "hilos", "gramaje_crudo", "gramaje_terminado",
                "malla_manual", "malla", "nota")
        if not fecha and not any(item[c] for c in util):
            continue
        salida.append(item)
    return salida, descartes


def _firma(a) -> tuple:
    """Cómo se reconoce que dos filas son el mismo ajuste: la máquina, el día y
    la tela. Sirve para no cargar dos veces lo que la hoja vieja repite."""
    return (a["id_maquina"], a["fecha"],
            (a["tela"] or "").upper().strip(),
            (a["hilos"] or "").upper().strip())


# --------------------------------------------------------------------------
# La ficha escrita arriba de la hoja
# --------------------------------------------------------------------------
# Trece hojas —las de las máquinas 51 a 63— tienen la ficha de la máquina
# escrita en una frase suelta arriba de todo, sin títulos ni columnas:
# «MAQUINA JIUNN LONG DIAMETRO 36 GALGA 28 100 ALIMENTADORES 6 TRAK». No la
# leía nadie. La 58 es el caso que más pesa: es la única máquina cuya ficha
# está VACÍA en la planilla de mantenimiento, y esta frase la completa entera.
#
# Cada número se busca pegado a su palabra, y del lado en que está escrito:
# los alimentadores llevan el número adelante («100 ALIMENTADORES») y las
# agujas atrás («AGUJAS 2976»). Las agujas NO aceptan el número de adelante a
# propósito: la 58 dice «MAYER OV 3,2 QC 2016 AGUJAS 2976» y el 2016 es el
# año, no la cantidad de agujas.
_DIAMETRO_ESCRITO = re.compile(r"di[aá]metro\s*(\d+(?:[.,]\d+)?)", re.I)
_GALGA_ESCRITA = re.compile(r"galga\s*(\d+)", re.I)
_ALIMENTADORES_ESCRITOS = re.compile(
    r"(?:(\d+)\s*alimentadores|alimentadores\s*(\d+))", re.I)
_AGUJAS_ESCRITAS = re.compile(r"agujas\s*(\d+)", re.I)

# El año va suelto adentro del nombre del modelo («OV 3,2 QC 2016»). Se busca
# sólo ahí, antes de la primera medida: más a la derecha, un año de cuatro
# cifras y una cantidad de agujas se parecen demasiado.
_ANIO_ESCRITO = re.compile(r"\b(19\d\d|20[0-3]\d)\b")

# Dónde corta la marca: en la primera palabra que ya es una medida.
_PALABRAS_DE_FICHA = ("diametro", "diámetro", "galga", "alimentadores",
                      "agujas", "trak")

# Cuántas filas de arriba se miran. La frase está arriba de los títulos, y
# sólo ahí: más abajo, «con galga 28» es una anotación de una fila (MQ 22).
_FILAS_DE_FICHA = 3

# Los campos que puede traer la frase. `modelo` va aparte de `marca` porque
# «MAYER OV 3,2 QC 2016» son las dos cosas escritas juntas.
CAMPOS_FICHA_ESCRITA = ("marca", "modelo", "anio", "diametro", "galga",
                        "alimentadores", "agujas")


def _hallado(busqueda) -> str | None:
    return busqueda.group(1) if busqueda else None


def _numero_escrito(regla, texto) -> str | None:
    """El número que encontró la regla. Algunas lo tienen adelante y otras
    atrás, así que se devuelve el grupo que haya matcheado."""
    hallado = regla.search(texto)
    if hallado is None:
        return None
    return next((g for g in hallado.groups() if g), None)


def _ficha_escrita(texto: str) -> dict | None:
    """La ficha que dice esa frase, o vacío si no es una frase de ficha."""
    bajo = texto.lower()
    if not bajo.lstrip().startswith("maquina"):
        return None
    if not any(p in bajo for p in _PALABRAS_DE_FICHA):
        return None

    # La marca es lo que va entre «MAQUINA» y la primera medida. Se pasa por
    # `excel._marca_pareja`, que es la misma lista de marcas que usa la
    # planilla de mantenimiento: dos listas de marcas terminan en dos plantas
    # distintas con la misma máquina escrita de dos formas.
    arranque = bajo.index("maquina") + len("maquina")
    corte = min((bajo.find(p, arranque) for p in _PALABRAS_DE_FICHA
                 if bajo.find(p, arranque) >= 0), default=len(texto))
    bruto = texto[arranque:corte].strip(" .,-")
    anio = _numero(_hallado(_ANIO_ESCRITO.search(bruto)))
    if anio is not None:
        bruto = _ANIO_ESCRITO.sub("", bruto, count=1).strip(" .,-")
    marca = excel._marca_pareja(bruto) if bruto else None
    modelo = None
    if marca and bruto and _apretado(bruto) != _apretado(marca):
        # Lo que sobra después de la marca es el modelo: «OV 3,2 QC».
        modelo = bruto[len(marca):].strip(" .,-") or None
    elif not marca:
        modelo = bruto or None

    return {
        "marca": _texto(marca),
        "modelo": _texto(modelo),
        "anio": anio,
        # El diámetro puede estar escrito con coma («3,2»): va con
        # `_decimal`, que la entiende. Los otros tres son enteros.
        "diametro": _decimal(_numero_escrito(_DIAMETRO_ESCRITO, texto)),
        "galga": _numero(_numero_escrito(_GALGA_ESCRITA, texto)),
        "alimentadores": _numero(_numero_escrito(_ALIMENTADORES_ESCRITOS, texto)),
        "agujas": _numero(_numero_escrito(_AGUJAS_ESCRITAS, texto)),
    }


def leer_fichas_escritas(wb, maquinas) -> tuple[dict[int, dict], list[dict]]:
    """La ficha escrita arriba de cada hoja de ajuste. {id_maquina: ficha}.

    Devuelve también si esa misma frase está escrita igual en otras hojas: en
    la planilla, las hojas 51 a 57 tienen las SIETE la misma frase, palabra
    por palabra, y la de mantenimiento dice que la 53 es de 34 pulgadas y 96
    alimentadores y no de 36 y 100. Está copiada y pegada. Se lee igual —el
    dato está escrito y hay que verlo— pero se marca, y `completar_ficha` no
    escribe una frase copiada en ninguna ficha.
    """
    por_numero = {m["numero"]: m for m in maquinas if m.get("numero") is not None}
    leidas: dict[int, dict] = {}
    descartes: list[dict] = []

    for nombre_hoja in wb.sheetnames:
        limpio = _apretado(nombre_hoja)
        if limpio in _HOJAS_PROPIAS or limpio in ("hoja3", "agustes"):
            continue
        numero = _numero(nombre_hoja)
        maquina = por_numero.get(numero) if numero is not None else None
        if not maquina:
            continue
        filas = _filas_de(wb[nombre_hoja])
        for fila in filas[:_FILAS_DE_FICHA]:
            frase = next((t for t in (_texto(c) for c in fila)
                          if t and _ficha_escrita(t)), None)
            if not frase:
                continue
            ficha = _ficha_escrita(frase)
            ficha.update({"id_maquina": maquina["id"], "numero": numero,
                          "hoja": nombre_hoja, "texto": frase,
                          "copiada": False})
            leidas[maquina["id"]] = ficha
            break

    # Una frase que está escrita igual en más de una hoja no es la ficha de
    # ninguna de ellas: es una copia. Se marcan todas, incluida la primera.
    cuantas: dict[str, int] = {}
    for ficha in leidas.values():
        clave = _apretado(ficha["texto"])
        cuantas[clave] = cuantas.get(clave, 0) + 1
    for ficha in leidas.values():
        repetida = cuantas[_apretado(ficha["texto"])]
        if repetida > 1:
            ficha["copiada"] = True
            descartes.append({
                "donde": ficha["hoja"],
                "motivo": f"La ficha escrita arriba de la hoja está copiada "
                          f"igual en {repetida} hojas: no se puede saber de "
                          "qué máquina es. No se completa nada con ella."})
    return leidas, descartes


def _mismo_numero(uno, otro) -> bool:
    """Dos números son el mismo aunque uno venga de la base y otro del Excel."""
    try:
        return abs(float(uno) - float(otro)) < 0.005
    except (TypeError, ValueError):
        return _apretado(uno) == _apretado(otro)


def completar_ficha(leidas: dict[int, dict],
                    fichas: dict[int, dict]) -> tuple[list[dict], list[dict]]:
    """Qué se puede completar de cada ficha con lo que dice la hoja.

    Devuelve (cambios, descartes). Cada cambio es
    {"id_maquina": ..., "campo": valor, ...} con SÓLO los campos que hoy
    están vacíos: lo que ya está cargado no se pisa nunca. La ficha de la
    máquina la cargó el mecánico desde la pantalla o desde la planilla de
    mantenimiento, y una frase suelta arriba de una hoja de ajustes no le
    puede ganar a eso.

    Lo que no coincide tampoco se corrige solo: sale en descartes con las dos
    versiones, para que el mecánico decida cuál vale.
    """
    cambios, descartes = [], []
    for id_maquina, escrita in sorted(leidas.items()):
        actual = fichas.get(id_maquina) or {}
        nuevo = {}
        for campo in CAMPOS_FICHA_ESCRITA:
            valor = escrita.get(campo)
            if valor in (None, ""):
                continue
            tiene = actual.get(campo)
            if tiene not in (None, ""):
                if not _mismo_numero(tiene, valor):
                    descartes.append({
                        "donde": escrita["hoja"],
                        "motivo": f"La hoja dice {campo} {valor} y la ficha de "
                                  f"la MQ {escrita['numero']} dice {tiene}. "
                                  "Queda lo que ya estaba cargado."})
                continue
            if escrita["copiada"]:
                continue
            nuevo[campo] = valor
        if nuevo:
            cambios.append({"id_maquina": id_maquina,
                            "numero": escrita["numero"],
                            "hoja": escrita["hoja"], **nuevo})
    return cambios, descartes


# --------------------------------------------------------------------------
# Las hojas chicas
# --------------------------------------------------------------------------
def _buscar_titulos(filas, etiquetas, hasta=8, desde=0) -> tuple[int | None, dict[str, int]]:
    """La fila que tiene esos títulos, y en qué columna cayó cada uno.

    `desde` es la primera columna donde mirar. Existe porque la hoja de bandas
    tiene dos tablas al lado de la otra y sus títulos se parecen: sin este
    freno, «BANDAS DE MEMMIGER» de la izquierda se lleva la columna «BANDA» de
    la tabla de la derecha.
    """
    for i, fila in enumerate(filas[:hasta]):
        limpias = [_apretado(c) if j >= desde else "" for j, c in enumerate(fila)]
        mapa = {}
        for campo, pistas in etiquetas.items():
            for j, t in enumerate(limpias):
                if t and any(t == p or t.startswith(p) for p in pistas):
                    mapa[campo] = j
                    break
        if len(mapa) >= max(2, len(etiquetas) - 2):
            return i, mapa
    return None, {}


def leer_agujas(wb, maquinas) -> tuple[list[dict], list[dict]]:
    """Qué aguja lleva cada máquina.

    Las cuatro columnas de cilindro son las cuatro pistas de la misma máquina:
    van juntas, igual que las poleas.
    """
    hoja = next((n for n in wb.sheetnames if _apretado(n) == "agujas"), None)
    if not hoja:
        return [], []
    por_numero = {m["numero"]: m for m in maquinas if m.get("numero") is not None}
    filas = _filas_de(wb[hoja])
    corte, mapa = _buscar_titulos(filas, {
        "numero": ("mq", "maquina", "maq"),
        "cilindro": ("cilindro",),
        "plato": ("plato",),
        "descripcion": ("maquina",),
        "platinas": ("platinas",),
    })
    if corte is None:
        return [], [{"donde": hoja, "motivo": "No se encontraron los títulos"}]

    # Las columnas de cilindro son varias seguidas: desde la que dice
    # «CILINDRO» hasta la siguiente que tiene título.
    inicio = mapa.get("cilindro")
    fin = min((v for v in mapa.values() if inicio is not None and v > inicio),
              default=(inicio + 4) if inicio is not None else 0)

    salida, descartes, vistas = [], [], set()
    for fila in filas[corte + 1:]:
        if _vacia(fila):
            continue
        numero = _numero(fila[mapa["numero"]] if "numero" in mapa else None)
        maquina = por_numero.get(numero)
        if not maquina:
            descartes.append({"donde": f"{hoja}, MQ {numero}",
                              "motivo": f"La máquina {numero} no está en Asinfo"})
            continue
        cilindro = _juntar(_celdas(fila, range(inicio, fin)))
        if cilindro and "vacio" in _apretado(cilindro):
            cilindro = None
        item = {
            "id_maquina": maquina["id"],
            "descripcion": _texto(_una(fila, {"d": [mapa["descripcion"]]}, "d")
                                  if "descripcion" in mapa else None),
            "cilindro": cilindro,
            "plato": _texto(fila[mapa["plato"]]) if "plato" in mapa else None,
            "platinas": _juntar(_celdas(fila, range(mapa["platinas"], len(fila))))
                        if "platinas" in mapa else None,
            "nota": None,
        }
        if not any((item["cilindro"], item["plato"], item["platinas"])):
            # La MQ 14 dice «VACIO EL LUGAR»: la máquina no tiene agujas
            # puestas. No es un error de lectura, pero tiene que verse.
            descartes.append({
                "donde": f"{hoja}, MQ {numero}",
                "motivo": "La planilla no le pone agujas a esta máquina"})
            continue
        if maquina["id"] in vistas:
            # La MQ 12 aparece dos veces: tiene un juego para galga 24 y otro
            # para galga 28. Los dos son de esa máquina, así que se guardan los
            # dos juntos; quedarse con uno sería perder la mitad de la ficha.
            previo = next(x for x in salida if x["id_maquina"] == maquina["id"])
            for campo in ("cilindro", "plato", "platinas"):
                partes = [p for p in (previo[campo], item[campo]) if p]
                previo[campo] = " · ".join(partes) or None
            previo["nota"] = "La planilla le pone dos juegos, uno por galga"
            descartes.append({
                "donde": f"{hoja}, MQ {numero}",
                "motivo": "La máquina aparece dos veces: los dos juegos se guardan juntos"})
            continue
        vistas.add(maquina["id"])
        salida.append(item)
    return salida, descartes


def leer_agujas_modelo(wb) -> tuple[list[dict], list[dict]]:
    """La hoja «CODIGO DE AGUJAS»: la aguja de cada MODELO, con la marca.

    Se parece a «AGUJAS», pero trae un dato que la otra no tiene: de qué marca
    es la aguja y de qué marca la platina. Va a su propia tabla en vez de
    mezclarse, porque acá la fila es un modelo («MAYER)1-2-3-4-5-7-9-10») y no
    una máquina, y repartir eso entre máquinas sería adivinar.
    """
    hoja = next((n for n in wb.sheetnames if _apretado(n) == "codigo de agujas"), None)
    if not hoja:
        return [], []
    filas = _filas_de(wb[hoja])
    corte, mapa = _buscar_titulos(filas, {
        "modelo": ("maquina",), "marca": ("marca",), "codigo": ("codigo",),
        "donde": ("cilindro",), "platinas": ("platinas",),
    })
    if corte is None:
        return [], [{"donde": hoja, "motivo": "No se encontraron los títulos"}]

    # Los códigos son varias columnas seguidas, todas tituladas «CODIGO».
    ini = mapa.get("codigo")
    fin = mapa.get("donde", (ini + 4) if ini is not None else 0)

    salida, descartes, vistos = [], [], set()
    for fila in filas[corte + 1:]:
        if _vacia(fila):
            continue
        modelo = _texto(fila[mapa["modelo"]]) if "modelo" in mapa else None
        if not modelo or _apretado(modelo) == "vacia":
            continue
        if _apretado(modelo) in vistos:
            descartes.append({"donde": f"{hoja} · {modelo}",
                              "motivo": "El modelo aparece dos veces; se guarda el primero"})
            continue
        vistos.add(_apretado(modelo))
        col_pl = mapa.get("platinas")
        salida.append({
            "modelo": modelo,
            "marca_aguja": _texto(fila[mapa["marca"]]) if "marca" in mapa else None,
            "codigos": _juntar(_celdas(fila, range(ini, fin))) if ini is not None else None,
            "donde": _juntar(_celdas(fila, range(fin, col_pl))) if col_pl else None,
            "platinas": _texto(fila[col_pl]) if col_pl is not None else None,
            "marca_platina": _texto(fila[col_pl + 1]) if col_pl is not None
                             and col_pl + 1 < len(fila) else None,
            "nota": _juntar(_celdas(fila, range(col_pl + 2, len(fila))))
                    if col_pl is not None else None,
        })
    return salida, descartes


def _inicio_levas_tela(filas) -> int | None:
    """En qué columna empieza la tabla de la derecha de «INVENTARIO LEVAS»."""
    for fila in filas[:6]:
        for j, celda in enumerate(fila):
            if isinstance(celda, str) and "levas por tela" in _apretado(celda):
                return j
    return None


def leer_levas(wb) -> tuple[list[dict], list[dict]]:
    hoja = next((n for n in wb.sheetnames if _apretado(n) == "inventario levas"), None)
    if not hoja:
        return [], []
    filas = _filas_de(wb[hoja])
    corte, mapa = _buscar_titulos(filas, {
        "maquinas": ("maquina",), "codigo": ("codigo",), "cantidad": ("cantidad",),
        "ubicacion": ("ubicacion",), "accionamiento": ("accionamiento",),
    })
    if corte is None:
        return [], [{"donde": hoja, "motivo": "No se encontraron los títulos"}]

    # Entre la última columna con título y la tabla de la derecha quedan dos
    # números sueltos (288 y 490) en una columna SIN título. No son una
    # fórmula ni tienen nota: nada en la hoja dice qué cuentan, y no coinciden
    # con ninguna cuenta de las otras tablas. No se inventa un nombre para
    # guardarlos: salen en descartes para que el mecánico diga qué son, y
    # recién ahí se les hace columna.
    ultima = max(mapa.values(), default=-1)
    tope = _inicio_levas_tela(filas)
    if tope is None:
        tope = max((len(f) for f in filas), default=ultima + 1)

    salida, descartes, vistas = [], [], set()
    for n, fila in enumerate(filas[corte + 1:], start=corte + 2):
        if _vacia(fila):
            continue
        suelto = _juntar(_celdas(fila, range(ultima + 1, tope)))
        if suelto:
            descartes.append({
                "donde": f"{hoja}, fila {n}",
                "motivo": f"Hay un número sin título al lado de la cantidad "
                          f"({suelto}): la hoja no dice qué cuenta, así que no "
                          "se guarda"})
        maquinas = _texto(fila[mapa["maquinas"]]) if "maquinas" in mapa else None
        codigo = _texto(fila[mapa["codigo"]]) if "codigo" in mapa else None
        if not maquinas or not codigo:
            continue
        if _apretado(codigo) in ("no", "nO".lower()):
            descartes.append({"donde": f"{hoja} · {maquinas}",
                              "motivo": "La leva no tiene código («NO»)"})
            continue
        accionamiento = _texto(fila[mapa["accionamiento"]]) if "accionamiento" in mapa else None
        clave = (maquinas, codigo, accionamiento or "")
        if clave in vistas:
            continue
        vistas.add(clave)
        salida.append({
            "maquinas": maquinas,
            "codigo": codigo,
            "cantidad": _numero(fila[mapa["cantidad"]]) if "cantidad" in mapa else None,
            "ubicacion": _texto(fila[mapa["ubicacion"]]) if "ubicacion" in mapa else None,
            "accionamiento": accionamiento,
        })
    return salida, descartes


def leer_levas_tela(wb) -> tuple[list[dict], list[dict]]:
    """Cuántas levas lleva cada TELA. La segunda tabla de «INVENTARIO LEVAS».

    Está pegada a la derecha del inventario, con su propio título («CANTIDAD DE
    LEVAS POR TELA») y sin fila de encabezados: cada celda dice qué es adentro
    del texto —«levas de trabajo 432»—, así que se lee por posición desde donde
    arranca el título. Eran cuarenta renglones que no entraban a ningún lado.

    El número se deja pegado a su texto: varias filas dicen «192 cilindro» o
    «16 cilindro», y separar el número perdería dónde van.
    """
    hoja = next((n for n in wb.sheetnames if _apretado(n) == "inventario levas"), None)
    if not hoja:
        return [], []
    filas = _filas_de(wb[hoja])
    inicio = _inicio_levas_tela(filas)
    if inicio is None:
        return [], []

    def _sin_etiqueta(valor):
        """«levas de trabajo 432» → «432». Lo que no es la etiqueta queda."""
        texto = _texto(valor)
        if not texto:
            return None
        limpio = re.sub(r"^\s*levas?\s+de\s+\w+\s*", "", texto, flags=re.IGNORECASE)
        return limpio.strip() or texto

    salida, vistas = [], set()
    for fila in filas:
        def col(desplazo):
            i = inicio + desplazo
            return fila[i] if i < len(fila) else None

        marca = _texto(col(0))
        if not marca or "levas por tela" in _apretado(marca):
            continue
        tela = _texto(col(3))
        # Una fila con la marca y nada más es un renglón empezado.
        if not any(_texto(col(d)) for d in (1, 2, 3, 4)):
            continue
        item = {
            "marca": marca,
            "diametro": _texto(col(1)),
            "alimentadores": _texto(col(2)),
            "tela": tela,
            "trabajo": _sin_etiqueta(col(4)),
            "retenido": _sin_etiqueta(col(5)),
            "anulacion": _sin_etiqueta(col(6)),
            # Lo que sigue a la derecha son aclaraciones sueltas: la fila de la
            # JIUNN LONG de 38 trae once celdas con el detalle de cada leva.
            "nota": _juntar(_celdas(fila, range(inicio + 7, len(fila)))),
        }
        clave = (item["marca"], item["diametro"], item["alimentadores"],
                 item["tela"], item["trabajo"])
        if clave in vistas:
            continue
        vistas.add(clave)
        salida.append(item)
    return salida, []


# «4 de 6.600» = cuatro bandas de esa medida. Sólo se lee el número de
# adelante: el de atrás es la medida, ya está en su propia columna, y encima
# está escrita de dos formas distintas («6.600» y «6,600») en la misma tabla.
_CUANTAS_BANDAS = re.compile(r"^\s*(\d+)\s*(?:de\b|$)", re.I)


def _cuantas_bandas(texto) -> int | None:
    t = _texto(texto)
    hallado = _CUANTAS_BANDAS.match(t) if t else None
    return int(hallado.group(1)) if hallado else None


def leer_bandas(wb) -> tuple[list[dict], list[dict], list[dict]]:
    """Las bandas Memminger y de motor, y el stock por medida.

    La hoja tiene cuatro tablas: las Memminger (izquierda), el stock por medida
    (al medio), las de motor (derecha) y, más abajo, el cuadro de lo que hay
    que pedir — ése lo lee `leer_bandas_pedido`.

    De la tabla de la izquierda se leen también las dos columnas que antes se
    tiraban: cuántas máquinas usan esa medida y cuántas bandas hacen falta. Son
    el único lugar donde está escrito para qué alcanza el stock.
    """
    hoja = next((n for n in wb.sheetnames if _apretado(n) == "bandas"), None)
    if not hoja:
        return [], [], []
    filas = _filas_de(wb[hoja])
    corte, mapa = _buscar_titulos(filas, {
        "maquinas": ("maquna", "maquina"), "cantidad_maquinas": ("cantidad de maquinas",),
        "diametro": ("diametro",), "media": ("banda 1 2",),
        "tres_cuartos": ("banda 3 4",), "lycra": ("banda lycra",),
        "medida": ("codigo",),
        "con_medida": ("cantidada de maquinas con", "cantidad de maquinas con"),
        "requerida": ("cantidad requerida",),
    })
    if corte is None:
        return [], [], [{"donde": hoja, "motivo": "No se encontraron los títulos"}]

    # El stock va en el par de columnas «CODIGO | CANTIDAD», pegado a la
    # derecha de la tabla de bandas. La cantidad se toma de la columna de al
    # lado del código, y no buscando el título «CANTIDAD»: ese título también
    # lo tiene «cantidad de maquinas», que está más a la izquierda y ganaría
    # siempre. El stock terminaría siendo la cantidad de máquinas.
    col_medida = mapa.get("medida")
    col_stock = col_medida + 1 if col_medida is not None else None

    # «CANTIDAD REQUERIDA» es un título combinado sobre DOS columnas: la de la
    # banda 1/2 y la de la 3/4. Excel deja el título sólo en la primera, así
    # que la segunda se toma por posición: no tiene título propio.
    col_req = mapa.get("requerida")
    col_req2 = col_req + 1 if col_req is not None else None

    def _celda(fila, col):
        return fila[col] if col is not None and col < len(fila) else None

    bandas, stock, descartes, vistas = [], [], [], set()
    for fila in filas[corte + 1:]:
        if _vacia(fila):
            continue
        maquinas = _texto(fila[mapa["maquinas"]]) if "maquinas" in mapa else None
        diametro = _decimal(fila[mapa["diametro"]]) if "diametro" in mapa else None
        medidas = {c: (_texto(fila[mapa[c]]) if c in mapa else None)
                   for c in ("media", "tres_cuartos", "lycra")}
        requerida = _texto(_celda(fila, col_req))
        requerida2 = _texto(_celda(fila, col_req2))
        # Un nombre de modelo suelto, sin ninguna medida y sin cuántas bandas
        # lleva, es una fila que quedó empezada. La máquina de 42 pulgadas
        # tiene las medidas en blanco pero sí dice cuántas necesita: antes se
        # caía por eso.
        entra = maquinas and (any(medidas.values()) or requerida or requerida2)
        if entra and (maquinas, diametro) not in vistas:
            vistas.add((maquinas, diametro))
            bandas.append({
                "clase": "memminger",
                "maquinas": maquinas,
                "cantidad_maquinas": _numero(fila[mapa["cantidad_maquinas"]])
                                     if "cantidad_maquinas" in mapa else None,
                "maquinas_con_medida": _texto(_celda(fila, mapa.get("con_medida"))),
                "diametro": diametro,
                "requerida_media": _cuantas_bandas(requerida),
                "requerida_media_texto": requerida,
                "requerida_tres_cuartos": _cuantas_bandas(requerida2),
                "requerida_tres_cuartos_texto": requerida2,
                "banda": None, "cobrador": None, "nota": None,
                **medidas,
            })
        # La fila «total» es una suma del Excel, no una medida de banda: como
        # no es un número, queda afuera sola.
        if col_stock is not None and col_stock < len(fila):
            medida = _decimal(fila[col_medida])
            if medida is not None:
                stock.append({"medida": medida, "cantidad": _numero(fila[col_stock])})

    # Las bandas de MOTOR, en su propia tabla a la derecha de la hoja. Otro
    # repuesto y otras columnas, así que van con clase propia: si se mezclaran
    # con las Memminger, una medida taparía a la otra.
    corte2, mapa2 = _buscar_titulos(filas, {
        "maquinas": ("maquina",), "cantidad_maquinas": ("cantidad",),
        "diametro": ("diametro",), "banda": ("banda",), "cobrador": ("cobrador",),
    }, hasta=4, desde=(col_stock + 1) if col_stock is not None else 0)
    if corte2 is not None and mapa2.get("cobrador"):
        vistas2 = set()
        for fila in filas[corte2 + 1:]:
            if _vacia(fila):
                continue
            maquinas = _texto(fila[mapa2["maquinas"]]) if "maquinas" in mapa2 else None
            banda = _texto(fila[mapa2["banda"]]) if "banda" in mapa2 else None
            if not maquinas or not banda:
                continue
            diametro = _decimal(fila[mapa2["diametro"]]) if "diametro" in mapa2 else None
            if (maquinas, diametro) in vistas2:
                continue
            vistas2.add((maquinas, diametro))
            col_cob = mapa2["cobrador"]
            bandas.append({
                "clase": "motor",
                "maquinas": maquinas,
                "cantidad_maquinas": _numero(fila[mapa2["cantidad_maquinas"]])
                                     if "cantidad_maquinas" in mapa2 else None,
                "maquinas_con_medida": None,
                "diametro": diametro,
                "requerida_media": None, "requerida_media_texto": None,
                "requerida_tres_cuartos": None, "requerida_tres_cuartos_texto": None,
                "media": None, "tres_cuartos": None, "lycra": None,
                "banda": banda,
                "cobrador": _texto(fila[col_cob]),
                "nota": _juntar(_celdas(fila, range(col_cob + 1, len(fila)))),
            })
    else:
        descartes.append({"donde": hoja,
                          "motivo": "No se encontró la tabla de bandas de motor"})
    return bandas, stock, descartes


def _titulos_de_una_fila(fila, etiquetas) -> dict[str, int]:
    """{campo: columna} mirando UNA fila sola."""
    mapa = {}
    limpias = [_apretado(c) for c in fila]
    for campo, pistas in etiquetas.items():
        for j, t in enumerate(limpias):
            if t and any(t == p or t.startswith(p) for p in pistas):
                mapa[campo] = j
                break
    return mapa


def leer_bandas_pedido(wb) -> tuple[list[dict], list[dict]]:
    """Cuántas bandas hacen falta, cuántas hay y cuántas hay que pedir.

    Es el cuadro del fondo de la hoja BANDAS, debajo de las otras tres tablas,
    y hasta ahora no lo leía nadie: es el único lugar de toda la planilla donde
    está escrito qué hay que comprar. Una medida por fila.

    Tres cosas de la hoja obligan a leerlo así:

      * **Los títulos están en la fila 26, no arriba.** `_buscar_titulos` mira
        sólo las primeras filas, que son de las otras tablas. Acá la fila se
        busca por «pedir», el único título que no se repite en la hoja:
        «codigo» y «stok» también los tiene la tabla de stock de más arriba.
      * **La última fila trae el total pegado en la columna de la cantidad.**
        En el Excel es un SUM que quedó encima de la fila de la medida 13 y
        dice 132 donde tendría que decir cuántas bandas de 13 hacen falta. Se
        reconoce porque es exactamente la suma de las de arriba: se descarta
        esa celda —no la fila, que trae stock de verdad— y se avisa.
      * **La fila de totales del pie tiene una celda de plata sin título.** No
        se guarda: un precio sin nombre no es un dato.

    Ojo con los números: este cuadro NO cuadra con la tabla de stock de arriba
    ni con la de la izquierda. Son tres cuentas de lo mismo, escritas a mano en
    momentos distintos. Se cargan las tres como están; cuál manda lo dice el
    mecánico, no el programa.
    """
    hoja = next((n for n in wb.sheetnames if _apretado(n) == "bandas"), None)
    if not hoja:
        return [], []
    filas = _filas_de(wb[hoja])

    corte, mapa = None, {}
    for i, fila in enumerate(filas):
        posible = _titulos_de_una_fila(fila, {
            "requeridas": ("cantidad",), "medida": ("codigo",),
            "stock": ("stok", "stock"), "pedir": ("pedir",),
            "metros": ("metros",)})
        if "pedir" in posible and "medida" in posible:
            corte, mapa = i, posible
            break
    if corte is None:
        return [], [{"donde": hoja,
                     "motivo": "No se encontró el cuadro de bandas para pedir"}]

    salida, descartes, vistas, suma = [], [], set(), 0
    for n, fila in enumerate(filas[corte + 1:], start=corte + 2):
        if _vacia(fila):
            continue
        col_medida = mapa["medida"]
        medida = _decimal(fila[col_medida]) if col_medida < len(fila) else None
        # Sin medida no hay fila: así quedan afuera solas la fila de totales
        # del pie y cualquier renglón empezado.
        if medida is None:
            continue
        if medida in vistas:
            descartes.append({"donde": f"{hoja}, fila {n}",
                              "motivo": f"La medida {medida} está dos veces en "
                                        "el cuadro de pedido"})
            continue
        vistas.add(medida)

        def valor(campo, como=_numero):
            col = mapa.get(campo)
            return como(fila[col]) if col is not None and col < len(fila) else None

        requeridas = valor("requeridas")
        if requeridas is not None and suma and requeridas == suma:
            descartes.append({
                "donde": f"{hoja}, fila {n}",
                "motivo": f"En la medida {medida} la columna «cantidad» trae el "
                          f"total del cuadro ({requeridas}), no cuántas bandas "
                          "hacen falta. Queda vacía."})
            requeridas = None
        elif requeridas is not None:
            suma += requeridas

        salida.append({"medida": medida, "requeridas": requeridas,
                       "stock": valor("stock"), "pedir": valor("pedir"),
                       "metros": valor("metros", _decimal)})

    return salida, descartes


def leer_eficiencia(wb, maquinas) -> tuple[list[dict], list[dict]]:
    """Cuánto debería dar cada máquina en 12 horas.

    La hoja tiene dos bloques, uno abajo del otro, con los mismos títulos: las
    jersey arriba y las doble fontura abajo. Se leen los dos.

    Es un CÁLCULO de la planilla, no lo que la máquina tejió. Los kilos reales
    los mide Asinfo y no tienen por qué coincidir.
    """
    hoja = next((n for n in wb.sheetnames
                 if _apretado(n).startswith("eficiencia produccion")), None)
    if not hoja:
        return [], []
    por_numero = {m["numero"]: m for m in maquinas if m.get("numero") is not None}
    filas = _filas_de(wb[hoja])

    salida, descartes, vistas = [], [], set()
    mapa: dict[str, int] = {}
    for fila in filas:
        cabecera, repetidas = {}, {"aproximacion": [], "peso kg": [],
                                   "produccion real": []}
        for j, celda in enumerate(fila):
            t = _apretado(celda)
            if not t:
                continue
            for etiqueta in repetidas:
                if t.startswith(etiqueta):
                    repetidas[etiqueta].append(j)
            for campo, pistas in (("numero", ("maquina",)), ("rpm", ("velocidad",)),
                                  ("sistemas", ("sistema",)), ("diametro", ("diametro",)),
                                  # La columna «F» es la GALGA, no los
                                  # alimentadores: los alimentadores son los
                                  # «sistemas» —102, 96, 62— y la F dice 24, 28.
                                  ("galga", ("f",)), ("tamano_rollo", ("tamano de rollo",)),
                                  ("minutos_rollo", ("tiempo",)), ("rollos_dia", ("aproximacion",)),
                                  ("kg_dia", ("peso kg",))):
                if campo not in cabecera and any(t == p or t.startswith(p) for p in pistas):
                    cabecera[campo] = j
                    break
        if len(cabecera) >= 6:
            # La hoja calcula dos turnos con los MISMOS títulos repetidos: el
            # de 12 horas primero y el de 24 después. No hay forma de
            # distinguirlos por el nombre, así que se toman por orden de
            # aparición — y abajo se chequea que el de 24 sea mayor que el de
            # 12 antes de guardarlo. Si no lo es, se deja vacío.
            if len(repetidas["aproximacion"]) >= 3:
                cabecera["rollos_dia_24"] = repetidas["aproximacion"][2]
            if len(repetidas["peso kg"]) >= 4:
                cabecera["kg_dia_24"] = repetidas["peso kg"][3]
            # Lo que la máquina DIO de verdad, medido en planta. Es la columna
            # más importante de la hoja y no se estaba guardando: al lado del
            # cálculo dice si la máquina está rindiendo o no. Los kilos van
            # en la columna de al lado de los rollos.
            for n, campo in enumerate(("real_rollos_dia", "real_rollos_24")):
                if len(repetidas["produccion real"]) > n:
                    columna = repetidas["produccion real"][n]
                    cabecera[campo] = columna
                    cabecera[campo.replace("rollos", "kg")] = columna + 1
            mapa = cabecera
            continue
        if not mapa:
            continue
        # Con `.get`, no con `[...]`: si a esa hoja le cambian el título de la
        # columna de la máquina, esto tiene que dejar la hoja de lado, no tirar
        # abajo la carga entera de la planilla.
        columna = mapa.get("numero")
        if columna is None:
            continue
        numero = _numero(fila[columna] if columna < len(fila) else None)
        if numero is None or numero in vistas:
            continue
        maquina = por_numero.get(numero)
        if not maquina:
            descartes.append({"donde": f"{hoja}, MQ {numero}",
                              "motivo": f"La máquina {numero} no está en Asinfo"})
            continue
        vistas.add(numero)

        def val(campo, como=_decimal):
            i = mapa.get(campo)
            return como(fila[i]) if i is not None and i < len(fila) else None

        kg_dia = val("kg_dia")
        kg_24 = val("kg_dia_24")
        rollos_24 = val("rollos_dia_24")
        # El turno de 24 horas tiene que dar más que el de 12. Si no da, los
        # títulos repetidos no cayeron donde creíamos: mejor vacío que un
        # número puesto en la columna equivocada.
        if not (kg_dia and kg_24 and kg_24 > kg_dia):
            kg_24 = rollos_24 = None

        salida.append({
            "id_maquina": maquina["id"],
            "rpm": val("rpm"),
            "sistemas": val("sistemas", _texto),
            "diametro": val("diametro"),
            "galga": val("galga", _numero),
            "tamano_rollo": val("tamano_rollo"),
            "minutos_rollo": val("minutos_rollo"),
            "rollos_dia": val("rollos_dia"),
            "kg_dia": kg_dia,
            "rollos_dia_24": rollos_24,
            "kg_dia_24": kg_24,
            "real_rollos_dia": val("real_rollos_dia"),
            "real_kg_dia": val("real_kg_dia"),
            "real_rollos_24": val("real_rollos_24"),
            "real_kg_24": val("real_kg_24"),
        })
    return salida, descartes


# El porcentaje escrito al final del nombre del hilo: «HILO 22/1  71 %».
# Anclado al FINAL a propósito: un «65/35» del medio es la mezcla del hilo, no
# cuánto lleva la tela.
_PORCENTAJE_HILO = re.compile(r"(\d+(?:[.,]\d+)?)\s*%\s*$")


def _en_blanco(fila) -> bool:
    """Una fila con celdas de sólo espacios está vacía.

    En «consumo de hilo» hay una fila con tres espacios en una columna del
    fondo. Para el ojo separa dos telas; si no cuenta como vacía, los dos
    bloques quedan pegados y la tela de abajo hereda la de arriba.
    """
    return all(c in (None, "") or (isinstance(c, str) and not c.strip())
               for c in fila)


def _nombre_de_tela(fila) -> str | None:
    t = _texto(fila[0] if fila else None)
    return t if t and _apretado(t) != "tela" else None


def _bloques_de_hilo(filas, desde: int):
    """Las filas agrupadas por bloque: lo que hay entre dos filas vacías."""
    bloque = []
    for i, fila in enumerate(filas[desde:], start=desde):
        if _en_blanco(fila):
            if bloque:
                yield bloque
            bloque = []
        else:
            bloque.append((i, fila))
    if bloque:
        yield bloque


def leer_consumo_hilo(wb) -> tuple[list[dict], list[dict]]:
    """Cuánto hilo lleva cada tela.

    La tela se escribe una sola vez y abajo van los hilos que la componen.

    Tres cosas que la hoja hace y antes se perdían enteras y en silencio:

      * **Las últimas ocho telas tienen el porcentaje adentro del nombre del
        hilo** («HILO 22/1  71 %») en vez de en la columna de rendimiento. Como
        el lector exigía esa columna, FALSO LYCRA, FALSO FLEECE 102, FALSO
        FLEECE 96, VITOR, FLEECEC 200, FRANELA, STEFI y BOXER NUEVO no entraban
        ni salían en los descartes: desaparecían. Ahora el número se lee donde
        está y va a `porcentaje`, NO a `rendimiento`: son la misma idea en dos
        unidades, y pasar 71 % a 0,71 sería corregir la planilla. El que mira
        la pantalla tiene que ver cuál de las dos le escribieron.
      * **Una tela tiene el nombre escrito en la segunda línea de su bloque**
        (FLEECEC 200, galga 22). Arrastrando sólo hacia abajo, su primer hilo
        —el 82 %— se caía. Cuando en un bloque hay una sola tela nombrada, esa
        tela vale para todo el bloque, esté escrita arriba o al medio.
      * **Cuatro renglones traen el rendimiento sin el nombre del hilo**: son
        las telas que tienen los hilos escritos a lo ancho (tres columnas) y
        los rendimientos a lo largo. No se puede saber qué número va con qué
        hilo sin adivinar, así que van a descartes y los muestra la pantalla.

    Devuelve (filas, descartes).
    """
    hoja = next((n for n in wb.sheetnames if _apretado(n) == "consumo de hilo"),
                None)
    if not hoja:
        return [], []
    filas = _filas_de(wb[hoja])

    # La fila de títulos se saltea a mano porque está corrida: dice
    # «rendimiento» arriba de una columna y el dato vive dos más a la
    # izquierda. Leída por título, la hoja entera saldría vacía.
    inicio = 0
    for i, fila in enumerate(filas[:_FILAS_CABECERA]):
        if _apretado(fila[0] if fila else None) == "tela":
            inicio = i + 1
            break

    salida, descartes, vistas = [], [], set()
    for bloque in _bloques_de_hilo(filas, inicio):
        nombradas = [n for _, f in bloque if (n := _nombre_de_tela(f))]
        unica = nombradas[0] if len(nombradas) == 1 else None
        tela = unica
        for n, fila in bloque:
            if not unica and (nombre := _nombre_de_tela(fila)):
                tela = nombre
            hilo = _juntar(_celdas(fila, (1, 2, 3)))
            rendimiento = _decimal(fila[4] if len(fila) > 4 else None)
            if not hilo:
                if rendimiento is not None:
                    descartes.append({
                        "donde": f"{hoja} · {tela or 'sin tela'}, fila {n + 1}",
                        "motivo": f"Hay un rendimiento ({rendimiento}) sin el "
                                  "nombre del hilo: la tela tiene los hilos "
                                  "escritos a lo ancho"})
                continue
            if not tela:
                descartes.append({"donde": f"{hoja}, fila {n + 1}",
                                  "motivo": f"«{hilo}» no dice de qué tela es"})
                continue
            codigo = _texto(fila[5] if len(fila) > 5 else None)
            porcentaje = None
            hallado = _PORCENTAJE_HILO.search(hilo)
            if hallado is not None:
                porcentaje = _decimal(hallado.group(1))
                hilo = hilo[:hallado.start()].strip() or hilo
            if rendimiento is None and porcentaje is None:
                descartes.append({"donde": f"{hoja} · {tela}",
                                  "motivo": f"«{hilo}» no dice cuánto lleva"})
                continue
            if (tela, hilo) in vistas:
                descartes.append({"donde": f"{hoja} · {tela}",
                                  "motivo": f"«{hilo}» aparece dos veces en la "
                                            "misma tela"})
                continue
            vistas.add((tela, hilo))
            salida.append({"tela": tela, "hilo": hilo, "codigo_hilo": codigo,
                           "rendimiento": rendimiento, "porcentaje": porcentaje})
    return salida, descartes


def leer_gramajes(wb, maquinas, hoy=None) -> tuple[list[dict], list[dict]]:
    """El peso medido de la tela de cada máquina (la hoja «Hoja3»)."""
    hoy = hoy or date.today()
    hoja = next((n for n in wb.sheetnames if _apretado(n) == "hoja3"), None)
    if not hoja:
        return [], []
    por_numero = {m["numero"]: m for m in maquinas if m.get("numero") is not None}
    filas = _filas_de(wb[hoja])
    corte, mapa = _buscar_titulos(filas, {
        "numero": ("mq", "maquina"), "fecha": ("fecha",), "tela": ("tela",),
        "peso": ("peso",),
    })
    if corte is None:
        return [], [{"donde": hoja, "motivo": "No se encontraron los títulos"}]

    salida, descartes = [], []
    for n, fila in enumerate(filas[corte + 1:], start=1):
        if _vacia(fila):
            continue
        numero = _numero(fila[mapa["numero"]] if "numero" in mapa else None)
        maquina = por_numero.get(numero)
        if not maquina:
            descartes.append({"donde": f"{hoja}, MQ {numero}",
                              "motivo": f"La máquina {numero} no está en Asinfo"})
            continue
        peso = _decimal(fila[mapa["peso"]]) if "peso" in mapa else None
        if peso is None:
            continue
        salida.append({
            "id_maquina": maquina["id"],
            "fecha": next((f for f in (_fecha(c, hoy) for c in fila) if f), None),
            "tela": _texto(fila[mapa["tela"]]) if "tela" in mapa else None,
            "hilos": _juntar(_celdas(fila, (3, 4, 5))),
            "peso": peso,
            "orden": n,
        })
    return salida, descartes


# --------------------------------------------------------------------------
# Todo junto
# --------------------------------------------------------------------------
# Cómo se llama cada bloque en pantalla. En castellano y como lo dirían en
# planta: nadie dice «registros de ajuste».
NOMBRES = {
    "ajustes": "Ajustes de máquina",
    "agujas": "Agujas por máquina",
    "agujas_modelo": "Agujas por modelo",
    "levas": "Levas",
    "levas_tela": "Cuántas levas lleva cada tela",
    "bandas": "Bandas",
    "banda_stock": "Bandas en stock",
    "banda_pedido": "Bandas que hay que pedir",
    "eficiencia": "Producción por máquina",
    "consumo_hilo": "Consumo de hilo",
    "gramajes": "Gramajes medidos",
}


def leer(ruta: str, maquinas: list[dict], hoy=None) -> tuple[dict, list[dict]]:
    """Lee la planilla entera. Devuelve (bloques, descartes).

    `bloques` es {nombre: [filas]} listo para guardar; `descartes` dice qué no
    entró y por qué, para mostrarlo antes de confirmar.
    """
    wb = load_workbook(ruta, data_only=True)
    try:
        ajustes_, d1 = leer_ajustes(wb, maquinas, hoy)
        viejos_, d8 = leer_agustes(wb, maquinas, hoy)
        agujas_, d2 = leer_agujas(wb, maquinas)
        modelos_, d9 = leer_agujas_modelo(wb)
        levas_, d3 = leer_levas(wb)
        levas_tela_, d10 = leer_levas_tela(wb)
        bandas_, stock_, d4 = leer_bandas(wb)
        pedido_, d11 = leer_bandas_pedido(wb)
        eficiencia_, d5 = leer_eficiencia(wb, maquinas)
        consumo_, d6 = leer_consumo_hilo(wb)
        gramajes_, d7 = leer_gramajes(wb, maquinas, hoy)
    finally:
        wb.close()

    # De la hoja vieja entra sólo lo que no está en las hojas por máquina. Lo
    # repetido se cuenta y se avisa, para que el número cierre.
    conocidas = {_firma(a) for a in ajustes_}
    nuevos = [a for a in viejos_ if _firma(a) not in conocidas]
    repetidos = len(viejos_) - len(nuevos)
    if repetidos:
        d8.append({
            "donde": "AGUSTES",
            "motivo": f"{repetidos} ajustes de la hoja vieja ya estaban en las "
                      "hojas de cada máquina"})

    bloques = {
        "ajustes": ajustes_ + nuevos,
        "agujas": agujas_,
        "agujas_modelo": modelos_,
        "levas": levas_,
        "levas_tela": levas_tela_,
        "bandas": bandas_,
        "banda_stock": stock_,
        "banda_pedido": pedido_,
        "eficiencia": eficiencia_,
        "consumo_hilo": consumo_,
        "gramajes": gramajes_,
    }
    return bloques, d1 + d8 + d2 + d9 + d3 + d10 + d4 + d11 + d5 + d6 + d7
